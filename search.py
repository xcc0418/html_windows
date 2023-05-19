import pymysql
import json
import global_var
import datetime
import time
import hashlib
import urllib
import requests
import base64
from Crypto.Cipher import AES
import openpyxl
from openpyxl.styles import PatternFill, Font


def find_storage(fnsku):
    get_url = "https://erp.lingxing.com/api/storage/lists"
    get_headers = {'Host': 'erp.lingxing.com',
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                   , 'Referer': 'https://erp.lingxing.com/erp/msupply/warehouseDetail',
                   'Accept': 'application/json, text/plain, */*',
                   'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                   'Accept-Encoding': 'gzip, deflate, br',
                   'Content-Type': 'application/json;charset=utf-8',
                   'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                   'X-AK-Company-Id': '90136229927150080',
                   'X-AK-Request-Source': 'erp',
                   'X-AK-ENV-KEY': 'SAAS-10',
                   'X-AK-Version': '1.0.0.0.0.023',
                   'X-AK-Zid': '109810',
                   'Content-Length': '909',
                   'Origin': 'https://erp.lingxing.com',
                   'Connection': 'keep-alive'}
    data = {"wid_list": "", "mid_list": "", "sid_list": "", "cid_list": "", "bid_list": "", "principal_list": "",
            "product_type_list": "", "product_attribute": "", "product_status": "", "search_field": "fnsku",
            "search_value": f"{fnsku}", "is_sku_merge_show": 0, "is_hide_zero_stock": 0, "offset": 0,
            "length": 200, "sort_field": "pre_total", "sort_type": "desc", "gtag_ids": "",
            "senior_search_list": "[]", "req_time_sequence": "/api/storage/lists$$"}
    data = json.dumps(data)
    get_msg = global_var.s.post(get_url, headers=get_headers, data=data)
    get_msg = json.loads(get_msg.text)
    num_prepare = 0
    # if sku == '[GCWL]635':
    # print(get_msg)
    # for i in get_msg['list']:
    if get_msg['code'] == 1 and get_msg['msg'] == '操作成功':
        num_prepare = get_msg['data']['total_info']['total']
    # print(num_prepare)
    return num_prepare


def get_warehouse(filename):
    ws = openpyxl.load_workbook(filename)
    ws_sheet = ws.active
    row1 = ws_sheet.max_row
    for i in range(row1, 0, -1):
        cell_value1 = ws_sheet.cell(row=i, column=1).value
        if cell_value1:
            row1 = i
            break
    list_fnsku = []
    for i in range(1, row1+1):
        fnsku = ws_sheet.cell(row=i, column=2).value
        if fnsku:
            fnsku = fnsku.strip('')
            if fnsku in list_fnsku:
                continue
            else:
                list_fnsku.append(fnsku)
    list_warehouse = []
    for i in list_fnsku:
        msg = get_po(i)
        list_warehouse.append(msg)
    if list_warehouse:
        wb = openpyxl.Workbook()
        wb_sheet = wb.active
        wb_sheet.append(['品名', 'FNSKU', '箱号信息', '装箱总数量'])
        row1 = 1
        for i in list_warehouse:
            for j in i:
                wb_sheet.append(j)
                row1 += 1
        starrow1 = 2
        last_cell_value = wb_sheet.cell(row=starrow1, column=2).value
        count = 0
        for q in range(0, row1):
            cell_value = wb_sheet.cell(row=starrow1 + count, column=2).value
            if cell_value == last_cell_value:
                count = count + 1
            else:
                wb_sheet.merge_cells('B%d:B%d' % (starrow1, starrow1 - 1 + count))
                wb_sheet.merge_cells('A%d:A%d' % (starrow1, starrow1 - 1 + count))
                wb_sheet.merge_cells('D%d:D%d' % (starrow1, starrow1 - 1 + count))
                starrow1 = starrow1 + count
                count = 1
                last_cell_value = cell_value
        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        file = f'./static/FNSKU装箱信息/{time_now}-箱号.xlsx'
        wb.save(file)
        return f'../FNSKU装箱信息/{time_now}-箱号.xlsx', f'{time_now}-箱号'


def get_po(fnsku):
    list_msg = []
    num_local = 0
    result = func("SELECT `箱号`, `存放位置`, `数量1` FROM `storage`.`warehouse` WHERE `FNSKU1` = '%s' and `状态` = '已打包'" % fnsku)
    if result:
        name = func("SELECT `品名` FROM `data_read`.`listing` where `FNSKU`='%s'" % fnsku)
        num = 0
        for i in result:
            num += int(i[2])
        for i in result:
            num_local += i[2]
            if i[1]:
                msg = {'箱号': i[0], 'fnsku': fnsku, '品名': name[0][0], '装箱数量': i[2], '装箱状态': '已打包', '装箱位置': i[1]}
                list_msg.append(msg)
            else:
                msg = {'箱号': i[0], 'fnsku': fnsku, '品名': name[0][0], '装箱数量': i[2], '装箱状态': '已打包', '装箱位置': '这箱还没有匹配位置'}
                list_msg.append(msg)
    return list_msg, num_local


def func(sql, m='r'):
    py = pymysql.connect(host='3354n8l084.goho.co', user='test_user', password='a123456', port=24824, db='storage')
    cursor = py.cursor()
    data = None
    # print(sql)
    try:
        cursor.execute(sql)
        if m == 'r':
            data = cursor.fetchall()
        elif m == 'w':
            py.commit()
            data = cursor.rowcount
    except:
        data = False
        py.rollback()
    py.close()
    return data


class Quantity (object):
    def __init__(self):
        self.app_id = "ak_nEfE94OSogf3x"
        app_secret = "g2BcerjK4fWmhGZoetCJHeVqJmHEmfLt3gWFjDrBLB1yxiapgUKH6kOVs2N9JH7SFuBKuOF8K/CrNNSeFO1KKtsL05z24j" \
                     "+AdWTW+V4op5QxDkmlTllvlprT8FfjctDdNDGrwHvBvE6s9h0pO0dNgopBAYiA7oosPzQhDF1A6XC1X/cZZmgBy3XRHyEv" \
                     "xTT40xzwVGish53R8dZt3YIxNtKSgrBloo/CRQsV01yU40nyQR9L9oML32VT0C16jBrxcoWthlGDwfBn+CtVUvws4imyZi" \
                     "+sG/CqZQeaVLkBCqLCgqw1VK4/a4jZws6HO3FMeBgvuf0aS5euNmfhkudQmg=="
        querystring = {"appId": f"{self.app_id}", "appSecret": f"{app_secret}"}
        url = "https://openapi.lingxing.com/api/auth-server/oauth/access-token"
        payload = ""
        headers = {
            "Content-Type": "application/x-www-form-urlencoded"
        }
        response = requests.post(url, data=payload, headers=headers, params=querystring)
        result = json.loads(response.text)
        # print(response.text)
        self.access_token = result['data']['access_token']
        self.refresh_token = result['data']['refresh_token']
        self.time = datetime.datetime.now().strftime("%Y-%m-%d")
        self.time = datetime.datetime.strptime(self.time, "%Y-%m-%d")
        # self.time = "2021-12-06"
        self.start_time = self.time - datetime.timedelta(days=1)
        self.time = self.time.strftime("%Y-%m-%d")
        self.start_time = self.start_time.strftime("%Y-%m-%d")
        # self.start_time = '2022-03-07'
        # print(self.time, self.start_time)

    def sql(self):
        self.connection = pymysql.connect (host='3354n8l084.goho.co',  # 数据库地址
                                           port=24824,
                                           user='test_user',  # 数据库用户名
                                           password='a123456',  # 数据库密码
                                           db='storage',  # 数据库名称
                                           charset='utf8',
                                           cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor ()

    def sql_close(self):
        self.cursor.close ()
        self.connection.close ()

    def get_sign(self, body):
        apikey = "ak_nEfE94OSogf3x"
        # body = {"access_token": self.access_token,
        #         "timestamp": int(time.time()),
        #         "start_date": f"{self.start_time}",
        #         "end_date": f"{self.time}",
        #         "app_key": apikey}
        # print(body)
        # bb = sign().get_sign(apikey, body)
        # print(bb)

        # 目标md5串
        str_parm = ''
        # 将字典中的key排序
        for p in sorted (body):
            # 每次排完序的加到串中
            # if body[p]:
            # str类型需要转化为url编码格式
            if isinstance (body[p], str):
                str_parm = str_parm + str (p) + "=" + str (urllib.parse.quote (body[p])) + "&"
                continue
            # if isinstance(body[p], dict):
            #     for i in body[p]:
            #         body[p][i] = str(body[p][i])
            # if isinstance(body[p], list) and isinstance(body[p][0], dict):
            #     for i in range(0, len(body[p])):
            #         for j in body[p][i]:
            #             body[p][i][j] = str(urllib.parse.quote(str(body[p][i][j])))
            # if p == "product_list":
            #     str_parm = str_parm + str(p) + "=" + '[{"sku":"GCWL.897","good_num":1,"bad_num":0,"seller_id":0,"fnsku":""}]' + "&"
            #     continue
            #     body[p][0] = str(urllib.parse.quote(body[p][0]))
            #     print(str(urllib.parse.quote(body[p][0])))
            str_value = str (body[p])
            str_value = str_value.replace (" ", "")
            # str_value = str_value.replace("?", " ")
            str_value = str_value.replace ("'", "\"")
            str_parm = str_parm + str (p) + "=" + str_value + "&"
        # 加上对应的key
        str_parm = str_parm.rstrip ('&')
        # print("字符串拼接1:", str_parm)
        str_parm = str_parm.replace ("?", " ")
        # print("字符串拼接2:", str_parm)
        if isinstance (str_parm, str):
            # 如果是unicode先转utf-8
            parmStr = str_parm.encode ("utf-8")
            # parmStr = str_parm
            m = hashlib.md5 ()
            m.update (parmStr)
            md5_sign = m.hexdigest ()
            # print(m.hexdigest())
            md5_sign = md5_sign.upper ()
            # print("MD5加密:", md5_sign)
        eg = EncryptDate (apikey)  # 这里密钥的长度必须是16的倍数
        res = eg.encrypt (md5_sign)
        # print("AES加密:", res)
        # print(eg.decrypt(res))
        return res

    def get_warehouse_num(self, sku):
        # 生成sign签名
        # API接口路径
        url = "https://openapi.lingxing.com/erp/sc/routing/data/local_inventory/inventoryDetails"
        body = {"access_token": self.access_token,
                "timestamp": int(time.time()),
                "app_key": "ak_nEfE94OSogf3x", "offset": 0, "length": 400
               }
        # "wid": "2156,1489,2382,1490,1461,1488,1476,1477,1399,1478,414", "offset": 0, "length": 400
        res = self.get_sign(body)
        # res = ''
        querystring = {"access_token": self.access_token,
                       "timestamp": int(time.time()),
                       "app_key": self.app_id,
                       "sign": res
                       }
        payload = {"wid": "", "offset": 0, "length": 400}
        payload = json.dumps(payload)
        headers = {
            "Content-Type": "application/json",
            "Authorization": "bearer {{access_token}}"
        }
        response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
        result1 = json.loads(response.text)
        num_total = 0
        # print(result1)
        # 判断数据是否正常获取
        if result1['code'] == 0 and result1['message'] == "success":
            # print(len(result1['data']))
            # 获取库存明细数据条数
            length = result1['total']
            # length = 620
            # print(length)
            i = int(length/400)
            # print(i)
            # 以四百条数据为一页分页查询
            for k in range(0, i+1):
                # 分页查询
                payload = {"wid": '', "offset": k*400, "length": 400}
                if k == i:
                    body['length'] = length-k*400
                    payload['length'] = length-k*400
                body['offset'] = k*400
                # print(payload)
                res = self.get_sign(body)
                querystring['sign'] = res
                payload = json.dumps(payload)
                response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
                result = json.loads(response.text)
                # print(result)
                # print(result['data'][0])
                for q in result['data']:
                    if len(sku) == 10:
                        if q['fnsku'] == sku:
                            # print(q)
                            num_total += q['product_total']
                    else:
                        if q['sku'] == sku:
                            # print(q)
                            num_total += q['product_total']
        else:
            print(result1)
        # print(num_total)
        return num_total

    def delect_box(self, pa_name):
        self.sql()
        sql1 = "SELECT `状态` from `storage`.`warehouse` WHERE `箱号` = '%s'" % pa_name
        self.cursor.execute(sql1)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            if result[0]['状态'] == "已删除":
                return False, '这箱物料已删除'
            elif result[0]['状态'] == "已出货":
                return False, "这箱物料已出货"
            else:
                self.sql()
                status = "已删除"
                sql2 = "UPDATE `storage`.`warehouse` SET `状态` = '%s' WHERE `箱号` = '%s'" % (status, pa_name)
                self.cursor.execute(sql2)
                self.connection.commit()
                self.sql_close()
                return True, '删除成功'
        else:
            return False, '请查看箱号是否正确'

    def get_msg(self, index, index_data):
        self.sql()
        data_msg = []
        if index == '箱号':
            sql = "select * from `storage`.`warehouse` where `箱号` = '%s' and 状态 = '已打包'" % index_data
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            self.sql_close()
            if result:
                for i in result:
                    self.sql()
                    sql1 = "SELECT * FROM `data_read`.`listing` where `FNSKU`='%s'" % i['FNSKU1']
                    self.cursor.execute(sql1)
                    name = self.cursor.fetchall()[0]['品名']
                    self.sql_close()
                    if i['存放位置']:
                        msg = {'箱号': i['箱号'], 'fnsku': index_data, '品名': name, '装箱数量': i['数量1'], '装箱状态': i['状态'], '装箱位置': i['存放位置']}
                    else:
                        msg = {'箱号': i['箱号'], 'fnsku': index_data, '品名': name, '装箱数量': i['数量1'], '装箱状态': i['状态'],
                               '装箱位置': '这箱还没有匹配位置'}
                    data_msg.append(msg)
                return {'data_list': data_msg}
            else:
                return False
        if index == 'FNSKU':
            list_re = self.get_pa([index_data])
            # print(list_re)
            if list_re:
                time_now = self.write_re(list_re)
                return [f'../FNSKU装箱信息/{time_now}-箱号.xlsx', f'{time_now}-箱号']
            else:
                return False
        if index == '位置':
            sql = "select * from `storage`.`warehouse` where `存放位置` = '%s' and 状态 = '已打包'" % index_data
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            self.sql_close()
            if result:
                for i in result:
                    self.sql()
                    sql1 = "SELECT * FROM `data_read`.`listing` where `FNSKU`='%s'" % i['存放位置']
                    self.cursor.execute(sql1)
                    name = self.cursor.fetchall()[0]['品名']
                    self.sql_close()
                    msg = {'箱号': i['箱号'], 'fnsku': index_data, '品名': name, '装箱数量': i['数量1'], '装箱状态': i['状态'], '装箱位置': i['存放位置']}
                    data_msg.append(msg)
                    # print(msg)
                return {'data_list': data_msg}
            else:
                return False

    def get_pa(self, list_fnsku):
        list_re = {}
        for k in list_fnsku:
            list_pa = []
            self.sql()
            sql = "select * from `storage`.`warehouse` where `FNSKU1` = '%s' and `状态` = '已打包'" % k
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            if result:
                for i in result:
                    list_pa.append(i['箱号'])
            for i in list_pa:
                sql = "select * from `storage`.`relevance_hakone` where `箱号` = '%s'" % i
                self.cursor.execute(sql)
                result = self.cursor.fetchall()
                if result:
                    if result[0]['关联箱标'] in list_re:
                        list_re[result[0]['关联箱标']][i] = 1
                    else:
                        list_re[result[0]['关联箱标']] = {}
                        sql1 = "select * from `storage`.`relevance_hakone` where `关联箱标` = '%s'" % result[0]['关联箱标']
                        self.cursor.execute(sql1)
                        result1 = self.cursor.fetchall()
                        for j in result1:
                            list_re[result[0]['关联箱标']][j['箱号']] = 0
                        list_re[result[0]['关联箱标']][i] = 1
                else:
                    list_re[i] = 1
            self.sql_close()
        return list_re

    def write_re(self, list_re):
        wb = openpyxl.Workbook()
        wb_sheet = wb.active
        wb_sheet.append(['关联箱号', 'FNSKU', 'SKU', '品名', '箱号', '存放位置', '数量', '总数量'])
        row1 = 1
        for i in list_re:
            if i.find('PA') >= 0:
                msg = self.get_pa_msg1(i, '', 0)
                del msg[7]
                wb_sheet.append(msg)
                for q in range(2, 8):
                    wb_sheet.cell(row=row1+1, column=q).font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color='DC143C')
                row1 += 1
            if i.find('RE') >= 0:
                list_pa = []
                for k in list_re[i]:
                    msg = self.get_pa_msg1(k, i, list_re[i][k])
                    if msg[7]:
                        del msg[7]
                        wb_sheet.append(msg)
                        for q in range(2, 8):
                            wb_sheet.cell(row=row1+1, column=q).font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color='DC143C')
                        row1 += 1
                    else:
                        del msg[7]
                        list_pa.append(msg)
                if list_pa:
                    for k in list_pa:
                        wb_sheet.append(k)
                        row1 += 1
        # print(row1)
        starrow1 = 2
        starrow2 = 2
        last_cell_value = wb_sheet.cell(row=starrow1, column=1).value
        last_cell_value1 = wb_sheet.cell(row=starrow2, column=2).value
        count = 0
        count1 = 0
        for q in range(0, row1):
            cell_value = wb_sheet.cell(row=starrow1 + count, column=1).value
            if cell_value == last_cell_value:
                count = count + 1
            else:
                if last_cell_value:
                    wb_sheet.merge_cells('A%d:A%d' % (starrow1, starrow1 - 1 + count))
                starrow1 = starrow1 + count
                count = 1
                last_cell_value = cell_value
            cell_value1 = wb_sheet.cell(row=starrow2 + count1, column=2).value
            if cell_value1 == last_cell_value1:
                count1 = count1 + 1
            else:
                if last_cell_value1:
                    num_total = 0
                    for i in range(starrow2, starrow2 + count1):
                        num = int(wb_sheet.cell(row=i, column=7).value)
                        num_total += num
                    for i in range(starrow2, starrow2 + count1):
                        wb_sheet.cell(row=i, column=8).value = num_total
                    wb_sheet.merge_cells('B%d:B%d' % (starrow2, starrow2 - 1 + count1))
                    wb_sheet.merge_cells('C%d:C%d' % (starrow2, starrow2 - 1 + count1))
                    wb_sheet.merge_cells('D%d:D%d' % (starrow2, starrow2 - 1 + count1))
                    wb_sheet.merge_cells('H%d:H%d' % (starrow2, starrow2 - 1 + count1))
                starrow2 = starrow2 + count1
                count1 = 1
                last_cell_value1 = cell_value1
        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        file = f'./static/FNSKU装箱信息/{time_now}-箱号.xlsx'
        wb.save(file)
        return time_now

    def get_pa_msg1(self, pa_name, re_name, index):
        self.sql()
        msg = ''
        sql = "select * from `storage`.`warehouse` where `箱号` = '%s' and 状态 = '已打包'" % pa_name
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            for i in result:
                self.sql()
                sql1 = "SELECT * FROM `data_read`.`listing` where `FNSKU`='%s'" % i['FNSKU1']
                self.cursor.execute(sql1)
                result1 = self.cursor.fetchall()
                name = result1[0]['品名']
                sku = result1[0]['SKU']
                self.sql_close()
                if i['存放位置']:
                    msg = [re_name, i['FNSKU1'], sku, name, i['箱号'], i['存放位置'], i['数量1'], index]
                else:
                    msg = [re_name, i['FNSKU1'], sku, name, i['箱号'], '这箱还没有匹配位置', i['数量1'], index]
            return msg
        else:
            return False

    def get_pa_msg(self, pa_name):
        self.sql()
        num_location = 0
        msg = ''
        sql = "select * from `storage`.`warehouse` where `箱号` = '%s' and 状态 = '已打包'" % pa_name
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            for i in result:
                num_location = int(i['数量1'])
                self.sql()
                sql1 = "SELECT * FROM `data_read`.`listing` where `FNSKU`='%s'" % i['FNSKU1']
                self.cursor.execute(sql1)
                name = self.cursor.fetchall()[0]['品名']
                self.sql_close()
                if i['存放位置']:
                    msg = {'箱号': i['箱号'], 'fnsku': i['FNSKU1'], '品名': name, '装箱数量': i['数量1'], '装箱状态': i['状态'],
                           '装箱位置': i['存放位置']}
                else:
                    msg = {'箱号': i['箱号'], 'fnsku': i['FNSKU1'], '品名': name, '装箱数量': i['数量1'], '装箱状态': i['状态'],
                           '装箱位置': '这箱还没有匹配位置'}
            return num_location, msg
        else:
            return False, False

    def get_warehouse(self, filename):
        ws = openpyxl.load_workbook(filename)
        ws_sheet = ws.active
        row1 = ws_sheet.max_row
        for i in range(row1, 0, -1):
            cell_value1 = ws_sheet.cell(row=i, column=1).value
            if cell_value1:
                row1 = i
                break
        list_fnsku = []
        for i in range(1, row1 + 1):
            fnsku = ws_sheet.cell(row=i, column=2).value
            if fnsku:
                fnsku = fnsku.strip('')
                if fnsku in list_fnsku:
                    continue
                else:
                    list_fnsku.append(fnsku)
        list_re = self.get_pa(list_fnsku)
        # print(list_re)
        time_now = self.write_re(list_re)
        return [f'../FNSKU装箱信息/{time_now}-箱号.xlsx', f'{time_now}-箱号']


######## AES-128-ECS 加密
class EncryptDate:
    def __init__(self, key):
        self.key = key.encode("utf-8")  # 初始化密钥
        self.length = AES.block_size  # 初始化数据块大小
        self.aes = AES.new(self.key, AES.MODE_ECB)  # 初始化AES,ECB模式的实例
        # 截断函数，去除填充的字符
        self.unpad = lambda date: date[0:-ord(date[-1])]

    def pad(self, text):
        """
        #填充函数，使被加密数据的字节码长度是block_size的整数倍
        """
        count = len(text.encode('utf-8'))
        add = self.length - (count % self.length)
        entext = text + (chr(add) * add)
        return entext

    def encrypt(self, encrData):  # 加密函数
        res = self.aes.encrypt(self.pad(encrData).encode("utf8"))
        msg = str(base64.b64encode(res), encoding="utf8")
        return msg

    def decrypt(self, decrData):  # 解密函数
        res = base64.decodebytes(decrData.encode("utf8"))
        msg = self.aes.decrypt(res).decode("utf8")
        return self.unpad(msg)
