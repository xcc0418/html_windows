import openpyxl
import pymysql
import datetime


class Quantity(object):
    def __init__(self):
        self.connection = None
        self.cursor = None

    def sql(self):
        self.connection = pymysql.connect(host='3354n8l084.goho.co',  # 数据库地址
                                          port=24824,
                                          user='test_user',  # 数据库用户名
                                          password='a123456',  # 数据库密码
                                          db='storage',  # 数据库名称
                                          charset='utf8',
                                          cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor()

    def sql_close(self):
        self.cursor.close()
        self.connection.close()

    def open_downloads(self):
        self.sql()
        sql = "update `flag`.`amazon_form_flag` set `flag_num` = 1 where `flag_name` = 'parent'"
        self.cursor.execute(sql)
        self.connection.commit()
        self.sql_close()

    def close_downloads(self):
        self.sql()
        sql = "update `flag`.`amazon_form_flag` set `flag_num` = 0 where `flag_name` = 'parent'"
        self.cursor.execute(sql)
        self.connection.commit()
        self.sql_close()

    # 本地父体新增
    def add_male(self, male_parent):
        self.sql()
        sql = "select * from `amazon_form`.`list_parent` where `本地父体` = '%s'" % male_parent
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            return False, "已经有这个本地父体了，请检查"
        else:
            try:
                sql1 = "insert into `amazon_form`.`list_parent`(`本地父体`)values('%s')" % male_parent
                self.cursor.execute(sql1)
                self.connection.commit()
                self.sql_close()
                return True, True
            except Exception as e:
                self.connection.rollback()
                self.sql_close()
                return False, str(e)

    # 批量新增本地父体
    def upload_male(self, filename):
        self.sql()
        sql = "select * from `flag`.`amazon_form_flag` where `flag_name` = 'parent'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result[0]['flag_num'] == 0:
            try:
                list_error = []
                self.open_downloads()
                wb = openpyxl.load_workbook(filename)
                wb_sheet = wb.active
                row1 = wb_sheet.max_row
                for i in range(row1, 0, -1):
                    cell_value1 = wb_sheet.cell(row=i, column=1).value
                    if cell_value1:
                        row1 = i
                        break
                for i in range(2, row1+1):
                    male_parent = wb_sheet.cell(row=i, column=1).value
                    if male_parent:
                        msg, message = self.add_male(male_parent)
                        if msg or message == "已经有这个本地父体了，请检查":
                            continue
                        else:
                            list_error.append(male_parent)
                        wb_sheet.cell(row=i, column=2).value = message
                wb.save(filename)
                self.close_downloads()
                if list_error:
                    msg = ''
                    for i in list_error:
                        msg = f"{msg}{i}添加失败, "
                    msg = f"{msg}其余添加成功。"
                    return False, msg
                return True, True
            except Exception as e:
                print(e)
                self.close_downloads()
                return False, str(e)
        else:
            return False, "当前正在进行添加操作，请稍后再试"

    # 批量新增本地品名
    def upload_sku(self, filename):
        self.sql()
        sql = "select * from `flag`.`amazon_form_flag` where `flag_name` = 'parent'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result[0]['flag_num'] == 0:
            try:
                list_error = []
                self.open_downloads()
                wb = openpyxl.load_workbook(filename)
                wb_sheet = wb.active
                row1 = wb_sheet.max_row
                for i in range(row1, 0, -1):
                    cell_value1 = wb_sheet.cell(row=i, column=1).value
                    if cell_value1:
                        row1 = i
                        break
                for i in range(2, row1+1):
                    male_sku = wb_sheet.cell(row=i, column=1).value
                    if male_sku:
                        msg, message = self.add_sku(male_sku)
                        # print(message)
                        if msg or message == "已经有这个本地品名了，请检查":
                            continue
                        else:
                            list_error.append(male_sku)
                        wb_sheet.cell(row=i, column=2).value = message
                wb.save(filename)
                self.close_downloads()
                if list_error:
                    msg = ''
                    for i in list_error:
                        msg = f"{msg}{i}添加失败, "
                    msg = f"{msg}其余添加成功。"
                    return False, msg
                return True, True
            except Exception as e:
                print(e)
                self.close_downloads()
                return False, str(e)
        else:
            return False, "当前正在进行添加操作，请稍后再试"

    # 批量关联本地父体
    def upload_parent(self, filename):
        self.sql()
        sql = "select * from `flag`.`amazon_form_flag` where `flag_name` = 'parent'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result[0]['flag_num'] == 0:
            try:
                list_error = []
                self.open_downloads()
                wb = openpyxl.load_workbook(filename)
                wb_sheet = wb.active
                row1 = wb_sheet.max_row
                for i in range(row1, 0, -1):
                    cell_value1 = wb_sheet.cell(row=i, column=1).value
                    if cell_value1:
                        row1 = i
                        break
                for i in range(2, row1+1):
                    male_parent = wb_sheet.cell(row=i, column=1).value
                    if male_parent:
                        sku = wb_sheet.cell(row=i, column=2).value
                        if sku:
                            msg, message = self.relevance_male(male_parent, sku)
                            if msg:
                                continue
                            else:
                                if message.find('已关联') >= 0:
                                    continue
                                list_error.append([male_parent, sku])
                        else:
                            male_sku = wb_sheet.cell(row=i, column=3).value
                            list_sku, message = self.get_male_sku("本地品名", male_sku)
                            if list_sku:
                                flag = 1
                                for j in list_sku:
                                    sku = j[1]
                                    msg, message = self.relevance_male(male_parent, sku)
                                    if msg:
                                        continue
                                    else:
                                        if message.find('已关联') >= 0:
                                            continue
                                        flag = 0
                                if flag:
                                    message = "关联成功"
                                else:
                                    list_error.append([male_parent, sku])
                                    message = "关联失败"
                            else:
                                message = "没有找到这个本地品名"
                        wb_sheet.cell(row=i, column=3).value = message
                time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                wb.save(f'./static/本地父体/关联本地父体_{time_now}.xlsx')
                self.close_downloads()
                if list_error:
                    msg = ''
                    for i in list_error:
                        msg = f"{msg}{i[0]}与{i[1]}关联失败, "
                    msg = f"{msg}其余关联成功。"
                    return False, msg
                return True, True
            except Exception as e:
                self.close_downloads()
                return False, str(e)
        else:
            return False, "当前正在进行关联操作，请稍后再试"

    # 批量关联本地品名
    def upload_name(self, filename):
        self.sql()
        sql = "select * from `flag`.`amazon_form_flag` where `flag_name` = 'parent'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result[0]['flag_num'] == 0:
            try:
                list_error = []
                list_success = []
                self.open_downloads()
                wb = openpyxl.load_workbook(filename)
                wb_sheet = wb.active
                row1 = wb_sheet.max_row
                for i in range(row1, 0, -1):
                    cell_value1 = wb_sheet.cell(row=i, column=1).value
                    if cell_value1:
                        row1 = i
                        break
                for i in range(2, row1+1):
                    male_parent = wb_sheet.cell(row=i, column=1).value
                    sku = wb_sheet.cell(row=i, column=2).value
                    msg, message = self.relevance_sku(male_parent, sku)
                    if msg:
                        list_success.append([sku, male_parent])
                    else:
                        list_error.append([male_parent, sku])
                    wb_sheet.cell(row=i, column=3).value = message
                if list_success:
                    self.add_image_msg(list_success)
                time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                wb.save(f'./static/本地品名/关联本地品名_{time_now}.xlsx')
                self.close_downloads()
                if list_error:
                    msg = ''
                    for i in list_error:
                        msg = f"{msg}{i[0]}与{i[1]}关联失败, "
                    msg = f"{msg}其余关联成功。"
                    return False, msg
                return True, True
            except Exception as e:
                print(e)
                self.close_downloads()
                return False, str(e)
        else:
            return False, "当前正在进行关联操作，请稍后再试"

    # 新增图片状态监控
    def add_image_msg(self, list_sku):
        self.sql()
        sql = "select * from `data_read`.`product_image`"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        list_image = []
        for i in result:
            list_image.append([i['SKU']])
        for i in list_sku:
            if i[0] in list_image:
                continue
            sql1 = "select * from `data_read`.`product_id` where `SKU` = '%s'" % i[0]
            self.cursor.execute(sql1)
            result = self.cursor.fetchall()
            name = '暂无品名'
            values = '无图片链接'
            if result:
                name = result[0]['品名']
                values = result[0]['图片状态']
            sql2 = "insert into `data_read`.`product_image`(`本地品名`, `SKU`,`品名`, `状态`) VALUES ('%s','%s','%s','%s')" % (
                i[1], i[0], name, values)
            self.cursor.execute(sql2)
        self.connection.commit()
        self.sql_close()

    # 删除图片状态监控
    def delete_image_msg(self, list_sku):
        self.sql()
        for i in list_sku:
            sql1 = "DELETE FROM `data_read`.`product_image` where `SKU` = '%s'" % i[0]
            self.cursor.execute(sql1)
        self.connection.commit()
        self.sql_close()

    # 本地品名新增
    def add_sku(self, local_sku):
        self.sql()
        sql = "select * from `amazon_form`.`list_local_sku` where `本地品名` = '%s'" % local_sku
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            return False, "已经有这个本地品名了，请检查"
        else:
            try:
                sql1 = "insert into `amazon_form`.`list_local_sku`(`本地品名`)values('%s')" % local_sku
                self.cursor.execute(sql1)
                self.connection.commit()
                self.sql_close()
                return True, True
            except Exception as e:
                self.connection.rollback()
                self.sql_close()
                return False, str(e)

    # 关联本地父体
    def relevance_male(self, male_parent, sku):
        self.sql()
        # 查询是否有这个本地父体
        sql = "select * from `amazon_form`.`list_parent` where `本地父体` = '%s'" % male_parent
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            # 查询这个本地父体与SKU是否已关联
            sql = "select * from `amazon_form`.`male_parent` where `本地父体` = '%s' and `SKU` = '%s'" % (male_parent, sku)
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            if result:
                print(f"{male_parent}与{sku}已关联")
                return False, f"{male_parent}与{sku}已关联"
            else:
                # 配对本地父体是与SKU关联还是与本地品名关联
                sql2 = "select * from `amazon_form`.`male_sku` where `SKU` = '%s'" % sku
                self.cursor.execute(sql2)
                result1 = self.cursor.fetchall()
                if result1:
                    # 获取这个SKU关联的本地品名关所联的所有SKU
                    sql3 = "select * from `amazon_form`.`male_sku` where `本地品名` = '%s'" % result1[0]['本地品名']
                    self.cursor.execute(sql3)
                    result2 = self.cursor.fetchall()
                    for i in result2:
                        # 防止本地父体与SKU重复关联
                        sql5 = "select * from `amazon_form`.`male_parent` where `本地父体` = '%s' and `SKU` = '%s'" % (male_parent, i['SKU'])
                        self.cursor.execute(sql5)
                        result = self.cursor.fetchall()
                        if result:
                            continue
                        else:
                            sql4 = "insert into `amazon_form`.`male_parent`(`本地父体`, `SKU`)values('%s', '%s')" % (male_parent, i['SKU'])
                            self.cursor.execute(sql4)
                else:
                    # 获取这个本地品名关所联的所有SKU
                    sql3 = "select * from `amazon_form`.`male_sku` where `本地品名` = '%s'" % sku
                    self.cursor.execute(sql3)
                    result2 = self.cursor.fetchall()
                    if result2:
                        # 获取这个本地父体关所联的所有SKU
                        sql5 = "select * from `amazon_form`.`male_parent` where `本地父体` = '%s'" % male_parent
                        self.cursor.execute(sql5)
                        result5 = self.cursor.fetchall()
                        list_male_sku = []
                        for i in result5:
                            list_male_sku.append(i['SKU'])
                        for i in result2:
                            # 防止本地父体与SKU重复关联
                            if i['SKU'] in list_male_sku:
                                continue
                            sql4 = "insert into `amazon_form`.`male_parent`(`本地父体`, `SKU`)values('%s', '%s')" % (male_parent, i['SKU'])
                            self.cursor.execute(sql4)
                    else:
                        # 这个sku没有与本地品名关联时
                        sql1 = "insert into `amazon_form`.`male_parent`(`本地父体`, `SKU`)values('%s', '%s')" % (male_parent, sku)
                        self.cursor.execute(sql1)
                try:
                    self.connection.commit()
                    return True, '成功'
                except Exception as e:
                    # print(e)
                    self.connection.rollback()
                    return False, str(e)
        else:
            self.sql_close()
            # print(f"没有{male_parent}这个本地父体，请检查")
            return False, f"没有{male_parent}这个本地父体，请检查"

    # 解除本地父体关联
    def relieve_male(self, male_parent, sku):
        self.sql()
        sql = "select * from `amazon_form`.`male_parent` where `本地父体` = '%s' and `SKU` = '%s'" % (male_parent, sku)
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        sql = "select * from `amazon_form`.`male_sku` where `本地品名` = '%s'" % sku
        self.cursor.execute(sql)
        result3 = self.cursor.fetchall()
        # 查询这个本地父体与SKU是否已关联或者这个是否是本地品名
        if result or result3:
            if result:
                sql2 = "select * from `amazon_form`.`male_sku` where `SKU` = '%s'" % sku
            else:
                sql2 = "select * from `amazon_form`.`male_sku` where `本地品名` = '%s'" % sku
            # 获取要解除配对的所有SKU
            self.cursor.execute(sql2)
            result1 = self.cursor.fetchall()
            if result1:
                sql3 = "select * from `amazon_form`.`male_sku` where `本地品名` = '%s'" % result1[0]['本地品名']
                self.cursor.execute(sql3)
                result2 = self.cursor.fetchall()
                for i in result2:
                    sql4 = "DELETE FROM `amazon_form`.`male_parent` WHERE `本地父体` = '%s' and `SKU` = '%s'" % (male_parent, i['SKU'])
                    self.cursor.execute(sql4)
            else:
                sql1 = "DELETE FROM `amazon_form`.`male_parent` WHERE `本地父体` = '%s' and `SKU` = '%s'" % (male_parent, sku)
                self.cursor.execute(sql1)
            try:
                self.connection.commit()
                self.sql_close()
                return True, '成功'
            except Exception as e:
                self.connection.rollback()
                self.sql_close()
                return False, str(e)
        else:
            self.sql_close()
            return False, "这个父体与sku没有相关信息"

    # 关联本地品名
    def relevance_sku(self, local_sku, sku):
        self.sql()
        sql = "select * from `amazon_form`.`list_local_sku` where `本地品名` = '%s'" % local_sku
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            sql = "select * from `amazon_form`.`male_sku` where `SKU` = '%s'" % sku
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            if result:
                self.sql_close()
                return False, f"{result[0]['本地品名']}与{sku}已关联"
            else:
                sql1 = "insert into `amazon_form`.`male_sku`(`本地品名`, `SKU`)values('%s', '%s')" % (local_sku, sku)
                try:
                    self.cursor.execute(sql1)
                    self.connection.commit()
                    self.sql_close()
                    return True, '成功'
                except Exception as e:
                    self.connection.rollback()
                    self.sql_close()
                    return False, str(e)
        else:
            self.sql_close()
            return False, f"没有{local_sku}这个本地品名，请检查"

    # 解除本地品名关联
    def relieve_sku(self, local_sku, sku):
        self.sql()
        sql = "select * from `amazon_form`.`male_sku` where `本地品名` = '%s' and `SKU` = '%s'" % (local_sku, sku)
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            sql1 = "DELETE FROM `amazon_form`.`male_sku` WHERE `本地品名` = '%s' and `SKU` = '%s'" % (local_sku, sku)
            try:
                self.cursor.execute(sql1)
                self.connection.commit()
                self.sql_close()
                return True, True
            except Exception as e:
                self.connection.rollback()
                self.sql_close()
                return False, str(e)
        else:
            self.sql_close()
            return False, "这个父体与sku没有相关信息"

    # 关联本地父体与亚马逊父体
    def relevance_amazon(self, male_parent, amazon_parent):
        self.sql()
        sql1 = "select * from `amazon_form`.`list_parent` where `本地父体` = '%s'" % male_parent
        self.cursor.execute(sql1)
        result1 = self.cursor.fetchall()
        # print(sql1)
        if result1:
            sql2 = "select * from `amazon_form`.`male_amazon` where `本地父体` = '%s' and `亚马逊父体` = '%s'" % (male_parent, amazon_parent)
            self.cursor.execute(sql2)
            result2 = self.cursor.fetchall()
            if result2:
                self.sql_close()
                return False, f'{male_parent}与{amazon_parent}已关联'
            else:
                try:
                    sql3 = "insert into `amazon_form`.`male_amazon`(`本地父体`, `亚马逊父体`)values('%s', '%s')" % (male_parent, amazon_parent)
                    self.cursor.execute(sql3)
                    self.connection.commit()
                    self.sql_close()
                    return True, True
                except Exception as e:
                    self.connection.rollback()
                    self.sql_close()
                    return False, str(e)
        else:
            return False, f'没有{male_parent}，请检查'

    # 解除本地父体与亚马逊父体关联
    def relieve_amazon(self, male_parent, amazon_parent):
        self.sql()
        sql1 = "select * from `amazon_form`.`male_amazon` where `本地父体` = '%s' and `亚马逊父体` = '%s'" % (male_parent, amazon_parent)
        self.cursor.execute(sql1)
        result1 = self.cursor.fetchall()
        if result1:
            try:
                sql2 = "DELETE FROM `amazon_form`.`male_amazon` WHERE `本地父体` = '%s' and `亚马逊父体` = '%s'" % (male_parent, amazon_parent)
                self.cursor.execute(sql2)
                self.connection.commit()
                self.sql_close()
                return True, True
            except Exception as e:
                self.connection.rollback()
                self.sql_close()
                return False, str(e)
        else:
            return False, f"{male_parent}与{amazon_parent}没有关联信息，请检查"

    # 批量解除本地父体关联
    def upload_relieve_male(self, filename):
        self.sql()
        sql = "select * from `flag`.`amazon_form_flag` where `flag_name` = 'parent'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result[0]['flag_num'] == 0:
            try:
                list_error = []
                self.open_downloads()
                wb = openpyxl.load_workbook(filename)
                wb_sheet = wb.active
                row1 = wb_sheet.max_row
                for i in range(row1, 0, -1):
                    cell_value1 = wb_sheet.cell(row=i, column=1).value
                    if cell_value1:
                        row1 = i
                        break
                for i in range(2, row1 + 1):
                    male_parent = wb_sheet.cell(row=i, column=1).value
                    sku = wb_sheet.cell(row=i, column=2).value
                    if sku:
                        msg, message = self.relieve_male(male_parent, sku)
                    else:
                        male_sku = wb_sheet.cell(row=i, column=3).value
                        list_sku, message = self.get_male_sku("本地品名", male_sku)
                        if list_sku:
                            flag = 1
                            for j in list_sku:
                                sku = j[1]
                                msg, message = self.relieve_male(male_parent, sku)
                                if msg:
                                    continue
                                else:
                                    flag = 0
                            if flag:
                                message = "解除成功"
                            else:
                                list_error.append([male_parent, sku])
                                message = "解除失败"
                        else:
                            message = "没有找到这个本地品名"
                    wb_sheet.cell(row=i, column=3).value = message
                time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                wb.save(f'./static/本地父体/解除本地父体关联_{time_now}.xlsx')
                self.close_downloads()
                if list_error:
                    msg = ''
                    for i in list_error:
                        msg = f"{msg}{i[0]}与{i[1]}解除关联失败, "
                    msg = f"{msg}其余解除关联成功。"
                    return False, msg
                return True, True
            except Exception as e:
                print(e)
                self.close_downloads()
                return False, str(e)
        else:
            return False, "当前正在进行关联操作，请稍后再试"

    # 批量解除本地品名关联
    def upload_relieve_sku(self, filename):
        self.sql()
        sql = "select * from `flag`.`amazon_form_flag` where `flag_name` = 'parent'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result[0]['flag_num'] == 0:
            try:
                list_error = []
                list_succes = []
                self.open_downloads()
                wb = openpyxl.load_workbook(filename)
                wb_sheet = wb.active
                row1 = wb_sheet.max_row
                for i in range(row1, 0, -1):
                    cell_value1 = wb_sheet.cell(row=i, column=1).value
                    if cell_value1:
                        row1 = i
                        break
                for i in range(2, row1 + 1):
                    male_parent = wb_sheet.cell(row=i, column=1).value
                    sku = wb_sheet.cell(row=i, column=2).value
                    msg, message = self.relieve_sku(male_parent, sku)
                    if msg:
                        list_succes.append([sku, male_parent])
                    else:
                        list_error.append([male_parent, sku])
                    wb_sheet.cell(row=i, column=3).value = message
                if list_succes:
                    self.delete_image_msg(list_succes)
                time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                wb.save(f'./static/本地品名/解除本地品名关联_{time_now}.xlsx')
                self.close_downloads()
                if list_error:
                    msg = ''
                    for i in list_error:
                        msg = f"{msg}{i[0]}与{i[1]}解除关联失败, "
                    msg = f"{msg}其余解除关联成功。"
                    return False, msg
                return True, True
            except Exception as e:
                print(e)
                self.close_downloads()
                return False, str(e)
        else:
            return False, "当前正在进行关联操作，请稍后再试"

    # 删除本地父体
    def delete_male(self, male_parent):
        self.sql()
        sql = "select * from `amazon_form`.`list_parent` where `本地父体` = '%s'" % male_parent
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            sql1 = "DELETE FROM `amazon_form`.`male_parent` WHERE `本地父体` = '%s'" % male_parent
            sql2 = "DELETE FROM `amazon_form`.`list_parent` WHERE `本地父体` = '%s'" % male_parent
            try:
                self.cursor.execute(sql1)
                self.cursor.execute(sql2)
                self.connection.commit()
                self.sql_close()
                return True, True
            except Exception as e:
                self.connection.rollback()
                self.sql_close()
                return False, str(e)
        else:
            self.sql_close()
            return False, f"没有{male_parent}这个父体"

    # 删除本地品名
    def delete_sku(self, male_parent):
        self.sql()
        sql = "select * from `amazon_form`.`list_local_sku` where `本地品名` = '%s'" % male_parent
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            sql1 = "DELETE FROM `amazon_form`.`male_sku` WHERE `本地品名` = '%s'" % male_parent
            sql2 = "DELETE FROM `amazon_form`.`list_local_sku` WHERE `本地品名` = '%s'" % male_parent
            try:
                self.cursor.execute(sql1)
                self.cursor.execute(sql2)
                self.connection.commit()
                self.sql_close()
                return True, True
            except Exception as e:
                self.connection.rollback()
                self.sql_close()
                return False, str(e)
        else:
            self.sql_close()
            return False, f"没有{male_parent}这个父体"

    # 批量删除本地父体
    def upload_delete_male(self, filename):
        self.sql()
        sql = "select * from `flag`.`amazon_form_flag` where `flag_name` = 'parent'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result[0]['flag_num'] == 0:
            try:
                list_error = []
                self.open_downloads()
                wb = openpyxl.load_workbook(filename)
                wb_sheet = wb.active
                row1 = wb_sheet.max_row
                for i in range(row1, 0, -1):
                    cell_value1 = wb_sheet.cell(row=i, column=1).value
                    if cell_value1:
                        row1 = i
                        break
                for i in range(2, row1 + 1):
                    male_parent = wb_sheet.cell(row=i, column=1).value
                    msg, message = self.delete_male(male_parent)
                    if msg:
                        continue
                    else:
                        list_error.append(i)
                    wb_sheet.cell(row=i, column=3).value = message
                time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                wb.save(f'./static/本地父体/删除本地父体_{time_now}.xlsx')
                self.close_downloads()
                if list_error:
                    msg = ''
                    for i in list_error:
                        msg = f"{msg}{i}删除本失败, "
                    msg = f"{msg}其余删除本成功。"
                    return False, msg
                return True, True
            except Exception as e:
                print(e)
                self.close_downloads()
                return False, str(e)
        else:
            return False, "当前正在进行关联操作，请稍后再试"

    # 批量删除本地品名
    def upload_delete_sku(self, filename):
        self.sql()
        sql = "select * from `flag`.`amazon_form_flag` where `flag_name` = 'parent'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result[0]['flag_num'] == 0:
            try:
                list_error = []
                self.open_downloads()
                wb = openpyxl.load_workbook(filename)
                wb_sheet = wb.active
                row1 = wb_sheet.max_row
                for i in range(row1, 0, -1):
                    cell_value1 = wb_sheet.cell(row=i, column=1).value
                    if cell_value1:
                        row1 = i
                        break
                for i in range(2, row1 + 1):
                    male_parent = wb_sheet.cell(row=i, column=1).value
                    msg, message = self.delete_sku(male_parent)
                    if msg:
                        continue
                    else:
                        list_error.append(i)
                    wb_sheet.cell(row=i, column=3).value = message
                time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                wb.save (f'./static/本地品名/删除本地品名_{time_now}.xlsx')
                self.close_downloads()
                if list_error:
                    msg = ''
                    for i in list_error:
                        msg = f"{msg}{i}删除本失败, "
                    msg = f"{msg}其余删除本成功。"
                    return False, msg
                return True, True
            except Exception as e:
                print(e)
                self.close_downloads()
                return False, str(e)
        else:
            return False, "当前正在进行关联操作，请稍后再试"

    # 修改本地父体
    def change_parent(self, parent, change_parent):
        self.sql()
        sql1 = "select * from `amazon_form`.`list_parent` where `本地父体` = '%s'" % parent
        self.cursor.execute(sql1)
        result = self.cursor.fetchall()
        if result:
            sql2 = "UPDATE `amazon_form`.`list_parent` SET `本地父体` = '%s' WHERE `本地父体` = '%s'" % (change_parent, parent)
            sql3 = "UPDATE `amazon_form`.`male_parent` SET `本地父体` = '%s' WHERE `本地父体` = '%s'" % (change_parent, parent)
            try:
                self.cursor.execute(sql2)
                self.cursor.execute(sql3)
                self.connection.commit()
                self.sql_close()
                return True, True
            except Exception as e:
                self.connection.rollback()
                self.sql_close()
                print(e)
                return False, str(e)
        else:
            self.sql_close()
            return False, f'没有找到{parent}, 请检查'

    # 修改本地品名
    def change_sku(self, sku, change_sku):
        self.sql()
        sql1 = "select * from `amazon_form`.`list_local_sku` where `本地品名` = '%s'" % sku
        self.cursor.execute(sql1)
        result = self.cursor.fetchall()
        if result:
            sql2 = "UPDATE `amazon_form`.`list_local_sku` SET `本地品名` = '%s' WHERE `本地品名` = '%s'" % (change_sku, sku)
            sql3 = "UPDATE `amazon_form`.`male_sku` SET `本地品名` = '%s' WHERE `本地品名` = '%s'" % (change_sku, sku)
            try:
                self.cursor.execute(sql2)
                self.cursor.execute(sql3)
                self.connection.commit()
                self.sql_close()
                return True, True
            except Exception as e:
                self.connection.rollback()
                self.sql_close()
                print(e)
                return False, str(e)
        else:
            self.sql_close()
            return False, f'没有找到{sku}, 请检查'

    # 生成批量上传功能详情表
    def write_xlsx(self, list_data, list_header):
        wb = openpyxl.Workbook()
        wb_sheet = wb.active
        if list_header:
            list_header.insert(0, '本地品名')
            list_header.insert(0, '父体')
            wb_sheet.append(list_header)
        for i in list_data:
            if i[1] and i[1] != '...':
                wb_sheet.append(i)
        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        if list_data[0][0] == "本地父体":
            path = f'./static/本地父体/本地父体详情_{time_now}.xlsx'
            file = f'../本地父体/本地父体详情_{time_now}.xlsx'
            filename = f'本地父体详情_{time_now}'
            download_name = "本地父体详情"
        elif list_data[0][0] == "本地品名":
            path = f'./static/本地品名/本地品名详情_{time_now}.xlsx'
            file = f'../本地品名/本地品名详情_{time_now}.xlsx'
            filename = f'本地品名详情_{time_now}'
            download_name = "本地品名详情"
        else:
            if list_data[0][0].find('B0') >= 0:
                path = f'./static/本地父体/亚马逊父体详情_{time_now}.xlsx'
                file = f'../本地父体/亚马逊父体详情_{time_now}.xlsx'
                filename = f'亚马逊父体详情_{time_now}'
                download_name = "亚马逊父体详情"
            else:
                path = f'./static/本地父体/本地父体详情_{time_now}.xlsx'
                file = f'../本地父体/本地父体详情_{time_now}.xlsx'
                filename = f'本地父体详情_{time_now}'
                download_name = "本地父体详情"
        wb.save(path)
        return file, filename, download_name

    # 获取本地父体详情
    def get_male_parent(self, index, index_data):
        self.sql()
        if index == "本地父体":
            if index_data:
                sql = f"select * from `amazon_form`.`male_parent` where `本地父体` like '%{index_data}%'"
            else:
                # 获取所有本地父体，不含SKU数据
                sql = "select * from `amazon_form`.`list_parent`"
                self.cursor.execute(sql)
                result = self.cursor.fetchall()
                self.sql_close()
                if result:
                    list_data = []
                    for i in result:
                        list_data.append([i['本地父体']])
                    return list_data, 1
                else:
                    return False, 0
        else:
            if index_data:
                sql = f"select * from `amazon_form`.`male_parent` where `SKU` like '%{index_data}%'"
            else:
                sql = "select * from `amazon_form`.`male_parent`"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            dict_sku = self.get_dict_sku ()
            list_data = []
            for i in result:
                if i['SKU'] in dict_sku:
                    list_data.append([i['本地父体'], i['SKU'], dict_sku[i['SKU']]])
                else:
                    list_data.append([i['本地父体'], i['SKU'], '当前没有本地品名'])
            self.sql_close()
            return list_data, 0
        else:
            self.sql_close()
            return False, 0

    # 获取SKU与本地品名的关系
    def get_dict_sku(self):
        dict_sku = {}
        sql = "select * from `amazon_form`.`male_sku`"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        for i in result:
            dict_sku[i['SKU']] = i['本地品名']
        return dict_sku

    # 获取本地品名详情
    def get_male_sku(self, index, index_data):
        self.sql()
        if index == "本地品名":
            if index_data:
                sql = f"select * from `amazon_form`.`male_sku` where `本地品名` like '%{index_data}%'"
            else:
                # 获取所有本地品名。不含SKU
                sql = "select * from `amazon_form`.`list_local_sku`"
                self.cursor.execute(sql)
                result = self.cursor.fetchall()
                self.sql_close()
                if result:
                    list_data = []
                    for i in result:
                        list_data.append([i['本地品名']])
                    return list_data, 1
                else:
                    return False, 0
        else:
            if index_data:
                sql = f"select * from `amazon_form`.`male_sku` where `SKU` like '%{index_data}%'"
            else:
                sql = "select * from `amazon_form`.`male_sku`"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            list_data = []
            for i in result:
                list_data.append([i['本地品名'], i['SKU']])
            return list_data, 0
        else:
            return False, 0

    def read_xlsx(self, filename):
        wb = openpyxl.load_workbook(filename)
        wb_sheet = wb.active
        list_sku = []
        for i in range(2, 400):
            male_parent = wb_sheet.cell(row=i, column=1).value
            if male_parent:
                name = wb_sheet.cell(row=i, column=2).value
                sku = self.get_sku(name)
                if sku:
                    continue
                else:
                    list_sku.append(name)
                    print(name)
        print(list_sku)

    def test_sql(self):
        self.sql()
        sql = "SELECT `SKU` FROM `amazon_form`.`male_parent` WHERE `本地父体` = 'K11-K22父体'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        print(result)
        self.sql_close()

    def get_sku(self, name):
        self.sql()
        sql = "select * from `data_read`.`product_id` where `品名` = '%s'" % name
        # print(sql)
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            return result[0]['SKU']
        else:
            return False


if __name__ == '__main__':
    quantity = Quantity()
    # quantity.read_xlsx("F:/html_windows/static/本地父体/关联本地父体_20230428172405.xlsx")
    # quantity.test_sql()
    quantity.add_image_msg(['CB.1459', 'KPW5-透明保护壳-打印-CB080-浅绿底花海B'])
