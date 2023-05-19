import requests
import json
import time
import hashlib
import urllib
import base64
import datetime
import pymysql
from Crypto.Cipher import AES
import openpyxl
from zipfile import ZipFile
import zipfile
import global_var
import shutil
import copy
import os


######## AES-128-ECS 加密
class EncryptDate:
    def __init__(self, key):
        self.key = key.encode("utf-8")  # 初始化密钥
        self.length = AES.block_size  # 初始化数据块大小
        self.aes = AES.new(self.key, AES.MODE_ECB)  # 初始化AES,ECB模式的实例
        # 截断函数，去除填充的字符
        self.unpad = lambda date: date[0:-ord(date[-1])]

    def pad(self, text):
        """
        #填充函数，使被加密数据的字节码长度是block_size的整数倍
        """
        count = len(text.encode('utf-8'))
        add = self.length - (count % self.length)
        entext = text + (chr(add) * add)
        return entext

    def encrypt(self, encrData):  # 加密函数
        res = self.aes.encrypt(self.pad(encrData).encode("utf8"))
        msg = str(base64.b64encode(res), encoding="utf8")
        return msg

    def decrypt(self, decrData):  # 解密函数
        res = base64.decodebytes(decrData.encode("utf8"))
        msg = self.aes.decrypt(res).decode("utf8")
        return self.unpad(msg)


class Quantity(object):
    def __init__(self):
        self.app_id = "ak_nEfE94OSogf3x"
        app_secret = "g2BcerjK4fWmhGZoetCJHeVqJmHEmfLt3gWFjDrBLB1yxiapgUKH6kOVs2N9JH7SFuBKuOF8K/CrNNSeFO1KKtsL05z24j" \
                     "+AdWTW+V4op5QxDkmlTllvlprT8FfjctDdNDGrwHvBvE6s9h0pO0dNgopBAYiA7oosPzQhDF1A6XC1X/cZZmgBy3XRHyEv" \
                     "xTT40xzwVGish53R8dZt3YIxNtKSgrBloo/CRQsV01yU40nyQR9L9oML32VT0C16jBrxcoWthlGDwfBn+CtVUvws4imyZi" \
                     "+sG/CqZQeaVLkBCqLCgqw1VK4/a4jZws6HO3FMeBgvuf0aS5euNmfhkudQmg=="
        querystring = {"appId": f"{self.app_id}", "appSecret": f"{app_secret}"}
        url = "https://openapi.lingxing.com/api/auth-server/oauth/access-token"
        payload = ""
        headers = {
            "Content-Type": "application/x-www-form-urlencoded"
        }
        response = requests.post(url, data=payload, headers=headers, params=querystring)
        result = json.loads(response.text)
        # print(response.text)
        self.access_token = result['data']['access_token']
        self.refresh_token = result['data']['refresh_token']
        self.time = datetime.datetime.now().strftime("%Y-%m-%d")
        self.time = datetime.datetime.strptime(self.time, "%Y-%m-%d")
        # self.time = "2021-12-06"
        self.start_time = self.time - datetime.timedelta(days=1)
        self.time = self.time.strftime("%Y-%m-%d")
        self.start_time = self.start_time.strftime("%Y-%m-%d")
        # self.start_time = '2022-03-07'

    def sql(self):
        self.connection = pymysql.connect(host='3354n8l084.goho.co',  # 数据库地址
                                          port=24824,
                                          user='test_user',  # 数据库用户名
                                          password='a123456',  # 数据库密码
                                          db='storage',  # 数据库名称
                                          charset='utf8',
                                          cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor()

    def sql_close(self):
        self.cursor.close()
        self.connection.close()

    def get_sign(self, body):
        apikey = "ak_nEfE94OSogf3x"
        # body = {"access_token": self.access_token,
        #         "timestamp": int(time.time()),
        #         "start_date": f"{self.start_time}",
        #         "end_date": f"{self.time}",
        #         "app_key": apikey}
        # print(body)
        # bb = sign().get_sign(apikey, body)
        # print(bb)

        # 目标md5串
        str_parm = ''
        # 将字典中的key排序
        for p in sorted(body):
            # 每次排完序的加到串中
            # if body[p]:
            # str类型需要转化为url编码格式
            if isinstance(body[p], str):
                str_parm = str_parm + str (p) + "=" + str(urllib.parse.quote (body[p])) + "&"
                continue
            # if isinstance(body[p], dict):
            #     for i in body[p]:
            #         body[p][i] = str(body[p][i])
            # if isinstance(body[p], list) and isinstance(body[p][0], dict):
            #     for i in range(0, len(body[p])):
            #         for j in body[p][i]:
            #             body[p][i][j] = str(urllib.parse.quote(str(body[p][i][j])))
            # if p == "product_list":
            #     str_parm = str_parm + str(p) + "=" + '[{"sku":"GCWL.897","good_num":1,"bad_num":0,"seller_id":0,"fnsku":""}]' + "&"
            #     continue
            #     body[p][0] = str(urllib.parse.quote(body[p][0]))
            #     print(str(urllib.parse.quote(body[p][0])))
            str_value = str(body[p])
            str_value = str_value.replace(" ", "")
            # str_value = str_value.replace("?", " ")
            str_value = str_value.replace("'", "\"")
            str_parm = str_parm + str(p) + "=" + str_value + "&"
        # 加上对应的key
        str_parm = str_parm.rstrip('&')
        # print("字符串拼接1:", str_parm)
        str_parm = str_parm.replace("?", " ")
        # print("字符串拼接2:", str_parm)
        if isinstance (str_parm, str):
            # 如果是unicode先转utf-8
            parmStr = str_parm.encode("utf-8")
            # parmStr = str_parm
            m = hashlib.md5 ()
            m.update (parmStr)
            md5_sign = m.hexdigest()
            # print(m.hexdigest())
            md5_sign = md5_sign.upper()
            # print("MD5加密:", md5_sign)
        eg = EncryptDate(apikey)  # 这里密钥的长度必须是16的倍数
        res = eg.encrypt(md5_sign)
        # print("AES加密:", res)
        # print(eg.decrypt(res))
        return res

    def create_msku(self, sku, country, supplier, num, msku):
        self.sql()
        msg = self.find_male(sku)
        if msg:
            time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
            list_msku = []
            for i in range(1, num + 1):
                msku_new = f'{msku}-{i}-{time_now[2:]}-T'
                list_msku.append(msku_new)
            # print(list_msku)
            path = self.write_excl(list_msku)
            msg, message = self.write_sql(list_msku, sku, country, supplier)
            if path and msg:
                self.connection.commit()
                self.sql_close()
                return True, [f'../MSKU生成/{path}.xlsx', path]
            else:
                self.sql_close()
                return False, False
        else:
            self.sql_close()
            return False, f"请先将{sku}关联本地品名与本地父体"

    def find_male(self, sku):
        sql1 = "select * from `amazon_form`.`male_parent` where `SKU` = '%s'" % sku
        self.cursor.execute(sql1)
        result1 = self.cursor.fetchall()
        sql2 = "select * from `amazon_form`.`male_sku` where `SKU` = '%s'" % sku
        self.cursor.execute(sql2)
        result2 = self.cursor.fetchall()
        if result1 and result2:
            return True
        else:
            return False

    def write_sql(self, list_msku, sku, country, supplier):
        try:
            for i in list_msku:
                sql1 = "select * from `amazon_form`.`pre_msku` where `MSKU` = '%s'" % i
                self.cursor.execute(sql1)
                result = self.cursor.fetchall()
                if result:
                   return False, f'{i}已重复'
                else:
                    sql = "insert into `amazon_form`.`pre_msku`(`ASIN`, `MSKU`, `SKU`, `国家`, `供应商`, `状态`)values" \
                          "('NULL', '%s', '%s', '%s', '%s', '未匹配')" % (i, sku, country, supplier)
                    self.cursor.execute(sql)
            return True, True
        except Exception as e:
            print(e)
            return False, e

    def write_excl(self, list_msku):
        try:
            time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
            wb = openpyxl.Workbook()
            wb_sheet = wb.active
            wb_sheet.append(['MSKU'])
            for i in list_msku:
                wb_sheet.append([i])
            wb.save(f'./static/MSKU生成/msku_{time_now}.xlsx')
            return f'msku_{time_now}'
        except Exception as e:
            print(e)
            return False

    def upload_excl(self, filename):
        wb = openpyxl.load_workbook(filename, data_only=True)
        wb_sheet = wb.active
        row1 = wb_sheet.max_row
        for i in range(row1, 0, -1):
            cell_value1 = wb_sheet.cell(row=i, column=1).value
            if cell_value1:
                row1 = i
                break
        list_msku = []
        self.sql()
        for i in range(2, row1 + 1):
            sku = wb_sheet.cell(row=i, column=1).value.strip()
            msg = self.find_male(sku)
            if msg:
                msku = wb_sheet.cell(row=i, column=2).value.strip()
                country = wb_sheet.cell(row=i, column=3).value.strip()
                supplier = wb_sheet.cell(row=i, column=4).value.strip()
                num = int(wb_sheet.cell(row=i, column=5).value)
                list_msku_index = []
                if sku and len(msku) < 23 and country and supplier and num:
                    time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
                    for j in range(1, num + 1):
                        msku_new = f'{msku}-{j}-{time_now[2:]}-T'
                        list_msku_index.append(msku_new)
                        list_msku.append(msku_new)
                    msg, message = self.write_sql(list_msku_index, sku, country, supplier)
                    if msg:
                        continue
                    else:
                        self.sql_close()
                        return False, False
                else:
                    self.sql_close()
                    return False, f"第{i}行数据不正确，请检查"
            else:
                self.sql_close()
                return False, f"请先将{sku}关联本地品名与本地父体"
        path = self.write_excl(list_msku)
        if path:
            self.connection.commit()
            self.sql_close()
            return True, [f'../MSKU生成/{path}.xlsx', path]
        else:
            self.sql_close()
            return False, "生成表格失败"

    def find_fnsku(self, sku, country):
        try:
            self.sql()
            sql = "select * from `amazon_form`.`pre_msku` where `SKU` = '%s' and `状态` = '未使用' and " \
                  "`国家` = '%s'" % (sku, country)
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            num = 0
            if result:
                num = len(result)
            self.sql_close()
            return True, f'{country}的{sku}有{num}个未使用的FNSKU'
        except Exception as e:
            return False, str(e)

    def get_fnsku(self, sku, country, num):
        try:
            self.sql()
            ws = openpyxl.Workbook()
            ws_sheet = ws.active
            ws_sheet.append(['*SKU', '品名', '原FNSKU', '调整FNSKU', '调整量'])
            sql = "select * from `amazon_form`.`pre_msku` where `SKU` = '%s' and `状态` = '未使用' and " \
                  "`国家` = '%s'" % (sku, country)
            # print(sql)
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            # print(result)
            if result and num <= len(result):
                k = 0
                product_name = self.get_product_name(sku)
                # print(product_name)
                # print(len(list_fnsku))
                for j in result:
                    k += 1
                    self.change_sql(j['FNSKU'])
                    ws_sheet.append([sku, product_name, '', j['FNSKU']])
                    if k == num:
                        break
                time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
                self.connection.commit()
                flag = self.check_sql(len(result), sku, country)
                self.sql_close()
                if flag:
                    ws.save(f'D:/html_windows/static/换标调整/换标调整-{sku}-{time_now}.xlsx')
                    return True, [f'../换标调整/换标调整-{sku}-{time_now}.xlsx', f"换标调整-{sku}-{time_now}"]
                else:
                    return False, 'FNSKU标记失败，请重试'
            else:
                self.sql_close()
                return False, f'{sku}的fnsku未使用个数不足'
        except Exception as e:
            self.sql_close()
            return False, str(e)

    def read_excl(self, filename):
        try:
            self.sql()
            wb = openpyxl.load_workbook(filename)
            wb_sheet = wb.active
            row1 = wb_sheet.max_row
            for i in range(row1, 0, -1):
                cell_value1 = wb_sheet.cell(row=i, column=1).value
                if cell_value1:
                    row1 = i
                    break
            ws = openpyxl.Workbook()
            ws_sheet = ws.active
            ws_sheet.append(['*SKU', '品名', '原FNSKU', '店铺名称', '调整FNSKU', '店铺名称', '调整量'])
            for i in range(2, row1 + 1):
                sku_excl = wb_sheet.cell(row=i, column=1).value.strip()
                num = int(wb_sheet.cell(row=i, column=2).value)
                country = wb_sheet.cell(row=i, column=3).value.strip()
                # num_order = self.get_inventory(sku)
                # if num <= num_order:
                sql = "select * from `amazon_form`.`pre_msku` where `SKU` = '%s' and `状态` = '未使用' and " \
                      "`国家` = '%s'" % (sku_excl, country)
                # print(sql)
                self.cursor.execute(sql)
                result = self.cursor.fetchall()
                # print(result)
                if result and num <= len(result):
                    k = 0
                    product_name = self.get_product_name(sku_excl)
                    for j in result:
                        k += 1
                        ws_sheet.append([sku_excl, product_name, '', '', j['FNSKU']])
                        if k == num:
                            break
                else:
                    return False, f'{sku_excl}的fnsku未使用个数不足'
            time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
            self.connection.commit()
            ws_row = ws_sheet.max_row
            for i in range(2, ws_row+1):
                fnsku = ws_sheet.cell(row=i, column=5).value
                self.change_sql(fnsku)
            self.sql_close()
            ws.save(f'D:/html_windows/static/换标调整/换标调整-{time_now}.xlsx')
            return True, [f'../换标调整/换标调整-{time_now}.xlsx', f"换标调整-{time_now}"]
        except Exception as e:
            self.sql_close()
            return False, str(e)

    def get_product_name(self, sku):
        sql = "select * from `data_read`.`product_id` where `SKU` = '%s'" % sku
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            return result[0]['品名']
        else:
            return False

    def change_sql(self, fnsku):
        sql = "update `amazon_form`.`pre_msku` set `状态` = '已使用' where `FNSKU` = '%s'" % fnsku
        # print(sql)
        self.cursor.execute(sql)
        self.connection.commit()

    def check_sql(self, length, sku, country):
        sql = "select * from `amazon_form`.`pre_msku` where `SKU` = '%s' and `状态` = '未使用' and " \
              "`国家` = '%s'" % (sku, country)
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        # print(len(result), length)
        if len(result) != length:
            return True
        else:
            return False

    def change_fnsku(self, filename, warehouse):
        try:
            dict_warehouse = {'工厂仓库': '2156', '横中路仓库-加拿大': '1489', '横中路仓库-日本': '1490',
                              '横中路仓库-美国': '1461', '横中路仓库-英国': '1488',
                              '横中路仓库-德国': '2382', '淘汰-横中路成品仓库-加拿大': '1476', '淘汰-横中路成品仓库-日本': '1477',
                              '淘汰-横中路成品仓库-美国': '1399', '淘汰-横中路成品仓库-英国': '1478', '百汇办公室': '414'}
            wb = openpyxl.load_workbook(filename)
            wb_sheet = wb.active
            row1 = wb_sheet.max_row
            for i in range(row1, 0, -1):
                cell_value1 = wb_sheet.cell(row=i, column=1).value
                if cell_value1:
                    row1 = i
                    break
            dict_product = {}
            dict_product['wid'] = dict_warehouse[warehouse]
            time_file = datetime.datetime.now().strftime("%Y%m%d")
            # self.sql()
            dict_product['product_list'] = []
            self.mkdir(f"fnsku_{time_file}")
            for i in range(2, row1 + 1):
                sku = wb_sheet.cell(row=i, column=1).value.strip()
                product_id = self.get_product_id1(sku)
                fnsku = wb_sheet.cell(row=i, column=3).value.strip()
                seller_name = wb_sheet.cell(row=i, column=4).value.strip()
                seller_id = self.get_seller_id1(seller_name)
                fnsku_new = wb_sheet.cell(row=i, column=5).value.strip()
                to_seller_name = wb_sheet.cell(row=i, column=6).value.strip()
                to_seller_id = self.get_seller_id1(to_seller_name)
                num = int(wb_sheet.cell(row=i, column=7).value)
                if seller_id and to_seller_id:
                    self.get_fnsku_lable(sku, fnsku_new, f'fnsku_{time_file}')
                    wb_sheet.cell(row=i, column=8).value = '换标成功'
                    dict_product['product_list'].append({'product_id': product_id, 'fnsku': fnsku, 'seller_id': seller_id, 'to_fnsku': fnsku_new, 'to_seller_id': to_seller_id, 'adjustment_valid_num': num})
                else:
                    return False, f"请检查{fnsku}和{fnsku_new}的店铺名称"
            msg, message = self.adjustment_order(dict_product)
            if msg:
                self.mkdir2(f'{time_file}-{message}')
                wb.save(f'D:/html_windows/static/换标调整/{time_file}-{message}/换标调整详情-{message}.xlsx')
                shutil.move(f"D:/FNSKU标签/fnsku_{time_file}/", rf"D:/html_windows/static/换标调整/{time_file}-{message}/")
                self.get_order_sn(message, time_file, dict_warehouse[warehouse])
                # path = f'D:/html_windows/static/换标调整/{time_file}-{message}'
                self.make_zip(f'D:/html_windows/static/换标调整/{time_file}-{message}', f'D:/html_windows/static/换标调整/{time_file}-{message}.zip')
                return msg, [f'../换标调整/{time_file}-{message}.zip', f'{time_file}-{message}']
            else:
                return msg, message
        except Exception as e:
            print(e)
            return False, str(e)

    def mkdir(self, file):
        folder = os.path.exists(f"D:/FNSKU标签/{file}")
        if not folder:
            os.makedirs(f"D:/FNSKU标签/{file}")

    def mkdir2(self, file):
        folder = os.path.exists(f"D:/html_windows/static/换标调整/{file}")
        if not folder:
            os.makedirs(f"D:/html_windows/static/换标调整/{file}")

    def make_zip(self, source_dir, output_filename):
        zipf = zipfile.ZipFile(output_filename, 'w')
        pre_len = len(os.path.dirname(source_dir))
        for parent, _, filenames in os.walk(source_dir):
            for filename in filenames:
                pathfile = os.path.join(parent, filename)
                arcname = pathfile[pre_len:].strip(os.path.sep)  # 相对路径
                zipf.write(pathfile, arcname)
        zipf.close ()

    def get_seller_id1(self, seller_name):
        self.sql()
        sql = "select * from `data_read`.`seller` where `店铺名` = '%s'" % seller_name
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            return result[0]['店铺ID']
        else:
            return False

    def get_seller_id2(self, seller_name):
        self.sql()
        sql = "select * from `data_read`.`seller` where `店铺名` = '%s'" % seller_name
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            return result[0]['亚马逊店铺ID'], result[0]['市场ID']
        else:
            return False, False

    def get_seller_country(self, seller_name):
        self.sql()
        sql = "select * from `data_read`.`seller` where `店铺名` = '%s'" % seller_name
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            return result[0]['国家']
        else:
            return False

    def get_product_id1(self, sku):
        self.sql()
        sql = "select * from `data_read`.`product_id` where `SKU` = '%s'" % sku
        # print(sql)
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            self.sql_close()
            return result[0]['ID']
        else:
            self.sql_close()
            return False

    def get_product_id2(self, sku):
        self.sql()
        sql = "select * from `data_read`.`product_id` where `SKU` = '%s'" % sku
        # print(sql)
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            return result[0]['品名'], result[0]['供应商']
        else:
            return False, False

    def adjustment_order(self, dict_product):
        time_stamp = int(time.time())
        url = "https://openapi.lingxing.com/erp/sc/routing/inventoryReceipt/StorageAdjustment/addRebrandAdjustmentOrder"
        body = {"access_token": self.access_token,
                "timestamp": time_stamp,
                "app_key": "ak_nEfE94OSogf3x",
                "wid": dict_product['wid'],
                "product_list": dict_product['product_list'],
                }
        res = self.get_sign(body)
        querystring = {"access_token": self.access_token,
                       "timestamp": time_stamp,
                       "app_key": self.app_id,
                       "sign": res
                       }
        payload = dict_product
        payload = json.dumps(payload)
        headers = {
            "Content-Type": "application/json",
            "Authorization": "bearer {{access_token}}"
        }
        response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
        result1 = json.loads(response.text)
        print(result1)
        if result1['code'] == 0 and result1['message'] == 'success':
            return True, result1['data']['order_sn']
        else:
            return False, result1['message']

    def get_order_sn(self, msg_id, time_now, wid):
        # order_sn = ''
        wb = openpyxl.Workbook()
        wb_sheet = wb.active
        wb_sheet.append(['单据编号', '仓库名称', '单据类型', '单据状态', '创建人', '创建时间', '操作人', '调整时间', 'SKU', '品名',
                         '店铺', 'FNSKU', '可用调整量', '调整数', '备注'])
        time_stamp = int(time.time())
        url = "https://openapi.lingxing.com/erp/sc/routing/inventoryReceipt/StorageAdjustment/getStorageAdjustOrderList"
        body = {"access_token": self.access_token,
                "timestamp": time_stamp,
                "app_key": "ak_nEfE94OSogf3x",
                "order_sn": msg_id,
                "wid": wid
                }
        res = self.get_sign (body)
        querystring = {"access_token": self.access_token,
                       "timestamp": time_stamp,
                       "app_key": self.app_id,
                       "sign": res
                       }
        payload = {"order_sn": msg_id, "wid": wid}
        payload = json.dumps(payload)
        headers = {
            "Content-Type": "application/json",
            "Authorization": "bearer {{access_token}}"
        }
        response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
        result1 = json.loads(response.text)
        # print(result1)
        if result1['code'] == 0 and result1['message'] == 'success':
            for i in result1['data']:
                if i['order_sn'] == msg_id:
                    order_sn = i['order_sn']
                    warehouse_name = i['ware_house_name']
                    type = i['type_text']
                    status = i['status_text']
                    realname = i['create_realname']
                    creat_time = i['create_time']
                    remark = i['remark']
                    for j in i['item_list']:
                        wb_sheet.append([order_sn, warehouse_name, type, status, realname, creat_time, None, None,
                                         j['sku'], j['product_name'], j['seller_name'], j['fnsku'],
                                         j['adjustment_valid_num'], -int(j['available_bin_list'][0]['product_num']), remark])
                        wb_sheet.append([order_sn, warehouse_name, type, status, realname, creat_time, None, None,
                                         j['sku'], j['product_name'], j['to_seller_name'], j['to_fnsku'],
                                         None, j['to_available_bin_list'][0]['product_num'], remark])
        wb.save(f'D:/html_windows/static/换标调整/{time_now}-{msg_id}/换标调整单-{msg_id}.xlsx')

    def login_asinking3(self, username, password):
        # 生成Session对象，用于保存Cookie
        global_var.s = requests.Session()
        # 登录url
        login_url = 'https://erp.lingxing.com/api/passport/login'
        # 请求头
        headers = {'Host': 'erp.lingxing.com',
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                   , 'Referer': 'https://erp.lingxing.com/login',
                   'Accept': 'application/json, text/plain, */*',
                   'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                   'Accept-Encoding': 'gzip, deflate, br',
                   'Content-Type': 'application/json;charset=utf-8',
                   'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                   'X-AK-Company-Id': '90136229927150080',
                   'X-AK-Request-Source': 'erp',
                   'X-AK-ENV-KEY': 'SAAS-10',
                   'X-AK-Version': '1.0.0.0.0.023',
                   'X-AK-Zid': '109810',
                   'Content-Length': '114',
                   'Origin': 'https://erp.lingxing.com',
                   'Connection': 'keep-alive'}
        # 传递用户名和密码
        data = {'account': username, 'pwd': password}
        data = json.dumps(data)
        try:
            r = global_var.s.post(login_url, headers=headers, data=data)
            # print(global_var.s.cookies.get('auth-token'))
            r.raise_for_status()
            r1 = r.text
            r2 = json.loads(r1)
            r3 = r2['msg']
        except:
            r3 = '网络错误失败'

        return r3

    def get_fnsku_lable(self, sku, fnsku, time_now):
        self.login_asinking3('IT-Test', 'IT-Test')
        get_url = f"https://erp.lingxing.com/api/product/showOnline?start_date=2021-10-19&end_date=2021-10-19&" \
                  f"search_field=fnsku&search_value={fnsku}&status=&" \
                  "is_pair=&fulfillment_channel_type=&offset=0&length=50&req_time_sequence=/api/product/showOnline$$27"
        headers = {'Host': 'erp.lingxing.com',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0',
            # , 'Referer': 'https://erp.lingxing.com/erp/warehouse_detail',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
            'Accept-Encoding': 'gzip, deflate, br',
            # 'Content-Type': 'application/json;charset=utf-8',
            'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
            'X-AK-Company-Id': '90136229927150080',
            'X-AK-Request-Source': 'erp',
            'X-AK-ENV-KEY': 'SAAS-10',
            'X-AK-Version': '1.0.0.0.0.023',
            'X-AK-Zid': '109810',
            # 'Content-Length': '114',
            # 'Origin': 'https://erp.lingxing.com',
            'Connection': 'keep-alive'}
        please_get = global_var.s.get(get_url, headers=headers)
        please_get = json.loads(please_get.text)
        # print(please_get)
        please_post = {}
        please_post['data'] = []
        please_post['data'].append(0)
        please_post['data'][0] = {}
        if please_get['msg'] and please_get['list']:
            for i in please_get['list']:
                if i['fnsku'] == fnsku and i['local_sku'] == sku:
                    data = {}
                    data['fnsku'] = fnsku
                    data['item_condition'] = i['item_condition']
                    data['item_name'] = i['item_name']
                    data['local_name'] = i['local_name']
                    data['local_sku'] = i['local_sku']
                    data['num'] = "1"
                    please_post['data'][0] = data
        please_post['is_content'] = 0
        please_post['is_self'] = 1
        please_post['name_type'] = 2
        please_post['page_type'] = 1
        please_post['req_time_sequence'] = "/api/print/printProduct$$8"
        please_post['self_content'] = "MADE IN CHINA"
        # print(please_post)
        please_post = json.dumps(please_post)
        print_url = "https://erp.lingxing.com/api/print/printProduct"
        post_headers = {'Host': 'erp.lingxing.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0',
                        'Referer': 'https://erp.lingxing.com/login',
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '114',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
        print_data = global_var.s.post(print_url, headers=post_headers, data=please_post)
        print_data = json.loads(print_data.text)
        # print(print_data)
        if print_data['msg'] == "操作成功":
            url2 = f"https://erp.lingxing.com/api/file/downloadById?id={int(print_data['file_id'])}&is_export=1"
            # 请求头
            headers = {'Host': 'erp.lingxing.com',
                       'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0',
                       # , 'Referer': 'https://erp.lingxing.com/erp/warehouse_detail',
                       'Accept': 'application/json, text/plain, */*',
                       'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                       'Accept-Encoding': 'gzip, deflate, br',
                       # 'Content-Type': 'application/json;charset=utf-8',
                       'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                       'X-AK-Company-Id': '90136229927150080',
                       'X-AK-Request-Source': 'erp',
                       'X-AK-ENV-KEY': 'SAAS-10',
                       'X-AK-Version': '1.0.0.0.0.023',
                       'X-AK-Zid': '109810',
                       # 'Content-Length': '114',
                       # 'Origin': 'https://erp.lingxing.com',
                       'Connection': 'keep-alive'}
            res2 = global_var.s.get(url2, headers=headers, stream=False)
            # print(res2)
            with open('D:/FNSKU标签/%s/%s.pdf' % (time_now, fnsku), 'wb') as wr:
                wr.write(res2.content)

    def pair_msku(self, sku, asin, store, version, msku, fnsku, male_parent=None, male_sku=None):
        index, index_data = self.get_male(sku, male_parent, male_sku)
        if index:
            country = self.get_seller_country (store)
            key = country + fnsku
            self.sql ()
            sql = "select * from `data_read`.`listing` where `key` = '%s'" % key
            self.cursor.execute (sql)
            result = self.cursor.fetchall ()
            if result:
                self.sql_close ()
                return False, f'{msku}已配对，请勿重复配对。'
            else:
                # print(store)
                seller_id, marketplace_id = self.get_seller_id2(store)
                data = {'data': [{'msku': msku, 'sku': sku, 'is_sync_pic': 0, 'seller_id': seller_id, 'marketplace_id': marketplace_id}]}
                # print(data)
                msg, message = self.pair(data)
                # print(message)
                if msg:
                    country = self.get_seller_country(store)
                    product_name, supplier = self.get_product_id2(sku)
                    if country and product_name:
                        key = country + fnsku
                        self.sql()
                        sql = "select * from `data_read`.`listing` where `key` = '%s'" % key
                        self.cursor.execute(sql)
                        result = self.cursor.fetchall()
                        if result:
                            self.sql_close()
                            return False, f'{msku}已配对，请勿重复配对。'
                        sql1 = f"INSERT INTO data_read.listing VALUES('{key}','{store}','{country}'," \
                              f"'{asin}','{msku}','{fnsku}','{product_name}','{sku}','可采购'," \
                              f"'可发货', '{version}',85,'60','{supplier}',DEFAULT,DEFAULT)"
                        self.cursor.execute(sql1)
                        self.connection.commit()
                        self.sql_close()
                        return True, '配对成功'
                    else:
                        print('sku: ', sku, product_name, supplier)
                        return False, '数据库错误'
                else:
                    return msg, message
        else:
            return index, index_data

    def batch_pair(self, path):
        wb = openpyxl.load_workbook(path)
        wb_sheet = wb.active
        row1 = wb_sheet.max_row
        for i in range(row1, 0, -1):
            cell_value1 = wb_sheet.cell(row=i, column=1).value
            if cell_value1:
                row1 = i
                break
        for i in range(2, row1+1):
            fnsku = wb_sheet.cell(row=i, column=1).value
            msku = wb_sheet.cell(row=i, column=2).value
            asin = wb_sheet.cell(row=i, column=3).value
            store = wb_sheet.cell(row=i, column=4).value
            version = wb_sheet.cell(row=i, column=5).value
            sku = wb_sheet.cell(row=i, column=6).value
            male_parent = wb_sheet.cell(row=i, column=7).value
            male_sku = wb_sheet.cell(row=i, column=8).value
            index, index_data = self.get_male(sku, male_parent, male_sku)
            if index:
                country = self.get_seller_country(store)
                key = country + fnsku
                self.sql()
                sql = "select * from `data_read`.`listing` where `key` = '%s'" % key
                self.cursor.execute(sql)
                result = self.cursor.fetchall()
                if result:
                    self.sql_close()
                    wb_sheet.cell(row=i, column=9).value = f'{msku}已配对，请勿重复配对。'
                else:
                    seller_id, marketplace_id = self.get_seller_id2(store)
                    data = {'data': [{'msku': msku, 'sku': sku, 'is_sync_pic': 0, 'seller_id': seller_id, 'marketplace_id': marketplace_id}]}
                    msg, message = self.pair(data)
                    if msg:
                        product_name, supplier = self.get_product_id2(sku)
                        if country and product_name and supplier:
                            sql1 = f"INSERT INTO data_read.listing VALUES('{key}','{store}','{country}'," \
                                   f"'{asin}','{msku}','{fnsku}','{product_name}','{sku}','可采购'," \
                                   f"'可发货', '{version}',85,'60','{supplier}',DEFAULT,DEFAULT)"
                            self.cursor.execute(sql1)
                            self.connection.commit()
                            self.sql_close()
                            wb_sheet.cell(row=i, column=9).value = '配对成功'
                        else:
                            wb_sheet.cell(row=i, column=9).value = f"错误：sku: , {sku}, {product_name}, {supplier}"
                    else:
                        wb_sheet.cell(row=i, column=9).value = f"{message}"
            else:
                wb_sheet.cell(row=i, column=9).value = f"{index_data}"
        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
        wb.save(f'./static/批量配对/批量配对详情_{time_now}.xlsx')
        return f"../批量配对/批量配对详情_{time_now}.xlsx", f"批量配对详情_{time_now}"

    def pair(self, data):
        payload = copy.deepcopy(data)
        for i in range(0, len(data['data'])):
            data['data'][i]['msku'] = data['data'][i]['msku'].replace(" ", "?")
        time_stamp = int(time.time())
        url = "https://openapi.lingxing.com/erp/sc/storage/product/link"
        body = {"access_token": self.access_token,
                "timestamp": time_stamp,
                "app_key": "ak_nEfE94OSogf3x",
                "data": data['data'],
                }
        # print(body)
        res = self.get_sign(body)
        querystring = {"access_token": self.access_token,
                       "timestamp": time_stamp,
                       "app_key": self.app_id,
                       "sign": res
                       }
        # print(payload)
        payload = json.dumps(payload)
        headers = {
            "Content-Type": "application/json",
            "Authorization": "bearer {{access_token}}"
        }
        response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
        result1 = json.loads(response.text)
        # print(result1)
        if result1['code'] == 0 and result1['message'] == 'success':
            return True, True
        else:
            print(result1)
            return False, result1['error_details'][0]['message']

    def get_male(self, sku, male_parent, male_sku):
        self.sql()
        sql1 = "select * from `amazon_form`.`male_parent` where `SKU` = '%s'" % sku
        self.cursor.execute(sql1)
        result1 = self.cursor.fetchall()
        sql2 = "select * from `amazon_form`.`male_sku` where `SKU` = '%s'" % sku
        self.cursor.execute(sql2)
        result2 = self.cursor.fetchall()
        # print(len(result1), len(result2))
        if result1 and result2:
            return True, True
        else:
            msg1, message1 = self.relevance_male(male_parent, sku)
            msg2, message2 = self.relevance_sku(male_sku, sku)
            if msg1 and msg2:
                return True, True
            else:
                if not msg1:
                    return False, message1
                if not msg2:
                    return False, message2

    def relevance_male(self, male_parent, sku):
        self.sql()
        sql = "select * from `amazon_form`.`list_parent` where `本地父体` = '%s'" % male_parent
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            sql2 = "select * from `amazon_form`.`male_parent` where `本地父体` = '%s' and `SKU` = '%s'" % (male_parent, sku)
            self.cursor.execute(sql2)
            result2 = self.cursor.fetchall()
            if result2:
                return True, True
            else:
                sql1 = "insert into `amazon_form`.`male_parent`(`本地父体`, `SKU`)values('%s', '%s')" % (male_parent, sku)
                try:
                    self.cursor.execute(sql1)
                    self.connection.commit()
                    return True, True
                except Exception as e:
                    self.connection.rollback()
                    return False, e
        else:
            self.sql_close()
            return False, f"{sku}没有{male_parent}本地父体，请检查"

    def relevance_sku(self, local_sku, sku):
        self.sql()
        sql = "select * from `amazon_form`.`list_local_sku` where `本地品名` = '%s'" % local_sku
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            sql2 = "select * from `amazon_form`.`male_sku` where `本地品名` = '%s' and `SKU` = '%s'" % (local_sku, sku)
            self.cursor.execute(sql2)
            result2 = self.cursor.fetchall()
            if result2:
                return True, True
            else:
                sql1 = "insert into `amazon_form`.`male_sku`(`本地品名`, `SKU`)values('%s', '%s')" % (local_sku, sku)
                try:
                    self.cursor.execute(sql1)
                    self.connection.commit()
                    self.sql_close()
                    return True, True
                except Exception as e:
                    self.connection.rollback()
                    self.sql_close()
                    return False, e
        else:
            self.sql_close()
            return False, f"没有{local_sku}这个本地品名，请检查"


class Find_order():
    def __init__(self):
        self.s = requests.Session()
        login_url = 'https://erp.lingxing.com/api/passport/login'
        # 请求头
        headers = {'Host': 'erp.lingxing.com',
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                    , 'Referer': 'https://erp.lingxing.com/login',
                   'Accept': 'application/json, text/plain, */*',
                   'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                   'Accept-Encoding': 'gzip, deflate, br',
                   'Content-Type': 'application/json;charset=utf-8',
                   'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                   'X-AK-Company-Id': '90136229927150080',
                   'X-AK-Request-Source': 'erp',
                   'X-AK-ENV-KEY': 'SAAS-10',
                   'X-AK-Version': '1.0.0.0.0.023',
                   'X-AK-Zid': '109810',
                   'Content-Length': '114',
                   'Origin': 'https://erp.lingxing.com',
                   'Connection': 'keep-alive'}
        # 传递用户名和密码
        data = {'account': 'IT-Test', 'pwd': 'IT-Test'}
        data = json.dumps(data)
        self.s.post(login_url, headers=headers, data=data)
        self.auth_token = None
        self.cty_dict = {'美国': 1, '加拿大': 2, '墨西哥': 3, '英国': 4, '意大利': 7, '德国': 5, '法国': 6, '西班牙': 8,
                         '印度': 9, '日本': 10, '澳洲': 11, '阿联酋': 12, '新加坡': 13, '荷兰': 14, '沙特阿拉伯': 15,
                         '巴西': 16, '瑞典': 17, '土耳其': 20, '波兰': 19}
        # self.find_sql()

    def sql(self):
        self.connection = pymysql.connect(host='3354n8l084.goho.co',  # 数据库地址
                                          port=24824,
                                          user='test_user',  # 数据库用户名
                                          password='a123456',  # 数据库密码
                                          db='storage',  # 数据库名称
                                          charset='utf8',
                                          cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor()

    def sql_close(self):
        self.cursor.close()
        self.connection.close()

    def get_msku(self):
        try:
            self.sql()
            sql = "update `flag`.`amazon_form_flag` set `flag_num` = 1 where `flag_name` = 'pre_msku'"
            self.cursor.execute(sql)
            self.connection.commit()
            self.delete_sql()
            sql = "select * from `amazon_form`.`pre_msku` where `状态` = '未匹配'"
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            self.sql_close()
            if result:
                auth_token = self.s.cookies.get('auth-token')
                auth_token = auth_token.replace('%25', '%')
                auth_token = auth_token.replace('%23', '#')
                auth_token = auth_token.replace('%26', '&')
                auth_token = auth_token.replace('%2B', '+')
                auth_token = auth_token.replace('%28', '(')
                auth_token = auth_token.replace('%29', ')')
                auth_token = auth_token.replace('%2F', '/')
                auth_token = auth_token.replace('%3D', '=')
                auth_token = auth_token.replace('%3F', '?')
                self.auth_token = auth_token
                self.dict_msku = {}
                    # print(auth_token)
                for i in result:
                    msku = i['MSKU'].strip()
                    sku = i['SKU'].strip()
                    # asin = i['ASIN']
                    country = i['国家'].strip()
                    supplier = i['供应商'].strip()
                    self.dict_msku[msku] = [sku, country, supplier]
                # print(self.dict_msku)
                index, list_asin = self.downloads(auth_token)
                # print(self.dict_msku)
                # list_asin = self.read_excl('517288065326436352')
                # print(self.dict_msku)
                if list_asin:
                    self.sql()
                    for i in self.dict_msku:
                        if len(self.dict_msku[i]) > 3:
                            if list_asin.count(self.dict_msku[i][5]) > 1:
                                sql = "update `amazon_form`.`pre_msku` set `FNSKU` = '%s' , `状态` = '存在相同ASIN', `ASIN` = '%s' " \
                                      "where `MSKU` = '%s'" % (self.dict_msku[i][4], self.dict_msku[i][5], i)
                                # print(111)
                            else:
                                # print(self.dict_msku[i])
                                quantity = Quantity()
                                msg, message = quantity.pair_msku(self.dict_msku[i][0], self.dict_msku[i][5], self.dict_msku[i][3], 1.0, i, self.dict_msku[i][4])
                                # print(message)
                                if msg:
                                    sql = "update `amazon_form`.`pre_msku` set `FNSKU` = '%s' , `状态` = '未使用', `ASIN` = '%s' " \
                                          "where `MSKU` = '%s'" % (self.dict_msku[i][4], self.dict_msku[i][5], i)
                                    list_asin.append(i[0])
                            self.cursor.execute(sql)
                            self.connection.commit()
                    self.sql_close()
                self.sql()
                sql = "update `flag`.`amazon_form_flag` set `flag_num` = 0 where `flag_name` = 'pre_msku'"
                self.cursor.execute(sql)
                self.connection.commit()
                self.sql_close()
            else:
                self.sql()
                sql = "update `flag`.`amazon_form_flag` set `flag_num` = 0 where `flag_name` = 'pre_msku'"
                self.cursor.execute(sql)
                self.connection.commit()
                self.sql_close()
        except Exception as e:
            print(e)
            self.sql()
            sql = "update `flag`.`amazon_form_flag` set `flag_num` = 0 where `flag_name` = 'pre_msku'"
            self.cursor.execute(sql)
            self.connection.commit()
            self.sql_close()
            return False

    def get_productname(self, sku):
        url = f"https://erp.lingxing.com/api/product/lists"
        get_headers = {'Host': 'erp.lingxing.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                        , 'Referer': 'https://erp.lingxing.com/erp/productManage',
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '909',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
        data = {"search_field_time": "create_time", "sort_field": "create_time", "sort_type": "desc",
                f"search_field": "sku", "search_value": f"{sku}", "attribute": "", "status": [], "gtag_ids": "",
                "senior_search_list": "[]", "is_matched_alibaba": "", "is_matched_listing": "", "relation_aux": "",
                "cg_package": "", "cg_product_gross_weight": {"left": "", "right": "", "symbol": "gt"},
                "cg_transport_costs": {"left": "", "right": "", "symbol": "gt", "country_code": "US"},
                "cg_price": {"left": "", "right": "", "symbol": "gt"}, "offset": 0, "is_combo": "", "length": 200,
                "is_aux": 0, "product_type": [1, 2], "selected_product_ids": "", "req_time_sequence": "/api/product/lists$$"}
        data = json.dumps(data)
        get_msg = self.s.post(url, headers=get_headers, data=data)
        get_msg = json.loads(get_msg.text)
        product_name = ''
        for i in get_msg['list']:
            if i['sku'] == sku:
                product_name = i['product_name']
        return product_name

    def downloads(self, auth_token):
        try:
            url = "https://gw.lingxingerp.com/listing-api/api/product/exportOnline"
            headers = {'Host': 'gw.lingxingerp.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                        , 'Referer': 'https://erp.lingxing.com',
                       'auth-token': auth_token,
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '909',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
            data = {"offset": 0, "length": 50, "search_field": "local_sku", "exact_search": 1, "sids": "", "status": "",
                    "is_pair": "", "fulfillment_channel_type": "", "global_tag_ids": "",
                    "req_time_sequence": "/listing-api/api/product/exportOnline$$2"}
            data = json.dumps(data)
            post_msg = self.s.post(url, headers=headers, data=data)
            post_msg = json.loads(post_msg.text)
            print(post_msg)
            if post_msg['code'] == 1 and post_msg['msg'] == "成功":
                report_id = post_msg['data']['data']['report_id']
                time.sleep(300)
                if report_id:
                    file_download_url = f"https://erp.lingxing.com/api/download/downloadCenterReport/downloadResource?report_id={report_id}"
                    # print(file_download_url)
                    get_headers = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/erp/muser/downloadCenter'}
                    download_file = self.s.get(file_download_url, headers=get_headers, stream=False)
                    with open('D:/listing/listing.zip', 'wb') as q:
                        q.write(download_file.content)
                    list_asin = self.read_excl(report_id)
                    if list_asin:
                        return True, list_asin
                    else:
                        return False, False
                else:
                    return False, False
            else:
                return False, False
        except Exception as e:
            print(e)
            return False, e

    def read_excl(self, report):
        file = zipfile.ZipFile('D:/listing/listing.zip')
        file.extractall('D:/listing/')
        file.close()
        time_data = datetime.datetime.now().strftime("%Y%m%d")
        filename = f"D:/listing/listing{time_data}-{report}.xlsx"
        wb = openpyxl.load_workbook(filename)
        wb_sheet = wb.active
        row_max = wb_sheet.max_row
        list_asin = []
        for i in range(2, row_max+1):
            asin = wb_sheet.cell(row=i, column=5).value
            msku = wb_sheet.cell(row=i, column=7).value
            country = wb_sheet.cell(row=i, column=2).value
            if asin:
                if asin in list_asin or len(asin) != 10 or asin.find('B0') < 0:
                    continue
                else:
                    list_asin.append(asin)
            if msku in self.dict_msku and country == self.dict_msku[msku][1]:
                shop_name = wb_sheet.cell(row=i, column=1).value
                fnsku = wb_sheet.cell(row=i, column=8).value
                self.dict_msku[msku].append(shop_name)
                self.dict_msku[msku].append(fnsku)
                self.dict_msku[msku].append(asin)
        if len(list_asin) > 5000 and 'B0BLCS9J5G' in list_asin:
            return list_asin
        else:
            return False

    def find_fnsku(self, msku, country, auth_token):
        post_url = "https://gw.lingxingerp.com/listing-api/api/product/showOnline"
        post_headers = {'Host': 'gw.lingxingerp.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                        , 'Referer': 'https://erp.lingxing.com',
                        'auth-token': auth_token,
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '909',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
        data = {}
        data['fulfillment_channel_type'] = ''
        data['is_pair'] = ''
        data['length'] = 200
        data['offset'] = 0
        data['req_time_sequence'] = '/listing-api/api/product/showOnline$$'
        data['search_field'] = 'msku'
        data['search_value'] = []
        data['search_value'].append(msku)
        data['sids'] = ''
        data['status'] = ''
        # print(data)
        data = json.dumps(data)
        post_msg = self.s.post(post_url, headers=post_headers, data=data)
        post_msg = json.loads(post_msg.text)
        # print(post_msg)
        # print(msku)
        if post_msg['code'] == 1 and post_msg['msg'] == '成功' and post_msg['data']['list']:
            for i in post_msg['data']['list']:
                # msku_ = i['msku'].strip()
                if i['msku'] == msku and country == i['marketplace']:
                    # print(i['fnsku'])
                    return i['fnsku'], i['asin'], i['seller_name']
                else:
                    print(i['msku'])
            return False, False, False
        return False, False, False

    def peidui(self, asin, country, fnsku, supplier, sku, version, shop_name, msku, product_name):
        key = country + fnsku
        sql3 = f"INSERT INTO data_read.listing VALUES('{key}','{shop_name}','{country}'," \
               f"'{asin}','{msku}','{fnsku}','{product_name}','{sku}','可采购'," \
               f"'可发货', '{version}',85,'60','{supplier}',DEFAULT,DEFAULT)"
        try:
            self.cursor.execute(sql3)
            file_url2 = f"https://erp.lingxing.com/api/product/lists"
            # 请求头
            headers2 = {'Host': 'erp.lingxing.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                        , 'Referer': 'https://erp.lingxing.com/erp/productManage',
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '909',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
            data_1 = {"search_field_time": "create_time", "sort_field": "create_time", "sort_type": "desc",
                      f"search_field": "sku", "search_value": f"{sku}", "attribute": "", "status": [], "gtag_ids": "",
                      "senior_search_list": "[]", "is_matched_alibaba": "", "is_matched_listing": "", "relation_aux": "",
                      "cg_package": "", "cg_product_gross_weight": {"left": "", "right": "", "symbol": "gt"},
                      "cg_transport_costs": {"left": "", "right": "", "symbol": "gt", "country_code": "US"},
                      "cg_price": {"left": "", "right": "", "symbol": "gt"}, "offset": 0, "is_combo": "", "length": 200,
                      "is_aux": 0, "product_type": [1,2], "selected_product_ids": "", "req_time_sequence": "/api/product/lists$$"}
            data_1 = json.dumps(data_1)
            res2 = self.s.post(file_url2, headers=headers2, data=data_1)
            res2 = res2.text
            res2 = json.loads(res2)
            print(res2)
            id = None
            for i in res2['list']:
                if sku == i['sku']:
                    id = i['id']
            url3 = "https://gw.lingxingerp.com/listing-api/api/product/batchLink"
            url4 = "https://gw.lingxingerp.com/listing-api/api/product/showOnline"
            # print(auth_token)
            headers3 = {
                        'Host': 'gw.lingxingerp.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:101.0) Gecko/20100101 Firefox/101.0'
                        , 'Referer': 'https://erp.lingxing.com/',
                        'Accept': 'application/json, text/plain, */*',
                        'auth-token': self.auth_token,
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '2.9.5.0.1.010',
                        'X-AK-Zid': '109810',
                        'Content-Length': '164',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
            data3 = {"sid_msku": [], "pid": id, "is_sync_pic": 0,
                     "req_time_sequence": "/listing-api/api/product/batchLink$$"}
            data4 = {}
            data4['fulfillment_channel_type'] = ''
            data4['is_pair'] = ''
            data4['req_time_sequence'] = '/listing-api/api/product/showOnline$$'
            data4['search_field'] = 'asin'
            data4['search_value'] = [asin]
            data4['status'] = ''
            data4 = json.dumps(data4)
            res4 = self.s.post(url4, headers=headers3, data=data4)
            res4 = json.loads(res4.text)
            # print(res4)
            if res4['code'] != 1:
                raise Exception('配对失败。')
            else:
                for i in res4['data']['list']:
                    if i['fnsku'] == fnsku:
                        data3['sid_msku'].append({'msku': i['msku'], 'store_id': i['store_id']})
            data3 = json.dumps(data3)
            res3 = self.s.post(url3, headers=headers3, data=data3)
            a3 = res3.text
            b3 = json.loads(a3)
            print(b3)
            if b3['code'] != 1:
                raise Exception('ERP中没配对成功。')
            self.connection.commit()
        except Exception as e:
            self.connection.rollback()
            print(e)
            print("配对失败！！！")
            return False, "配对失败！！！"
        return 2, b3['msg']

    def max_num(self, a, b):
        if float(a) > float(b):
            return float(a)
        else:
            return float(b)

    def delete_sql(self):
        time_now = datetime.datetime.now().strftime("%Y-%m-%d")
        time_now = datetime.datetime.strptime(time_now, '%Y-%m-%d')
        sql = "select * from `amazon_form`.`pre_msku` where `状态` = '未匹配'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        # print(result)
        for i in result:
            if i['ASIN'] and i['FNSKU']:
                continue
            else:
                time_data = str(i['创建时间'])
                time_data = time_data[0:10]
                # print(time_data)
                time_data = datetime.datetime.strptime(time_data, '%Y-%m-%d')
                # time_data = datetime.datetime.strptime(time_data, '%Y-%m-%d-%H-%M-%S')
                # time_data = time_data[0:10]
                # print(time_data, time_now)
                data_day = time_now - time_data
                # print(data_day.days)
                if data_day.days >= 7:
                    msku = i['MSKU']
                    sql1 = "update `amazon_form`.`pre_msku` set `状态` = '已删除'" \
                           "where `MSKU` = '%s'" % msku
                    self.cursor.execute(sql1)
        self.connection.commit()

    def test(self):
        time.sleep(10)
        print(111)

    def grt_flag(self):
        self.sql()
        sql = "select * from `flag`.`amazon_form_flag` where `flag_name` = 'pre_msku'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        return result[0]['flag_num']


if __name__ == '__main__':
    quantity = Find_order()
    quantity.get_msku()