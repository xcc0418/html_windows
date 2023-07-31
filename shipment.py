import os
import global_var
import shutil
import requests
import json
import time
import hashlib
import urllib
import base64
import datetime
import pymysql
from Crypto.Cipher import AES
import openpyxl
from shutil import copy
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


def func(sql, m='r'):
    py = pymysql.connect(host='3354n8l084.goho.co', user='test_user', password='a123456', port=24824, db='storage')
    cursor = py.cursor()
    data = None
    # print(sql)
    try:
        cursor.execute(sql)
        if m == 'r':
            data = cursor.fetchall()
        elif m == 'w':
            py.commit()
            data = cursor.rowcount
    except:
        data = False
        py.rollback()
    py.close()
    return data


def get_order_msg(index, order):
    try:
        if index == "SKU":
            if order.find('GC') >= 0:
                quantity_msg = Quantity()
                product_id = quantity_msg.get_warehouse_num(order)
                product_name = quantity_msg.get_product(product_id, order)
                return {'FNSKU': order, 'name': product_name}
            else:
                name = func("SELECT `品名` FROM `data_read`.`listing` where `SKU`='%s'" % order)
                if name[0]:
                    return {'FNSKU': order, 'name': name[0][0]}
                else:
                    return {'FNSKU': order, 'name': '-'}
        if index == "FNSKU":
            name = func("SELECT `品名` FROM `data_read`.`listing` where `FNSKU`='%s'" % order)
            if name[0]:
                return {'FNSKU': order, 'name': name[0][0]}
            else:
                return {'FNSKU': order, 'name': '-'}
        if index == "次品出库":
            if len(order) == 10:
                name = func("SELECT `品名` FROM `data_read`.`listing` where `SKU`='%s'" % order)
                return {'SKU': order, 'name': name[0][0]}
            else:
                name = func("SELECT `品名` FROM `data_read`.`listing` where `FNSKU`='%s'" % order)
                if name[0]:
                    return {'FNSKU': order, 'name': name[0][0]}
                else:
                    return {'FNSKU': order, 'name': '-'}
        if index == "计划单号":
            sql = ''
        if index == "PDD/ZX":
            if order.find('XNWL') > 0:
                sku = func(f"select `SKU` from `storage`.`sku_map` where `条形码` = '{order}'")[0]
                return {'SKU': sku[0], 'name': sku[1]}
            else:
                name = get_productname(order)
                return {'SKU': order, 'name': name}
    except Exception as e:
        print(e)
        return False


def get_productname(sku):
    url = f"https://erp.lingxing.com/api/product/lists?search_field_time=create_time&sort_field=create_time&" \
          f"sort_type=desc&search_field=sku&search_value={sku}&attribute=&status=&is_matched_alibaba=&" \
          f"senior_search_list=[]&offset=0&is_combo=&length=500&is_aux=0&product_type[]=1&product_type[]=2&" \
          f"selected_product_ids=&req_time_sequence=%2Fapi%2Fproduct%2Flists$$"
    get_headers = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/erp/productManage'}
    get_msg = global_var.s.get(url, headers=get_headers)
    get_msg = json.loads(get_msg.text)
    productname = ''
    for i in get_msg['list']:
        if i['sku'] == sku:
            productname = i['product_name']
    return productname


def uploading_pdd(list_sku, list_num):
    dict_product = []
    complete_flag = 0
    please_url = "https://erp.lingxing.com/api/module/inventoryReceipt/Outbound/completeOrder"
    find_headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
        , 'Referer': 'https://erp.lingxing.com/erp/outbound_manual', 'Accept': 'application/json, text/plain, */*',
                    'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                    'Accept-Encoding': 'gzip, deflate, br',
                    'Content-Type': 'application/json;charset=utf-8',
                    'X-AK-Request-Id': '03cabda1-739d-4cd7-bd94-6650f97775a8',
                    'X-AK-Company-Id': '90136229927150080',
                    'X-AK-Request-Source': 'erp',
                    'X-AK-ENV-KEY': 'SAAS-10',
                    'X-AK-Version': '2.8.5.1.2.033',
                    'X-AK-Zid': '109810',
                    'Content-Length': '481',
                    'Origin': 'https://erp.lingxing.com',
                    'Connection': 'keep-alive'}
    get_headers = {'Host': 'erp.lingxing.com',
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
        , 'Referer': 'https://erp.lingxing.com/erp/msupply/warehouseDetail',
                   'Accept': 'application/json, text/plain, */*',
                   'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                   'Accept-Encoding': 'gzip, deflate, br',
                   'Content-Type': 'application/json;charset=utf-8',
                   'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                   'X-AK-Company-Id': '90136229927150080',
                   'X-AK-Request-Source': 'erp',
                   'X-AK-ENV-KEY': 'SAAS-10',
                   'X-AK-Version': '1.0.0.0.0.023',
                   'X-AK-Zid': '109810',
                   'Content-Length': '909',
                   'Origin': 'https://erp.lingxing.com',
                   'Connection': 'keep-alive'}
    for j in range(0, len (list_sku)):
        # dict_product.append(j)
        # dict_product[j] = {}
        # print(dict_product)
        sku = list_sku[j]
        get_url = "https://erp.lingxing.com/api/storage/lists"
        data = {"wid_list": "", "mid_list": "", "sid_list": "", "cid_list": "", "bid_list": "",
                "principal_list": "",
                "product_type_list": "", "product_attribute": "", "product_status": "", "search_field": "sku",
                "search_value": f"{sku}", "is_sku_merge_show": 0, "is_hide_zero_stock": 0, "offset": 0,
                "length": 200, "sort_field": "pre_total", "sort_type": "desc", "gtag_ids": "",
                "senior_search_list": "[]", "req_time_sequence": "/api/storage/lists$$"}
        data = json.dumps (data)
        get_msg = global_var.s.post (get_url, headers=get_headers, data=data)
        get_msg = json.loads (get_msg.text)
        # print(text)
        list_text = []
        list_list = get_msg['data']['list']
        # print(get_msg['data']['list'])
        for i in range(0, len(list_list)):
            if list_list[i]['wh_name'] == '横中路仓库-美国' and list_list[i]['sku'] == sku:
                list_text.append(list_list[i])
        # print(list_text)
        if list_text:
            sku_name = list_text[0]['product_name']
            sku_name = sku_name.replace('/', '%2F')
            sku_name = sku_name.replace('-', '%2D')
            # sku_name = sku_name.replace('%', '%25')
            sku_name = sku_name.replace('#', '%23')
            sku_name = sku_name.replace('&', '%26')
            sku_name = sku_name.replace('+', '%2B')
            sku_name = sku_name.replace('(', '%28')
            sku_name = sku_name.replace(')', '%29')
            if list_text[0]['total'] >= int(list_num[j]):
                dict1 = {}
                dict1['bad_num'] = 0
                dict1['fnsku'] = ""
                dict1['good_num'] = str(list_num[j])
                dict1['item_id'] = 0
                dict1['product_id'] = list_text[0]['product_id']
                dict1['product_name'] = sku_name
                dict1['seller_id'] = '0'
                dict1['sid_str'] = ""
                dict1['sku'] = sku
                # print(dict1)
                dict_product.append (dict1)
                # print(dict_product)
            else:
                complete_flag = 1
        else:
            complete_flag = 1
    if complete_flag == 0:
        dict_pdd = {}
        dict_pdd['audit_user'] = []
        dict_pdd['cg_uid'] = ""
        dict_pdd['fee_part_type'] = 0
        dict_pdd['flow_id'] = ""
        dict_pdd['bin_type'] = 0
        dict_pdd['order_sn'] = ""
        dict_pdd['other_fee'] = ""
        dict_pdd['product_list'] = dict_product
        dict_pdd['purchase_order_sn'] = ""
        dict_pdd['remark'] = ""
        dict_pdd['req_time_sequence'] = "/api/module/inventoryReceipt/Outbound/completeOrder$$1"
        dict_pdd['return_price'] = ""
        dict_pdd['supplier_id'] = ""
        dict_pdd['to_wid'] = ""
        dict_pdd['type'] = 11
        dict_pdd['wid'] = 1461
        # please_pdd = urlencode(dict_pdd, encoding='gb2312')
        please_pdd = json.dumps(dict_pdd, ensure_ascii=False).encode ("utf-8")
        # please_pdd = json.dumps(dict_pdd)
        # please_pdd = dict_pdd
        print(please_pdd)
        # please_pdd = please_pdd.encode('utf-8')
        please = global_var.s.post(please_url, headers=find_headers, data=please_pdd)
        print(please.text)
        print(please.status_code)
        net_msg = json.loads(please.text)
        if net_msg['code'] == 1 and net_msg['msg'] == '请求成功':
            return True
        else:
            return False
    else:
        return False


def download_fab(sid, boxs):
    quantity = Quantity()
    msg = quantity.create_excl(sid, boxs)
    # msg = True
    if msg:
        return [True, "装箱清单下载完成"]
    else:
        return [False, False]


def read_excl():
    path = "./static/装箱信息单/上传装箱信息.xlsx"#C:/装箱信息单/上传装箱信息单.xlsx， ./static/装箱信息单/上传装箱信息.xlsx
    wb = openpyxl.load_workbook(path)
    wb_sheet = wb.active
    row = wb_sheet.max_row
    column = wb_sheet.max_column
    list_data = []
    list_fnsku = []
    for i in range(4, row+1):
        id = wb_sheet.cell(row=i, column=1).value
        if id:
            msku = wb_sheet.cell(row=i, column=2).value
            fnsku = wb_sheet.cell(row=i, column=3).value
            name = wb_sheet.cell(row=i, column=4).value
            sku = wb_sheet.cell(row=i, column=5).value
            num = wb_sheet.cell(row=i, column=6).value
            list_data.append({'id': id, 'msku': msku, 'fnsku': fnsku, 'name': name, 'sku': sku, 'num_shipment': num})
            list_fnsku.append(fnsku)
    return list_data, list_fnsku


def read_upload_excl(path):
    wb = openpyxl.load_workbook(path)
    wb_sheet = wb.active
    row = wb_sheet.max_row
    column = wb_sheet.max_column
    column1 = column
    list_data = []
    list_num = []
    fba_id = wb_sheet.cell(row=1, column=1).value
    # print(fba_id)
    for i in range(column, 0, -1):
        cell_value1 = wb_sheet.cell(row=3, column=i).value
        if cell_value1:
            column1 = i
            break
    box_num = wb_sheet.cell(row=1, column=2).value
    list_pa = []
    k = 2
    if wb_sheet.cell(row=2, column=1).value.find('PA') and wb_sheet.cell(row=2, column=1).value.find('PA') >= 0:
        k = 3
        for i in range(1, column):
            pa_name = wb_sheet.cell(row=2, column=i).value
            if pa_name:
                list_pa.append(pa_name)
    for i in range(k+1, row+1):
        id = wb_sheet.cell(row=i, column=1).value
        if id:
            list_row = []
            for j in range(1, int(box_num)+8):
                list_row.append(wb_sheet.cell(row=i, column=j).value)
            list_data.append(list_row)
    # print(list_pa)
    return list_data, fba_id, box_num, list_pa


def submmit(list_data, shipment_id, list_pa):
    # print("not finish")
    # change_excl(list_data)
    path = "./static/装箱信息单/上传装箱信息.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    count = 3
    finish_code = 0
    message = ''
    for i in range(0, len(list_data)):
        if count >= 4:
            long = len(list_data[i])
            total = 0
            for t in range(7, long):
                total = total + int(list_data[i][t])
            if total < int(list_data[i][5]):
                message = "%s号货物总数量不足，没有完全装箱！" % list_data[i][3]
                finish_code = 0
                break
            elif total > int(list_data[i][5]):
                message = "%s号货物总数量超出申报量，请仔细检查！" % list_data[i][3]
                finish_code = 0
                break
            else:
                for c in range(7, len(list_data[i]) + 1):
                    # print(count, c, list_data[i][(c - 1)])
                    sheet.cell(row=count, column=c, value=list_data[i][(c - 1)])
                count = count + 1
        else:
            for c in range(7, len(list_data[i]) + 1):
                sheet.cell(row=count, column=c, value=list_data[i][(c - 1)])
            count = count + 1
        finish_code = 1
    num_all = 0
    for i in range(0, len(list_data[0])-7):
        sum_num = 0
        for j in range(0, len(list_data)):
            if sheet.cell(row=j+4, column=i+8).value:
                sum_num += int(sheet.cell(row=j+4, column=i+8).value)
        num_all += sum_num
        sheet.cell(row=len(list_data)+3, column=i + 8, value=sum_num)
    sheet.cell(row=len(list_data)+3, column=len(list_data[0])+1, value=num_all)
    if finish_code:
        # 更新数据库
        if list_pa:
            quantity = Quantity()
            for i in list_pa:
                sql1 = "select * from `storage`.`relevance_hakone` where `箱号` = '%s'" % i
                result = func(sql1)
                if result:
                    sql2 = "DELETE FROM `storage`.`relevance_hakone` WHERE `箱号` = '%s'" % i
                    func(sql2, 'w')
                state = "已出货"
                func("UPDATE `storage`.`warehouse` SET `状态` = '%s' WHERE `箱号` = '%s'" % (state, str(i)), 'w')
            quantity.delect_sql(list_pa)
            wb.save("./static/装箱信息单/上传装箱信息.xlsx")
            # 可以自动上传后，下面的内容就不必了
        time1 = time.strftime("%Y_%m_%d", time.localtime())
        shutil.copyfile('./static/装箱信息单/上传装箱信息.xlsx',
                        './static/装箱信息单/%s上传装箱信息%s.xlsx' % (shipment_id, time1))
        return True, ['../装箱信息单/%s上传装箱信息%s.xlsx' % (shipment_id, time1), time1]
    else:
        return False, message


def change_excl(list_data, fba_id, box_num, list_pa=None):
    time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    path = f"./static/装箱信息单/{fba_id}_{time_now}装箱信息.xlsx"
    if os.path.exists(r'./static/装箱信息单/上传装箱信息.xlsx') == False:
        return False, "上传装箱信息文件丢失，请重新下载装箱信息。"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append([fba_id, box_num])
    if list_pa:
        sheet.append(list_pa)
    for i in list_data:
        sheet.append(i)
    wb.save(path)
    return True, [f"../装箱信息单/{fba_id}_{time_now}装箱信息.xlsx", time_now]


# python实现sign签名
class sign:
    def get_time(self):
        t1 = time.time()
        t = int(t1)
        return t

    def get_str(self, t, apikey):
        st = str(t)
        c = st + apikey
        return c

    def get_md5(self, c):
        md5 = hashlib.md5()
        md5.update(c.encode('UTF-8'))
        m = md5.hexdigest()
        return m

    def get_sign(self, apikey, body):
        s = sign()
        t = s.get_time()
        stt = s.get_str(t, apikey)
        m = s.get_md5(stt)
        # body['sign'] = m
        return m


######## AES-128-ECS 加密
class EncryptDate:
    def __init__(self, key):
        self.key = key.encode("utf-8")  # 初始化密钥
        self.length = AES.block_size  # 初始化数据块大小
        self.aes = AES.new(self.key, AES.MODE_ECB)  # 初始化AES,ECB模式的实例
        # 截断函数，去除填充的字符
        self.unpad = lambda date: date[0:-ord(date[-1])]

    def pad(self, text):
        """
        #填充函数，使被加密数据的字节码长度是block_size的整数倍
        """
        count = len(text.encode('utf-8'))
        add = self.length - (count % self.length)
        entext = text + (chr(add) * add)
        return entext

    def encrypt(self, encrData):  # 加密函数
        res = self.aes.encrypt(self.pad(encrData).encode("utf8"))
        msg = str(base64.b64encode(res), encoding="utf8")
        return msg

    def decrypt(self, decrData):  # 解密函数
        res = base64.decodebytes(decrData.encode("utf8"))
        msg = self.aes.decrypt(res).decode("utf8")
        return self.unpad(msg)


class Quantity(object):
    def __init__(self):
        self.app_id = "ak_nEfE94OSogf3x"
        app_secret = "g2BcerjK4fWmhGZoetCJHeVqJmHEmfLt3gWFjDrBLB1yxiapgUKH6kOVs2N9JH7SFuBKuOF8K/CrNNSeFO1KKtsL05z24j" \
                     "+AdWTW+V4op5QxDkmlTllvlprT8FfjctDdNDGrwHvBvE6s9h0pO0dNgopBAYiA7oosPzQhDF1A6XC1X/cZZmgBy3XRHyEv" \
                     "xTT40xzwVGish53R8dZt3YIxNtKSgrBloo/CRQsV01yU40nyQR9L9oML32VT0C16jBrxcoWthlGDwfBn+CtVUvws4imyZi" \
                     "+sG/CqZQeaVLkBCqLCgqw1VK4/a4jZws6HO3FMeBgvuf0aS5euNmfhkudQmg=="
        querystring = {"appId": f"{self.app_id}", "appSecret": f"{app_secret}"}
        url = "https://openapi.lingxing.com/api/auth-server/oauth/access-token"
        payload = ""
        headers = {
            "Content-Type": "application/x-www-form-urlencoded"
        }
        response = requests.post(url, data=payload, headers=headers, params=querystring)
        result = json.loads(response.text)
        # print(response.text)
        self.access_token = result['data']['access_token']
        self.refresh_token = result['data']['refresh_token']
        self.time = datetime.datetime.now().strftime("%Y-%m-%d")
        self.time = datetime.datetime.strptime(self.time, "%Y-%m-%d")
        # self.time = "2021-12-06"
        self.start_time = self.time - datetime.timedelta(days=1)
        self.time = self.time.strftime("%Y-%m-%d")
        self.start_time = self.start_time.strftime("%Y-%m-%d")
        # self.start_time = '2022-03-07'
        # print(self.time, self.start_time)

    def sql(self):
        self.connection = pymysql.connect(host='3354n8l084.goho.co',  # 数据库地址
                                          port=24824,
                                          user='test_user',  # 数据库用户名
                                          password='a123456',  # 数据库密码
                                          db='storage',  # 数据库名称
                                          charset='utf8',
                                          cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor()

    def sql_close(self):
        self.cursor.close()
        self.connection.close()

    def get_sign(self, body):
        apikey = "ak_nEfE94OSogf3x"
        # body = {"access_token": self.access_token,
        #         "timestamp": int(time.time()),
        #         "start_date": f"{self.start_time}",
        #         "end_date": f"{self.time}",
        #         "app_key": apikey}
        # print(body)
        # bb = sign().get_sign(apikey, body)
        # print(bb)

        # 目标md5串
        str_parm = ''
        # 将字典中的key排序
        for p in sorted (body):
            # 每次排完序的加到串中
            # if body[p]:
            # str类型需要转化为url编码格式
            if isinstance (body[p], str):
                str_parm = str_parm + str (p) + "=" + str (urllib.parse.quote (body[p])) + "&"
                continue
            # if isinstance(body[p], dict):
            #     for i in body[p]:
            #         body[p][i] = str(body[p][i])
            # if isinstance(body[p], list) and isinstance(body[p][0], dict):
            #     for i in range(0, len(body[p])):
            #         for j in body[p][i]:
            #             body[p][i][j] = str(urllib.parse.quote(str(body[p][i][j])))
            # if p == "product_list":
            #     str_parm = str_parm + str(p) + "=" + '[{"sku":"GCWL.897","good_num":1,"bad_num":0,"seller_id":0,"fnsku":""}]' + "&"
            #     continue
            #     body[p][0] = str(urllib.parse.quote(body[p][0]))
            #     print(str(urllib.parse.quote(body[p][0])))
            str_value = str (body[p])
            str_value = str_value.replace (" ", "")
            # str_value = str_value.replace("?", " ")
            str_value = str_value.replace ("'", "\"")
            str_parm = str_parm + str (p) + "=" + str_value + "&"
        # 加上对应的key
        str_parm = str_parm.rstrip ('&')
        # print("字符串拼接1:", str_parm)
        str_parm = str_parm.replace ("?", " ")
        # print("字符串拼接2:", str_parm)
        if isinstance (str_parm, str):
            # 如果是unicode先转utf-8
            parmStr = str_parm.encode ("utf-8")
            # parmStr = str_parm
            m = hashlib.md5 ()
            m.update (parmStr)
            md5_sign = m.hexdigest ()
            # print(m.hexdigest())
            md5_sign = md5_sign.upper ()
            # print("MD5加密:", md5_sign)
        eg = EncryptDate (apikey)  # 这里密钥的长度必须是16的倍数
        res = eg.encrypt (md5_sign)
        # print("AES加密:", res)
        # print(eg.decrypt(res))
        return res

    def create_excl(self, fba_shipment, box_num):
        url = "https://openapi.lingxing.com/erp/sc/data/fba_report/shipmentList"
        body = {"access_token": self.access_token,
                "timestamp": int(time.time()),
                "app_key": self.app_id,
                "shipment_id": fba_shipment,
                "sid": '462', "start_date": f"{self.start_time}", "end_date": f"{self.time}"
                }
        res = self.get_sign(body)
        querystring = {"access_token": self.access_token,
                       "timestamp": int(time.time()),
                       "app_key": self.app_id,
                       "sign": res
                       }
        payload = {"start_date": f"{self.start_time}", "shipment_id": fba_shipment, "sid": '462',
                   "end_date": f"{self.time}"}
        payload = json.dumps(payload)
        headers = {
            "Content-Type": "application/json",
            "Authorization": "bearer {{access_token}}"
        }
        response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
        result1 = json.loads(response.text)
        #print(result1)
        # 判断数据是否正常获取
        list_sku = []
        if result1['code'] == 0 and result1['message'] == 'success':
            # 判断查询的FBA号是否符合装箱出货要求
            for k in result1['data']['list']:
                if k['is_closed'] == 0 and k['shipment_status'] == 'WORKING' and k['shipment_id'] == fba_shipment:
                    for i in k['item_list']:
                        # # 根据fnsku获取sku和品名
                        list_msg = self.fnsku_to_sku(i['fnsku'])
                        list_sku.append([i['msku'], i['fnsku'], list_msg[1], list_msg[0], i['quantity_shipped']])
        else:
            return False
        #print(list_sku)
        self.excl_test(list_sku, box_num, fba_shipment)
        # wb.save('C:/装箱信息单/上传装箱信息单.xlsx')
        return './static/装箱信息单/上传装箱信息.xlsx'

    def fnsku_to_sku(self, fnsku):
        # API接口路径
        list_sid = [400, 1587, 1588, 273, 461, 272, 2463, 463, 821, 822, 3294, 2638, 2637, 2636, 974, 973, 972, 971,
                    462, 3000]
        # list_sid = json.dumps(list_sid)
        url = "https://openapi.lingxing.com/erp/sc/routing/fba/shipment/getFbaProductList"
        body = {"access_token": self.access_token,
                "timestamp": int(time.time()),
                "app_key": self.app_id,
                "search_field": 'fnsku',
                "offset": 0,
                "length": 20,
                "sids": list_sid,
                "search_value": fnsku
                }
        # print ("生成sign签名参数:", body)
        # 生成sign签名
        res = self.get_sign(body)
        querystring = {"access_token": self.access_token,
                       "timestamp": int(time.time()),
                       "app_key": self.app_id,
                       "sign": res
                       }
        payload = {"search_field": 'fnsku', "search_value": fnsku, "sids": list_sid, "offset": 0, "length": 20
                   }
        payload = json.dumps(payload)
        headers = {
            "Content-Type": "application/json",
            "Authorization": "bearer {{access_token}}"
        }
        response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
        result1 = json.loads(response.text)
        list_sku = []
        # 判断数据是否正常获取
        #print (result1)
        if result1['code'] == 0 and result1['message'] == 'success':
            for i in result1['data']:
                if i['fnsku'] == fnsku and i['sku']:
                    list_sku = [i['sku'], i['local_name']]
                    break
        if list_sku:
            return list_sku
        else:
            return [False, False]

    def excl_test(self, list_sku, box_num, shipment_id):
        # 复制模板表格
        copy('C:/装箱信息单/上传装箱信息模板.xlsx', './static/装箱信息单/上传装箱信息.xlsx')
        wb = openpyxl.load_workbook('./static/装箱信息单/上传装箱信息.xlsx')
        wb_sheet = wb.active
        alignment = Alignment(horizontal="center")
        # 插入行
        for i in range(4, len(list_sku) + 4):
            if i > 4:
                wb_sheet.insert_rows(4)
        # 修改单元格格式并写入表头
        for i in range(1, int(box_num) + 8):
            # 设置单元格输入对齐方式：居中
            for k in range(4, 3 + len(list_sku)):
                wb_sheet.cell(row=k, column=i).alignment = alignment
            # 设置单元格背景颜色：灰色
            wb_sheet.cell(row=4 + len(list_sku), column=i).fill = PatternFill("solid", fgColor="808080")
            # 输入表头
            if i >= 8:
                wb_sheet.cell(row=3, column=i).value = f'第{i - 7}箱'
        wb_sheet.cell(row=4 + len (list_sku), column=int(box_num) + 8).fill = PatternFill("solid", fgColor="808080")
        # 写入FBA货件单号
        wb_sheet.cell(row=2, column=2).value = shipment_id
        # 合并单元格
        for i in range(5, 9):
            merged = f'E{i + len (list_sku)}:G{i + len (list_sku)}'
            # print(merged)
            wb_sheet.merge_cells(merged)
        # 写入FBA出货信息
        row = 4
        sum_num = 0
        #print(list_sku)
        for i in list_sku:
            #print(row, i)
            wb_sheet.cell(row=row, column=1).value = row - 3
            wb_sheet.cell(row=row, column=2).value = i[0]
            wb_sheet.cell(row=row, column=3).value = i[1]
            wb_sheet.cell(row=row, column=4).value = i[2]
            wb_sheet.cell(row=row, column=5).value = i[3]
            wb_sheet.cell(row=row, column=6).value = i[4]
            wb_sheet.cell(row=row, column=7).value = 0
            row += 1
            sum_num += i[4]
        # 写入合计发货数量
        wb_sheet.cell(row=4 + len(list_sku), column=6).value = sum_num
        wb.save('./static/装箱信息单/上传装箱信息.xlsx')

    def order_out(self, list_index, list_num, warehouse):
        # print(sku_list)
        dict_warehouse = {'工厂仓库': '2156', '横中路仓库-加拿大': '1489', '横中路仓库-日本': '1490',
                          '横中路仓库-美国': '1461', '横中路仓库-英国': '1488',
                          '横中路仓库-德国': '2382', '淘汰-横中路成品仓库-加拿大': '1476', '淘汰-横中路成品仓库-日本': '1477',
                          '淘汰-横中路成品仓库-美国': '1399', '淘汰-横中路成品仓库-英国': '1478', '百汇办公室': '414'}
        product_list = []
        for i in range(0, len(list_index)):
            product_list.append({"sku": list_index[i], "good_num": list_num[i], "bad_num": 0, "seller_id": 0, "fnsku": ''})
        url = "https://openapi.lingxing.com/erp/sc/routing/storage/storage/orderAddOut"
        body = {"access_token": self.access_token,
                "timestamp": int(time.time()),
                "app_key": self.app_id,
                # "wid": dict_warehouse['工厂仓库'],
                "type": 11,
                "product_list": product_list,
                "sys_wid": dict_warehouse[warehouse]
                }
        # print("生成sign签名参数：", body)
        res = self.get_sign(body)
        querystring = {"access_token": self.access_token,
                       "timestamp": int(time.time()),
                       "app_key": self.app_id,
                       "sign": res
                       }
        payload = {"product_list": product_list, "type": 11, "sys_wid": dict_warehouse[warehouse]}
        # print(payload)
        payload = json.dumps(payload)
        headers = {
            "Content-Type": "application/json",
            "Authorization": "bearer {{access_token}}"
        }
        response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
        result1 = json.loads(response.text)
        if result1['code'] == 0 and result1['message'] == 'success':
            return True
        else:
            return False

    def order_put(self, list_index, list_num, warehouse):

        # print(sku_list)
        dict_warehouse = {'工厂仓库': '2156', '横中路仓库-加拿大': '1489', '横中路仓库-日本': '1490',
                          '横中路仓库-美国': '1461', '横中路仓库-英国': '1488',
                          '横中路仓库-德国': '2382', '淘汰-横中路成品仓库-加拿大': '1476', '淘汰-横中路成品仓库-日本': '1477',
                          '淘汰-横中路成品仓库-美国': '1399', '淘汰-横中路成品仓库-英国': '1478', '百汇办公室': '414'}
        product_list = []
        for i in range(0, len(list_index)):
            product_list.append({"sku": list_index[i], "good_num": list_num[i], "bad_num": 0, "seller_id": 0, "fnsku": '', 'price': 0.00})
        url = "https://openapi.lingxing.com/erp/sc/routing/storage/storage/orderAdd"
        body = {"access_token": self.access_token,
                "timestamp": int(time.time()),
                "app_key": self.app_id,
                "type": 1,
                "product_list": product_list,
                "sys_wid": dict_warehouse[warehouse]
                }
        res = self.get_sign(body)
        querystring = {"access_token": self.access_token,
                       "timestamp": int(time.time()),
                       "app_key": self.app_id,
                       "sign": res
                       }
        payload = {"product_list": product_list, "type": 1, "sys_wid": dict_warehouse[warehouse]}
        # print(payload)
        payload = json.dumps(payload)
        headers = {
            "Content-Type": "application/json",
            "Authorization": "bearer {{access_token}}"
        }
        response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
        result1 = json.loads(response.text)
        if result1['code'] == 0 and result1['message'] == 'success':
            return True
        else:
            return False

    def get_warehouse_num(self, sku):
        # 生成sign签名
        # API接口路径
        url = "https://openapi.lingxing.com/erp/sc/routing/data/local_inventory/inventoryDetails"
        body = {"access_token": self.access_token,
                "timestamp": int(time.time()),
                "app_key": "ak_nEfE94OSogf3x", "offset": 0, "length": 400, "wid": "2156"
               }
        # "wid": "2156,1489,2382,1490,1461,1488,1476,1477,1399,1478,414", "offset": 0, "length": 400
        res = self.get_sign(body)
        # res = ''
        querystring = {"access_token": self.access_token,
                       "timestamp": int(time.time()),
                       "app_key": self.app_id,
                       "sign": res
                       }
        payload = {"wid": "2156", "offset": 0, "length": 400}
        payload = json.dumps(payload)
        headers = {
            "Content-Type": "application/json",
            "Authorization": "bearer {{access_token}}"
        }
        response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
        result1 = json.loads(response.text)
        product_id = ''
        # 判断数据是否正常获取
        if result1['code'] == 0 and result1['message'] == "success":
            # 获取库存明细数据条数
            # 以四百条数据为一页分页查询
            for i in result1['data']:
                if i['sku'] == sku:
                    product_id = i['product_id']
                    break
        if product_id:
            return product_id
        else:
            payload = {"wid": '2156', "offset": 400, "length": 400}
            body['offset'] = 400
            res = self.get_sign(body)
            querystring['sign'] = res
            payload = json.dumps(payload)
            response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
            result = json.loads(response.text)
            if result['code'] == 0 and result['message'] == "success":
                for i in result['data']:
                    if i['sku'] == sku:
                        product_id = i['product_id']
                        break
            if product_id:
                return product_id
            else:
                return False

    def get_product(self, product_id, sku):
        url = "https://openapi.lingxing.com/erp/sc/routing/data/local_inventory/productInfo"
        body = {"access_token": self.access_token,
                "timestamp": int(time.time()),
                "app_key": self.app_id,
                "id": product_id
                }
        # print(querystring)
        res = self.get_sign(body)
        querystring = {"access_token": self.access_token,
                       "timestamp": int(time.time()),
                       "app_key": self.app_id,
                       "sign": res
                       }
        payload = {"id": product_id}
        payload = json.dumps(payload)
        headers = {
            "Content-Type": "application/json",
            "Authorization": "bearer {{access_token}}"
        }
        # time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        # print("访问API：", time_now)
        response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
        result1 = json.loads(response.text)
        product_name = ''
        if result1['code'] == 0 and result1['message'] == 'success':
            if result1['data']['sku'] == sku:
                product_name = result1['data']['product_name']
        return product_name

    def get_fnsku(self, index):
        self.sql()
        sql = f"select `FNSKU1`, `数量1` from `storage`.`warehouse` where `箱号` = '{index}' and `状态` = '已打包'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            return result[0]['FNSKU1'], result[0]['数量1']
        else:
            return False, False

    def delect_sql(self, list_pa):
        self.sql()
        for i in list_pa:
            sql3 = "select * from `storage`.`relevance_hakone` where `箱号` = '%s'" % i
            self.cursor.execute(sql3)
            result = self.cursor.fetchall()
            if result:
                sql4 = "DELETE FROM `storage`.`relevance_hakone` WHERE `箱号` = '%s'" % i
                self.cursor.execute(sql4)
        self.connection.commit()
        self.sql_close()