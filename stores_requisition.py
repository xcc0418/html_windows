# import os
# import shutil
import requests
import json
import time
import hashlib
import urllib
import base64
import datetime
import pymysql
from Crypto.Cipher import AES
import openpyxl
# from shutil import copy
# from openpyxl.styles import PatternFill
# from openpyxl.styles import Alignment


######## AES-128-ECS 加密
class EncryptDate:
    def __init__(self, key):
        self.key = key.encode("utf-8")  # 初始化密钥
        self.length = AES.block_size  # 初始化数据块大小
        self.aes = AES.new(self.key, AES.MODE_ECB)  # 初始化AES,ECB模式的实例
        # 截断函数，去除填充的字符
        self.unpad = lambda date: date[0:-ord(date[-1])]

    def pad(self, text):
        """
        #填充函数，使被加密数据的字节码长度是block_size的整数倍
        """
        count = len(text.encode('utf-8'))
        add = self.length - (count % self.length)
        entext = text + (chr(add) * add)
        return entext

    def encrypt(self, encrData):  # 加密函数
        res = self.aes.encrypt(self.pad(encrData).encode("utf8"))
        msg = str(base64.b64encode(res), encoding="utf8")
        return msg

    def decrypt(self, decrData):  # 解密函数
        res = base64.decodebytes(decrData.encode("utf8"))
        msg = self.aes.decrypt(res).decode("utf8")
        return self.unpad(msg)


class Quantity(object):
    def __init__(self):
        self.app_id = "ak_nEfE94OSogf3x"
        app_secret = "g2BcerjK4fWmhGZoetCJHeVqJmHEmfLt3gWFjDrBLB1yxiapgUKH6kOVs2N9JH7SFuBKuOF8K/CrNNSeFO1KKtsL05z24j" \
                     "+AdWTW+V4op5QxDkmlTllvlprT8FfjctDdNDGrwHvBvE6s9h0pO0dNgopBAYiA7oosPzQhDF1A6XC1X/cZZmgBy3XRHyEv" \
                     "xTT40xzwVGish53R8dZt3YIxNtKSgrBloo/CRQsV01yU40nyQR9L9oML32VT0C16jBrxcoWthlGDwfBn+CtVUvws4imyZi" \
                     "+sG/CqZQeaVLkBCqLCgqw1VK4/a4jZws6HO3FMeBgvuf0aS5euNmfhkudQmg=="
        querystring = {"appId": f"{self.app_id}", "appSecret": f"{app_secret}"}
        url = "https://openapi.lingxing.com/api/auth-server/oauth/access-token"
        payload = ""
        headers = {
            "Content-Type": "application/x-www-form-urlencoded"
        }
        response = requests.post(url, data=payload, headers=headers, params=querystring)
        result = json.loads(response.text)
        print(result)
        self.access_token = result['data']['access_token']
        self.refresh_token = result['data']['refresh_token']
        self.time = datetime.datetime.now().strftime("%Y-%m-%d")
        self.time = datetime.datetime.strptime(self.time, "%Y-%m-%d")
        # self.time = "2021-12-06"
        self.start_time = self.time - datetime.timedelta(days=1)
        self.time = self.time.strftime("%Y-%m-%d")
        self.start_time = self.start_time.strftime("%Y-%m-%d")
        # self.start_time = '2022-03-07'

    def get_access_token(self):
        try:
            self.app_id = "ak_nEfE94OSogf3x"
            app_secret = "g2BcerjK4fWmhGZoetCJHeVqJmHEmfLt3gWFjDrBLB1yxiapgUKH6kOVs2N9JH7SFuBKuOF8K/CrNNSeFO1KKtsL05z24j" \
                         "+AdWTW+V4op5QxDkmlTllvlprT8FfjctDdNDGrwHvBvE6s9h0pO0dNgopBAYiA7oosPzQhDF1A6XC1X/cZZmgBy3XRHyEv" \
                         "xTT40xzwVGish53R8dZt3YIxNtKSgrBloo/CRQsV01yU40nyQR9L9oML32VT0C16jBrxcoWthlGDwfBn+CtVUvws4imyZi" \
                         "+sG/CqZQeaVLkBCqLCgqw1VK4/a4jZws6HO3FMeBgvuf0aS5euNmfhkudQmg=="
            querystring = {"appId": f"{self.app_id}", "appSecret": f"{app_secret}"}
            url = "https://openapi.lingxing.com/api/auth-server/oauth/access-token"
            payload = ""
            headers = {
                "Content-Type": "application/x-www-form-urlencoded"
            }
            response = requests.post(url, data=payload, headers=headers, params=querystring)
            result = json.loads(response.text)
            # print(response.text)
            self.access_token = result['data']['access_token']
        except Exception as e:
            print(e)

    def sql(self):
        self.connection = pymysql.connect(host='3354n8l084.goho.co',  # 数据库地址
                                          port=24824,
                                          user='test_user',  # 数据库用户名
                                          password='a123456',  # 数据库密码
                                          db='storage',  # 数据库名称
                                          charset='utf8',
                                          cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor()

    def sql_close(self):
        self.cursor.close()
        self.connection.close()

    def get_sign(self, body):
        apikey = "ak_nEfE94OSogf3x"
        # body = {"access_token": self.access_token,
        #         "timestamp": int(time.time()),
        #         "start_date": f"{self.start_time}",
        #         "end_date": f"{self.time}",
        #         "app_key": apikey}
        # print(body)
        # bb = sign().get_sign(apikey, body)
        # print(bb)

        # 目标md5串
        str_parm = ''
        # 将字典中的key排序
        for p in sorted (body):
            # 每次排完序的加到串中
            # if body[p]:
            # str类型需要转化为url编码格式
            if isinstance (body[p], str):
                str_parm = str_parm + str (p) + "=" + str (urllib.parse.quote (body[p])) + "&"
                continue
            # if isinstance(body[p], dict):
            #     for i in body[p]:
            #         body[p][i] = str(body[p][i])
            # if isinstance(body[p], list) and isinstance(body[p][0], dict):
            #     for i in range(0, len(body[p])):
            #         for j in body[p][i]:
            #             body[p][i][j] = str(urllib.parse.quote(str(body[p][i][j])))
            # if p == "product_list":
            #     str_parm = str_parm + str(p) + "=" + '[{"sku":"GCWL.897","good_num":1,"bad_num":0,"seller_id":0,"fnsku":""}]' + "&"
            #     continue
            #     body[p][0] = str(urllib.parse.quote(body[p][0]))
            #     print(str(urllib.parse.quote(body[p][0])))
            str_value = str (body[p])
            str_value = str_value.replace (" ", "")
            # str_value = str_value.replace("?", " ")
            str_value = str_value.replace ("'", "\"")
            str_parm = str_parm + str (p) + "=" + str_value + "&"
        # 加上对应的key
        str_parm = str_parm.rstrip ('&')
        # print("字符串拼接1:", str_parm)
        str_parm = str_parm.replace ("?", " ")
        # print("字符串拼接2:", str_parm)
        if isinstance (str_parm, str):
            # 如果是unicode先转utf-8
            parmStr = str_parm.encode ("utf-8")
            # parmStr = str_parm
            m = hashlib.md5 ()
            m.update (parmStr)
            md5_sign = m.hexdigest ()
            # print(m.hexdigest())
            md5_sign = md5_sign.upper ()
            # print("MD5加密:", md5_sign)
        eg = EncryptDate (apikey)  # 这里密钥的长度必须是16的倍数
        res = eg.encrypt (md5_sign)
        # print("AES加密:", res)
        # print(eg.decrypt(res))
        return res

    def get_class(self):
        self.sql()
        sql = "select * from `storage`.`category`"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        list_class = []
        for i in result:
            list_class.append(i['品类'])
        self.sql_close()
        return list_class

    def add_class(self, class_name):
        try:
            self.sql()
            sql = "select * from `storage`.`category` where `品类` = '%s'" % class_name
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            if result:
                return False, "这个品类已存在"
            else:
                sql = "insert into `storage`.`category`(`品类`)values('%s')" % class_name
                self.cursor.execute(sql)
                self.connection.commit()
                self.sql_close()
                list_class = self.get_class()
                return True, list_class
        except Exception as e:
            return False, e

    def find_po(self, po_number, po_name):
        if po_number and po_name:
            time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
            self.sql()
            sql = "select * from `storage`.`stores_requisition` where `领料单号` = '%s'" % f'{po_number}-{po_name}-{time_now}'
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            if result:
                self.sql_close()
                return False, '这个领料单号已被创建，请检查'
            else:
                sql = "select * from `storage`.`po_sku` where `PO` = '%s'" % po_number
                self.cursor.execute(sql)
                result1 = self.cursor.fetchall()
                self.sql_close()
                if result1:
                    return True, '领料单号创建成功'
                else:
                    return False, f"没有找到{po_number}这个订单号，请检查"
        else:
            if not po_number:
                return False, '请先输入领料单号'
            elif not po_name:
                return False, '请选择品类'
            else:
                return False, '正在生成领料单'

    def write_sql(self, po_name, class_name, list_sku):
        try:
            dict_order = {}
            list_order = []
            for i in range(0, len(list_sku)):
                if list_sku[i][0] in dict_order:
                    continue
                dict_order[list_sku[i][0]] = list_sku[i][1]
                list_order.append(list_sku[i][0])
            list_order = self.get_order_msg(list_order, dict_order)
            if list_order:
                time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
                po_id = f'{po_name}-{class_name}-{time_now}'
                for i in range(0, len(list_sku)):
                    product_name = self.get_product_name('sku', list_sku[i][0])
                    list_sku[i].append(product_name)
                self.sql()
                for i in list_order:
                    num = list_order[i][0]
                    sql = "insert into `storage`.`stores_requisition`(`领料单号`, `SKU`, `品名`, `需求数量`, `实际" \
                          "领料数量`, `实际出库数量`, `状态`)values('%s', '%s', '%s', %d, %d,0,'%s')" \
                          "" % (po_id, i, list_order[i][1], num, num, '未完成')
                    self.cursor.execute(sql)
                    self.writer_sku_demand(i, num, list_order[i][1])
                for i in list_sku:
                    if i[0].find('GCWL') >= 0:
                        continue
                    sql = "SELECT * FROM `storage`.`po_sku` WHERE `PO` = '%s' AND `SKU` = '%s'" % (po_name, i[0])
                    self.cursor.execute(sql)
                    result = self.cursor.fetchall()
                    if result:
                        sql1 = "update `storage`.`po_sku` set `已领料数量` = %d where `PO` = '%s' and `SKU` = '" \
                               "%s'" % (int(i[1])+int(result[0]['已领料数量']), po_name, i[0])
                    else:
                        sql1 = "insert into `storage`.`po_sku`(`PO`, `SKU`, `品名`, `订单数量`, `已领料数量`)values" \
                               "('%s', '%s', '%s', %d, %d)" % (po_name, i[0], i[2], i[1], 0)
                    self.cursor.execute(sql1)
                self.connection.commit()
                self.sql_close()
                self.writer_pdf(po_id, list_order)
                return True, [f'../领料单/{po_id}.xlsx', f'{po_id}']
            else:
                return False, '没有找到对应信息，请查看输入sku是否正确'
        except Exception as e:
            print(e)
            self.connection.rollback()
            return False, str(e)

    def get_product_id(self, list_sku):
        list_id = []
        self.sql()
        for i in range(0, len(list_sku)):
            sql = "select * from `data_read`.`product_id` where `SKU` = '%s'" % list_sku[i]
            # print(sql)
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            if result:
                list_id.append(result[0]['ID'])
            else:
                self.sql_close()
                return False
        self.sql_close()
        return list_id

    def get_order_msg(self, list_order, dict_order=None):
        list_sku = {}
        list_id = self.get_product_id(list_order)
        if list_id:
            self.get_access_token()
            time_stamp = int(time.time())
            url = "https://openapi.lingxing.com/erp/sc/routing/data/local_inventory/batchGetProductInfo"
            body = {"access_token": self.access_token,
                    "timestamp": time_stamp,
                    "app_key": "ak_nEfE94OSogf3x",
                    "productIds": list_id
                    }
            res = self.get_sign(body)
            querystring = {"access_token": self.access_token,
                           "timestamp": time_stamp,
                           "app_key": self.app_id,
                           "sign": res
                           }
            payload = {"productIds": list_id}
            payload = json.dumps(payload)
            headers = {
                "Content-Type": "application/json",
                "Authorization": "bearer {{access_token}}"
            }
            response = requests.request("POST", url, data=payload, headers=headers, params=querystring)
            result1 = json.loads(response.text)
            # print(result1)
            if result1['code'] == 0 and result1['message'] == 'success':
                for i in result1['data']:
                    warehouse_sku = i['sku']
                    if warehouse_sku.find('GCWL') >= 0:
                        warehouse_num = dict_order[warehouse_sku]
                        if warehouse_sku not in list_sku:
                            list_sku[warehouse_sku] = [warehouse_num, i['product_name']]
                        else:
                            num_order = int(list_sku[warehouse_sku][0])
                            list_sku[warehouse_sku][0] = num_order + int(warehouse_num)
                    else:
                        for j in i['combo_product_list']:
                            product_id = j['product_id']
                            product_name = self.get_product_name('ID', product_id)
                            # print(product_name, product_id)
                            sku = j['sku']
                            if product_name.find('/100码') >= 0:
                                warehouse_num = int(j['quantity']) * dict_order[warehouse_sku]
                            elif product_name.find('0.01米') >= 0:
                                if product_name.find('肩带') >= 0 or product_name.find('松紧带') >= 0:
                                    warehouse_num = int(j['quantity']) * dict_order[warehouse_sku]
                                else:
                                    warehouse_num = int(j['quantity']) * dict_order[warehouse_sku] * 0.01
                            elif product_name.find('0.001方') >= 0:
                                warehouse_num = int(j['quantity']) * dict_order[warehouse_sku] * 0.001
                            else:
                                warehouse_num = int(j['quantity']) * dict_order[warehouse_sku]
                            if sku not in list_sku:
                                list_sku[sku] = [warehouse_num, product_name]
                            else:
                                num_order = int(list_sku[sku][0])
                                list_sku[sku][0] = num_order + int(warehouse_num)
                # print(list_sku)
                return list_sku
            else:
                print(result1)
                return False
        else:
            print(list_sku)
            return False

    def get_product_name(self, index, id):
        self.sql()
        sql = ''
        if index == 'ID':
            sql = "select * from `data_read`.`product_id` where `ID` = '%s'" % id
        if index == 'sku':
            sql = "select * from `data_read`.`product_id` where `SKU` = '%s'" % id
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            return result[0]['品名']
        else:
            return False

    def find_sql(self, list_sku, po_number):
        self.sql()
        list_order = []
        list_sku_order = []
        sql = "select `SKU` from `storage`.`po_order` where `PO` = '%s'" % po_number
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        for i in result:
            list_order.append(i['SKU'])
        for i in list_sku:
            list_sku_order.append(i[0])
        if list_sku_order in list_order:
            return True
        else:
            return False

    def writer_sku_demand(self, sku, num, name):
        sql = "select * from `storage`.`sku_demand` where `SKU` = '%s'" % sku
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            sql1 = "update `storage`.`sku_demand` set `需求量累计` = %d where `SKU` = '%s'" % (int(result[0]['需求量累计'])+int(num), sku)
        else:
            sql1 = "insert into `storage`.`sku_demand`(`SKU`, `品名`, `需求量累计`, `出库量累计`)values('%s', '%s', %d, %d)" % (sku, name, num, 0)
        # print(sql1)
        self.cursor.execute(sql1)

    def writer_pdf(self, po_name, list_order):
        wb = openpyxl.Workbook()
        wb_sheet = wb.active
        wb_sheet.column_dimensions['A'].width = 15.0
        wb_sheet.column_dimensions['B'].width = 55.0
        wb_sheet.column_dimensions['C'].width = 15.0
        wb_sheet.column_dimensions['D'].width = 15.0
        wb_sheet.append([po_name])
        wb_sheet.append(['SKU', '品名', '需求数量', '实际领料数量'])
        for i in list_order:
            wb_sheet.append([i, list_order[i][1], list_order[i][0], list_order[i][0]])
        excel_path = f'./static/领料单/{po_name}.xlsx'
        wb.save(excel_path)

    def find_sku(self, str_html, flag):
        if flag == 'sku':
            index_final = 0
            index = str_html.find('GC')
            if str_html[index-1] == '[':
                index_heard = index-1
            else:
                index_heard = index
            for i in range(1, 8):
                if str_html[index+i] == '<':
                    index_final = index+i
                    break
            # print(index_heard, index_final)
            sku = str_html[index_heard: index_final]
            return sku
        if flag == 'num':
            index_final = str_html.find('</td>')
            index_heard = 0
            for i in range(1, 8):
                if str_html[index_final - i] == '>':
                    index_heard = index_final - i + 1
                    break
            # print(index_heard, index_final)
            sku = str_html[index_heard: index_final]
            return sku

    def get_po(self, po_name):
        self.sql()
        sql = "select * from `storage`.`po_sku` where `PO` = '%s'" % po_name
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        list_msg = []
        if result:
            for i in result:
                list_msg.append([i['SKU'], i['品名'], i['订单数量'], i['已领料数量'], 0])
            return list_msg
        else:
            return False

    def upload_sql(self, filename):
        try:
            dict_sku = {}
            wb = openpyxl.load_workbook(filename, data_only=True)
            wb_sheet = wb.active
            row1 = wb_sheet.max_row
            column1 = wb_sheet.max_column
            for i in range(row1, 0, -1):
                cell_value1 = wb_sheet.cell(row=i, column=1).value
                if cell_value1:
                    row1 = i
                    break
            list_heard = []
            for i in range(1, column1 + 1):
                heard = wb_sheet.cell(row=1, column=i).value
                if heard:
                    list_heard.append(heard)
            # print(self.list_wb_heard)
            for i in range(2, row1 + 1):
                po_id = wb_sheet.cell(row=i, column=(list_heard.index('PO号') + 1)).value
                if po_id:
                    index = wb_sheet.cell(row=i, column=(list_heard.index('是否结束') + 1)).value
                    if index == '结束':
                        if po_id in dict_sku:
                            del dict_sku[po_id]
                        continue
                    else:
                        sku = wb_sheet.cell(row=i, column=(list_heard.index('ErpSKU\n(公式）') + 1)).value
                        if sku.find('GC') >= 0:
                            name = wb_sheet.cell(row=i, column=(list_heard.index('产品型号') + 1)).value
                            num = wb_sheet.cell(row=i, column=(list_heard.index('订单数量') + 1)).value
                            if po_id not in dict_sku:
                                dict_sku[po_id] = []
                                dict_sku[po_id].append([sku, num, name])
                            else:
                                dict_sku[po_id].append([sku, num, name])
            print(len(dict_sku))
            dict_gcwl = {}
            for i in dict_sku:
                list_sku = []
                dict_order = {}
                for j in dict_sku[i]:
                    if j[0] not in dict_order:
                        dict_order[j[0]] = j[1]
                        list_sku.append(j[0])
                # print(list_sku, dict_order)
                list_gcwl = self.get_order_msg(list_sku, dict_order)
                if list_gcwl:
                    dict_gcwl[i] = list_gcwl
                    continue
                else:
                    return False, False
                # print(list_gcwl)
            print(len(dict_gcwl))
            msg, message = self.insert_sql(dict_gcwl, dict_sku)
            if msg:
                return True, True
            else:
                return False, message
        except Exception as e:
            print(e)
            return False, e

    def insert_sql(self, dict_gcwl, dict_sku):
        try:
            self.sql()
            for i in dict_gcwl:
                sql = "select * from `storage`.`po_order` where `PO` = '%s'" % i
                self.cursor.execute(sql)
                result = self.cursor.fetchall()
                if result:
                    for j in dict_gcwl[i]:
                        # print(i, j)
                        sql1 = "select * from `storage`.`po_order` where `PO` = '%s' and `SKU` = '%s'" % (i, j)
                        self.cursor.execute(sql1)
                        result1 = self.cursor.fetchall()
                        if result1:
                            if dict_gcwl[i][j] == result1[0]['订单数量']:
                                # messagebox.showinfo(message=f'{i}重复，已在数据库中有记录')
                                continue
                            else:
                                sql2 = "update `storage`.`po_order` set `订单数量` = %d where `PO` = '%s' and" \
                                       " `SKU` = '%s'" % (dict_gcwl[i][j][0], i, j)
                                self.cursor.execute(sql2)
                else:
                    for j in dict_gcwl[i]:
                        sql = "insert into `storage`.`po_order`(`PO`, `SKU`, `订单数量`)values('%s', '%s', %d)" % (i, j, dict_gcwl[i][j][0])
                        self.cursor.execute(sql)
            for i in dict_sku:
                for j in dict_sku[i]:
                    sql1 = "select * from `storage`.`po_sku` where `PO` = '%s' and `SKU` = '%s'" % (i, j[0])
                    self.cursor.execute(sql1)
                    result1 = self.cursor.fetchall()
                    if result1:
                        sql2 = "update `storage`.`po_sku` set `订单数量` = %d where `PO` = '%s' and" \
                               " `SKU` = '%s'" % (int(j[1]), i, j[0])
                        self.cursor.execute(sql2)
                    else:
                        sql = "insert into `storage`.`po_sku`(`PO`, `SKU`, `品名`, `订单数量`, `已领料数量`)values" \
                              "('%s', '%s', '%s', %d, %d) " % (i, j[0], j[2], int(j[1]), 0)
                        # print(sql)
                        self.cursor.execute(sql)
            self.connection.commit()
            self.sql_close()
            return True, True
        except Exception as e:
            print(e)
            return False, e


if __name__ == '__main__':
    quantity = Quantity()
    # quantity.get_order_msg(['GC.410'], {'GC.410': 800})
    quantity.upload_sql('D:/生产日程表/生产日程表202305171418.xlsx')
