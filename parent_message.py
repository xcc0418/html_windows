import openpyxl
import requests
import pymysql
import json
# import datetime
import hashlib
import urllib
import base64
from Crypto.Cipher import AES
import zipfile
import time
import os
from datetime import datetime
import pandas as pd
import string_html


class EncryptDate:
    def __init__(self, key):
        self.key = key.encode("utf-8")  # 初始化密钥
        self.length = AES.block_size  # 初始化数据块大小
        self.aes = AES.new(self.key, AES.MODE_ECB)  # 初始化AES,ECB模式的实例
        # 截断函数，去除填充的字符
        self.unpad = lambda date: date[0:-ord(date[-1])]

    def pad(self, text):
        """
        #填充函数，使被加密数据的字节码长度是block_size的整数倍
        """
        count = len(text.encode('utf-8'))
        add = self.length - (count % self.length)
        entext = text + (chr(add) * add)
        return entext

    def encrypt(self, encrData):  # 加密函数
        res = self.aes.encrypt(self.pad(encrData).encode("utf8"))
        msg = str(base64.b64encode(res), encoding="utf8")
        return msg

    def decrypt(self, decrData):  # 解密函数
        res = base64.decodebytes(decrData.encode("utf8"))
        msg = self.aes.decrypt(res).decode("utf8")
        return self.unpad(msg)


class Quantity(object):
    def __init__(self):
        self.app_id = "ak_nEfE94OSogf3x"
        app_secret = "g2BcerjK4fWmhGZoetCJHeVqJmHEmfLt3gWFjDrBLB1yxiapgUKH6kOVs2N9JH7SFuBKuOF8K/CrNNSeFO1KKtsL05z24j" \
                     "+AdWTW+V4op5QxDkmlTllvlprT8FfjctDdNDGrwHvBvE6s9h0pO0dNgopBAYiA7oosPzQhDF1A6XC1X/cZZmgBy3XRHyEv" \
                     "xTT40xzwVGish53R8dZt3YIxNtKSgrBloo/CRQsV01yU40nyQR9L9oML32VT0C16jBrxcoWthlGDwfBn+CtVUvws4imyZi" \
                     "+sG/CqZQeaVLkBCqLCgqw1VK4/a4jZws6HO3FMeBgvuf0aS5euNmfhkudQmg=="
        querystring = {"appId": f"{self.app_id}", "appSecret": f"{app_secret}"}
        url = "https://openapi.lingxing.com/api/auth-server/oauth/access-token"
        payload = ""
        headers = {
            "Content-Type": "application/x-www-form-urlencoded"
        }
        response = requests.post(url, data=payload, headers=headers, params=querystring)
        result = json.loads(response.text)
        self.access_token = result['data']['access_token']
        self.refresh_token = result['data']['refresh_token']
        self.time = datetime.now().strftime("%Y-%m-%d")
        self.time = datetime.strptime(self.time, "%Y-%m-%d")
        # self.time = "2021-12-06"
        self.start_time = self.time
        self.time = self.time.strftime("%Y-%m-%d")
        self.start_time = self.start_time.strftime("%Y-%m-%d")
        # self.start_time = '2022-03-07'

    def sql(self):
        self.connection = pymysql.connect(host='3354n8l084.goho.co',  # 数据库地址
                                          port=24824,
                                          user='test_user',  # 数据库用户名
                                          password='a123456',  # 数据库密码
                                          db='storage',  # 数据库名称
                                          charset='utf8',
                                          cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor()

    def sql_close(self):
        self.cursor.close()
        self.connection.close()

    def get_sign(self, body):
        apikey = "ak_nEfE94OSogf3x"

        # 目标md5串
        str_parm = ''
        # 将字典中的key排序
        for p in sorted (body):
            # 每次排完序的加到串中
            # if body[p]:
            # str类型需要转化为url编码格式
            if isinstance(body[p], str):
                str_parm = str_parm + str(p) + "=" + str(urllib.parse.quote (body[p])) + "&"
                continue
            str_value = str (body[p])
            str_value = str_value.replace(" ", "")
            # str_value = str_value.replace("?", " ")
            str_value = str_value.replace("'", "\"")
            str_parm = str_parm + str(p) + "=" + str_value + "&"
        # 加上对应的key
        str_parm = str_parm.rstrip('&')
        str_parm = str_parm.replace("?", " ")
        if isinstance (str_parm, str):
            # 如果是unicode先转utf-8
            parmStr = str_parm.encode ("utf-8")
            # parmStr = str_parm
            m = hashlib.md5 ()
            m.update (parmStr)
            md5_sign = m.hexdigest ()
            # print(m.hexdigest())
            md5_sign = md5_sign.upper ()
            # print("MD5加密:", md5_sign)
        eg = EncryptDate (apikey)  # 这里密钥的长度必须是16的倍数
        res = eg.encrypt (md5_sign)
        # print("AES加密:", res)
        # print(eg.decrypt(res))
        return res

    def get_list_male(self, username, parent=None):
        self.sql()
        sql1 = "select * from `amazon_form`.`user_headers` where `账号` = '%s'" % username
        self.cursor.execute(sql1)
        result1 = self.cursor.fetchall()
        list_collect = []
        if result1:
            parent_collect = result1[0]['父体收藏']
            if parent_collect:
                list_collect = parent_collect.split('&')
        if parent:
            sql2 = f"select * from `amazon_form`.`list_parent` where `本地父体` like '%{parent}%'"
        else:
            sql2 = "select * from `amazon_form`.`list_parent`"
        self.cursor.execute(sql2)
        result2 = self.cursor.fetchall()
        self.sql_close()
        if result2:
            list_msg = []
            for i in result2:
                if i['本地父体'] in list_collect:
                    continue
                else:
                    list_msg.append(i['本地父体'])
            return [list_collect, list_msg]
        else:
            return False

    def parent_collect(self, username, parent):
        self.sql()
        sql1 = "select * from `amazon_form`.`user_headers` where `账号` = '%s'" % username
        self.cursor.execute(sql1)
        result1 = self.cursor.fetchall()
        if result1:
            list_collect = []
            if result1[0]['父体收藏']:
                list_collect = result1[0]['父体收藏'].split('&')
            list_collect.append(parent)
            str_collect = '&'.join(list_collect)
            sql2 = "update `amazon_form`.`user_headers` set `父体收藏` = '%s' where `账号` = '%s'" % (str_collect, username)
        else:
            str_header = "FBA库存.FBA在途.本地库存.预计库存.30天销量"
            sql2 = "insert into `amazon_form`.`user_headers`(`账号`, `list_header`, `父体收藏`)values('%s', '%s', '%s')" % (username, str_header, parent)
        self.cursor.execute(sql2)
        self.connection.commit()
        self.sql_close()

    def parent_unbind(self, username, parent):
        self.sql()
        parent = parent[1:]
        sql1 = "select * from `amazon_form`.`user_headers` where `账号` = '%s'" % username
        self.cursor.execute(sql1)
        result1 = self.cursor.fetchall()
        list_collect = result1[0]['父体收藏'].split('&')
        list_collect.remove(parent)
        str_collect = '&'.join(list_collect)
        sql2 = "update `amazon_form`.`user_headers` set `父体收藏` = '%s' where `账号` = '%s'" % (str_collect, username)
        self.cursor.execute(sql2)
        self.connection.commit()
        self.sql_close()

    def download_asin(self):
        self.sql()
        sql = "select * from `flag`.`amazon_form_flag` where `flag_name` = 'downloads'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result[0]['flag_num'] == 0:
            try:
                self.open_downloads()
                self.s = requests.Session()
                login_url = 'https://erp.lingxing.com/api/passport/login'
                # 请求头
                headers = {'Host': 'erp.lingxing.com',
                           'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                    , 'Referer': 'https://erp.lingxing.com/login',
                           'Accept': 'application/json, text/plain, */*',
                           'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                           'Accept-Encoding': 'gzip, deflate, br',
                           'Content-Type': 'application/json;charset=utf-8',
                           'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                           'X-AK-Company-Id': '90136229927150080',
                           'X-AK-Request-Source': 'erp',
                           'X-AK-ENV-KEY': 'SAAS-10',
                           'X-AK-Version': '1.0.0.0.0.023',
                           'X-AK-Zid': '109810',
                           'Content-Length': '114',
                           'Origin': 'https://erp.lingxing.com',
                           'Connection': 'keep-alive'}
                # 传递用户名和密码
                data = {'account': 'IT-Test', 'pwd': 'IT-Test'}
                data = json.dumps(data)
                self.s.post(login_url, headers=headers, data=data)
                auth_token = self.s.cookies.get('auth-token')
                auth_token = auth_token.replace('%25', '%')
                auth_token = auth_token.replace('%23', '#')
                auth_token = auth_token.replace('%26', '&')
                auth_token = auth_token.replace('%2B', '+')
                auth_token = auth_token.replace('%28', '(')
                auth_token = auth_token.replace('%29', ')')
                auth_token = auth_token.replace('%2F', '/')
                auth_token = auth_token.replace('%3D', '=')
                auth_token = auth_token.replace('%3F', '?')
                msg1, report = self.download1(auth_token)
                if msg1:
                    msg2 = self.download2(auth_token)
                    if msg2:
                        self.write_asin(report)
                        # self.write_order()
                        self.close_downloads()
                    else:
                        self.close_downloads()
                else:
                    self.close_downloads()
            except Exception as e:
                self.close_downloads()
                print(e)
        else:
            self.close_downloads()
            return False

    def open_downloads(self):
        self.sql()
        sql = "update `flag`.`amazon_form_flag` set `flag_num` = 1 where `flag_name` = 'downloads'"
        self.cursor.execute(sql)
        self.connection.commit()
        self.sql_close()

    def close_downloads(self):
        self.sql()
        sql = "update `flag`.`amazon_form_flag` set `flag_num` = 0 where `flag_name` = 'downloads'"
        self.cursor.execute(sql)
        self.connection.commit()
        self.sql_close()

    def download1(self, auth_token):
        try:
            url = "https://gw.lingxingerp.com/listing-api/api/product/exportOnline"
            headers = {'Host': 'gw.lingxingerp.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                        , 'Referer': 'https://erp.lingxing.com',
                       'auth-token': auth_token,
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '909',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
            data = {"offset": 0, "length": 50, "search_field": "local_sku", "exact_search": 1, "sids": "", "status": "",
                    "is_pair": "", "fulfillment_channel_type": "", "global_tag_ids": "",
                    "req_time_sequence": "/listing-api/api/product/exportOnline$$2"}
            data = json.dumps(data)
            post_msg = self.s.post(url, headers=headers, data=data)
            post_msg = json.loads(post_msg.text)
            print(post_msg)
            if post_msg['code'] == 1 and post_msg['msg'] == "成功":
                report_id = post_msg['data']['data']['report_id']
                time.sleep(300)
                if report_id:
                    file_download_url = f"https://erp.lingxing.com/api/download/downloadCenterReport/downloadResource?report_id={report_id}"
                    # print(file_download_url)
                    get_headers = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/erp/muser/downloadCenter'}
                    download_file = self.s.get(file_download_url, headers=get_headers, stream=False)
                    with open('D:/listing/listing.zip', 'wb') as q:
                        q.write(download_file.content)
                    return True, report_id
                else:
                    return False, False
            else:
                return False, False
        except Exception as e:
            print(e)
            return False, False

    def download2(self, auth_token):
        try:
            url = "https://erp.lingxing.com/api/storage/export"
            headers = {'Host': 'erp.lingxing.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                        , 'Referer': 'https://erp.lingxing.com/erp/muser/downloadCenter',
                       'auth-token': auth_token,
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '909',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
            data = {"wid_list": "", "mid_list": "", "sid_list": "", "cid_list": "", "bid_list": "",
                    "principal_list": "", "product_type_list": "", "product_attribute": "", "product_status": "",
                    "search_field": "sku", "search_value": "", "is_sku_merge_show": 0, "is_hide_zero_stock": 0,
                    "offset": 0, "length": 200, "sort_field": "", "sort_type": "", "gtag_ids": "",
                    "senior_search_list": "[]", "req_time_sequence": "/api/storage/export$$3"}
            data = json.dumps(data)
            post_msg = self.s.post(url, headers=headers, data=data)
            post_msg = json.loads(post_msg.text)
            print(post_msg)
            if post_msg['code'] == 1 and post_msg['msg'] == "操作成功":
                get_url = "https://erp.lingxing.com/api/download/downloadCenterReport/getReportData?offset=0&length=100" \
                          "&report_time_type=0&req_time_sequence=/api/download/downloadCenterReport/getReportData$$"
                get_headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                               , 'Referer': 'https://erp.lingxing.com/erp/msupply/warehouseDetail',}
                time.sleep(100)
                report_id = ''
                get_msg = self.s.get(get_url, headers=get_headers)
                get_msg = json.loads(get_msg.text)
                print(get_msg['data']['list'][0])
                if get_msg['data']['list'][0]['report_name'].find('库存明细_仓库库存') >= 0:
                    report_id = get_msg['data']['list'][0]['report_id']
                if report_id:
                    file_download_url = f"https://erp.lingxing.com/api/download/downloadCenterReport/downloadResource?report_id={report_id}"
                    # print(file_download_url)
                    get_headers = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/erp/muser/downloadCenter'}
                    download_file = self.s.get(file_download_url, headers=get_headers, stream=False)
                    with open('D:/listing/库存明细_仓库库存.xlsx', 'wb') as q:
                        q.write(download_file.content)
                    return True
                else:
                    return False
            else:
                return False
        except Exception as e:
            print(e)
            return False

    def write_asin(self, report):
        file = zipfile.ZipFile('D:/listing/listing.zip')
        file.extractall('D:/listing/')
        file.close()
        time_data = datetime.now().strftime("%Y%m%d")
        filename = f"D:/listing/listing{time_data}-{report}.xlsx"
        dict_asin = {}
        dict_sku = {}
        dict_asin_son = {}
        wb = openpyxl.load_workbook(filename)
        wb_sheet = wb.active
        max_row = wb_sheet.max_row
        column1 = wb_sheet.max_column
        list_heard = []
        for i in range(1, column1 + 1):
            heard = wb_sheet.cell(row=1, column=i).value
            if heard:
                list_heard.append(heard)
        list_header_html = ['父ASIN', 'SKU', 'ASIN', '店铺', 'FNSKU', 'FBA可售', 'FBA待调仓', 'FBA调仓中', 'FBA在途', '30日销量', '7日销量' ,'14日销量', '创建时间']
        for i in list_header_html:
            if i in list_heard:
                continue
            else:
                return False
        for i in range(2, max_row+1):
            asin = wb_sheet.cell(row=i, column=(list_heard.index('父ASIN') + 1)).value
            sku = wb_sheet.cell(row=i, column=(list_heard.index('SKU') + 1)).value
            if asin and sku:
                asin_son = wb_sheet.cell(row=i, column=(list_heard.index('ASIN') + 1)).value
                shop = wb_sheet.cell(row=i, column=(list_heard.index('店铺') + 1)).value
                fnsku = wb_sheet.cell(row=i, column=(list_heard.index('FNSKU') + 1)).value
                fba_order1 = int(wb_sheet.cell(row=i, column=(list_heard.index('FBA可售') + 1)).value)
                fba_order2 = int(wb_sheet.cell(row=i, column=(list_heard.index('FBA待调仓') + 1)).value)
                fba_order3 = int(wb_sheet.cell(row=i, column=(list_heard.index('FBA调仓中') + 1)).value)
                fba_order4 = int(wb_sheet.cell(row=i, column=(list_heard.index('FBA在途') + 1)).value)
                fba_order = fba_order3 + fba_order2 + fba_order1
                sale_num = int(wb_sheet.cell(row=i, column=(list_heard.index('30日销量') + 1)).value)
                sale_num_7 = int(wb_sheet.cell(row=i, column=(list_heard.index('7日销量') + 1)).value)
                sale_num_14 = int(wb_sheet.cell(row=i, column=(list_heard.index('14日销量') + 1)).value)
                create_time = str(wb_sheet.cell(row=i, column=(list_heard.index('创建时间') + 1)).value)
                if asin in dict_asin:
                    if sku in dict_asin[asin]:
                        if shop in dict_asin[asin][sku]:
                            dict_asin[asin][sku][shop].append([fnsku, fba_order, fba_order4, sale_num, shop, asin_son, asin, create_time, sale_num_14, sale_num_7])
                    else:
                        dict_asin[asin][sku] = {shop: [[fnsku, fba_order, fba_order4, sale_num, shop, asin_son, asin, create_time, sale_num_14, sale_num_7]]}
                else:
                    dict_asin[asin] = {sku: {shop: [[fnsku, fba_order, fba_order4, sale_num, shop, asin_son, asin, create_time, sale_num_14, sale_num_7]]}}
                if sku in dict_sku:
                    if shop in dict_sku[sku]:
                        dict_sku[sku][shop].append([fnsku, fba_order, fba_order4, sale_num, shop, asin_son, asin, create_time, sale_num_14, sale_num_7])
                    else:
                        dict_sku[sku][shop] = [[fnsku, fba_order, fba_order4, sale_num, shop, asin_son, asin, create_time, sale_num_14, sale_num_7]]
                else:
                    dict_sku[sku] = {shop: [[fnsku, fba_order, fba_order4, sale_num, shop, asin_son, asin, create_time, sale_num_14, sale_num_7]]}
                if asin in dict_asin_son:
                    if asin_son not in dict_asin_son[asin]:
                        dict_asin_son[asin].append(asin_son)
                else:
                    dict_asin_son[asin] = [asin_son]
        str_json = json.dumps(dict_asin)
        sku_json = json.dumps(dict_sku)
        asin_son_json = json.dumps(dict_asin_son)
        f1 = open('./static/json/asin_info.json', 'w')
        f1.write(str_json)
        f2 = open('./static/json/sku_info.json', 'w')
        f2.write(sku_json)
        f3 = open('./static/json/asin_son_info.json', 'w')
        f3.write(asin_son_json)
        self.read_json(asin_son_json)
        # print(len(dict_sku))
        dict_order = self.write_order()
        # print(len(dict_order))
        self.parent_json(dict_sku, dict_order)
        # print(len(dict_asin))
        self.asin_json(dict_asin, dict_order)

    def write_order(self):
        filename = f"D:/listing/库存明细_仓库库存.xlsx"
        dict_order = {}
        dict_fnsku = {}
        wb = openpyxl.load_workbook(filename)
        wb_sheet = wb.active
        max_row = wb_sheet.max_row
        column1 = wb_sheet.max_column
        list_heard = []
        for i in range(1, column1 + 1):
            heard = wb_sheet.cell(row=1, column=i).value
            if heard:
                list_heard.append(heard)
        for i in range(2, max_row+1):
            sku = wb_sheet.cell(row=i, column=(list_heard.index('SKU') + 1)).value
            if sku:
                fnsku = wb_sheet.cell(row=i, column=(list_heard.index('FNSKU') + 1)).value
                num = int(wb_sheet.cell(row=i, column=(list_heard.index('实际总量') + 1)).value)
                predict_num = int(wb_sheet.cell(row=i, column=(list_heard.index('预计总量') + 1)).value)
                shop = wb_sheet.cell(row=i, column=(list_heard.index('店铺') + 1)).value
                if fnsku:
                    if sku in dict_order:
                        if shop in dict_order[sku]:
                            dict_order[sku][shop][0] += num
                            dict_order[sku][shop][1] += predict_num
                        else:
                            dict_order[sku][shop] = [num, predict_num]
                    else:
                        dict_order[sku] = {shop: [num, predict_num]}
                    if sku in dict_fnsku:
                        if fnsku in dict_fnsku[sku]:
                            flag = 1
                            for j in range(0, len(dict_fnsku[sku][fnsku])):
                                if dict_fnsku[sku][fnsku][j][2] == shop:
                                    dict_fnsku[sku][fnsku][j][0] += num
                                    dict_fnsku[sku][fnsku][j][1] += predict_num
                                    flag = 0
                            if flag:
                                dict_fnsku[sku][fnsku].append([num, predict_num, shop])
                        else:
                            dict_fnsku[sku][fnsku] = [[num, predict_num, shop]]
                    else:
                        dict_fnsku[sku] = {fnsku: [[num, predict_num, shop]]}
        fnsku_json = json.dumps(dict_fnsku)
        str_json = json.dumps(dict_order)
        f1 = open('./static/json/order_info.json', 'w')
        f1.write(str_json)
        f2 = open('./static/json/fnsku_info.json', 'w')
        f2.write(fnsku_json)
        return dict_order

    def parent_json(self, dict_sku, dict_order):
        self.sql()
        sql = "select * from `amazon_form`.`male_parent`"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        dict_male = {}
        dict_parent = {}
        for i in result:
            parent = i['本地父体']
            if parent in dict_parent:
                dict_parent[parent][i['SKU']] = ''
            else:
                dict_male[parent] = {}
                dict_parent[parent] = {}
                dict_parent[parent][i['SKU']] = ''
        self.sql()
        sql2 = "select * from `amazon_form`.`male_sku`"
        self.cursor.execute(sql2)
        result2 = self.cursor.fetchall()
        self.sql_close()
        dict_local = {}
        for i in result2:
            sku = i['SKU']
            dict_local[sku] = i['本地品名']
        for i in dict_parent:
            for j in dict_parent[i]:
                if j in dict_local:
                    dict_parent[i][j] = dict_local[j]
        dict_parent_new = {}
        for i in dict_parent:
            dict_parent_new[i] = {}
            for j in dict_parent[i]:
                if dict_parent[i][j] in dict_parent_new[i]:
                    dict_parent_new[i][dict_parent[i][j]].append(j)
                else:
                    dict_parent_new[i][dict_parent[i][j]] = [j]
        # print(len(dict_parent_new))
        # print(dict_sku['[CB]322-COMBO'])
        for i in dict_parent_new:
            for p in dict_parent_new[i]:
                for j in dict_parent_new[i][p]:
                    if j in dict_sku:
                        for k in dict_sku[j]:
                            if k in dict_male[i]:
                                if p in dict_male[i][k]:
                                    for q in dict_sku[j][k]:
                                        dict_male[i][k][p]['FBA库存'] += q[1]
                                        dict_male[i][k][p]['FBA在途'] += q[2]
                                        dict_male[i][k][p]['30天销量'] += q[3]
                                        dict_male[i][k][p]['14天销量'] += q[8]
                                        dict_male[i][k][p]['7天销量'] += q[9]
                                else:
                                    dict_male[i][k][p] = {'FBA库存': 0, 'FBA在途': 0, '7天销量': 0, '14天销量': 0, '30天销量': 0, '本地库存': 0, '预计库存': 0}
                                    for q in dict_sku[j][k]:
                                        dict_male[i][k][p]['FBA库存'] += q[1]
                                        dict_male[i][k][p]['FBA在途'] += q[2]
                                        dict_male[i][k][p]['30天销量'] += q[3]
                                        dict_male[i][k][p]['14天销量'] += q[8]
                                        dict_male[i][k][p]['7天销量'] += q[9]
                            else:
                                dict_male[i][k] = {p: {'FBA库存': 0, 'FBA在途': 0, '7天销量': 0, '14天销量': 0, '30天销量': 0, '本地库存': 0, '预计库存': 0}}
                                for q in dict_sku[j][k]:
                                    dict_male[i][k][p]['FBA库存'] += q[1]
                                    dict_male[i][k][p]['FBA在途'] += q[2]
                                    dict_male[i][k][p]['30天销量'] += q[3]
                                    dict_male[i][k][p]['14天销量'] += q[8]
                                    dict_male[i][k][p]['7天销量'] += q[9]
                            if j in dict_order and k in dict_order[j]:
                                dict_male[i][k][p]['本地库存'] += dict_order[j][k][0]
                                dict_male[i][k][p]['预计库存'] += dict_order[j][k][1]
        dict_parent_json = json.dumps(dict_male)
        f1 = open('./static/json/dict_male.json', 'w')
        f1.write(dict_parent_json)

    def asin_json(self, dict_asin, dict_order):
        dict_amazon = {}
        dict_sku = {}
        self.sql()
        sql = "select * from `amazon_form`.`male_sku`"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        for i in result:
            dict_sku[i['SKU']] = i['本地品名']
        for i in dict_asin:
            dict_amazon[i] = {}
            for j in dict_asin[i]:
                if j in dict_sku:
                    local_sku = dict_sku[j]
                else:
                    local_sku = j
                for k in dict_asin[i][j]:
                    if k in dict_amazon[i]:
                        if local_sku in dict_amazon[i][k]:
                            for q in dict_asin[i][j][k]:
                                dict_amazon[i][k][local_sku]['FBA库存'] += q[1]
                                dict_amazon[i][k][local_sku]['FBA在途'] += q[2]
                                dict_amazon[i][k][local_sku]['30天销量'] += q[3]
                                dict_amazon[i][k][local_sku]['14天销量'] += q[8]
                                dict_amazon[i][k][local_sku]['7天销量'] += q[9]
                        else:
                            dict_amazon[i][k][local_sku] = {'FBA库存': 0, 'FBA在途': 0, '7天销量': 0, '14天销量': 0, '30天销量': 0, '本地库存': 0, '预计库存': 0}
                            for q in dict_asin[i][j][k]:
                                dict_amazon[i][k][local_sku]['FBA库存'] += q[1]
                                dict_amazon[i][k][local_sku]['FBA在途'] += q[2]
                                dict_amazon[i][k][local_sku]['30天销量'] += q[3]
                                dict_amazon[i][k][local_sku]['14天销量'] += q[8]
                                dict_amazon[i][k][local_sku]['7天销量'] += q[9]
                    else:
                        dict_amazon[i][k] = {local_sku: {'FBA库存': 0, 'FBA在途': 0, '7天销量': 0, '14天销量': 0, '30天销量': 0, '本地库存': 0, '预计库存': 0}}
                        for q in dict_asin[i][j][k]:
                            dict_amazon[i][k][local_sku]['FBA库存'] += q[1]
                            dict_amazon[i][k][local_sku]['FBA在途'] += q[2]
                            dict_amazon[i][k][local_sku]['30天销量'] += q[3]
                            dict_amazon[i][k][local_sku]['14天销量'] += q[8]
                            dict_amazon[i][k][local_sku]['7天销量'] += q[9]
                    if j in dict_order and k in dict_order[j]:
                        dict_amazon[i][k][local_sku]['本地库存'] += dict_order[j][k][0]
                        dict_amazon[i][k][local_sku]['预计库存'] += dict_order[j][k][1]
        dict_amazon_json = json.dumps(dict_amazon)
        f1 = open('./static/json/dict_amazon.json', 'w')
        f1.write(dict_amazon_json)

    def get_male_msg(self, parent, shop, list_header):
        f = open('./static/json/dict_male.json', 'r')
        dict_parent = json.load(f)
        list_msg = []
        if parent in dict_parent and shop in dict_parent[parent]:
            k = 0
            for i in dict_parent[parent][shop]:
                list_msg.append([])
                list_msg[k] = [parent, i]
                for j in list_header:
                    list_msg[k].append(dict_parent[parent][shop][i][j])
                k += 1
            return list_msg
        else:
            return False

    def get_amazon_msg(self, parent, shop, list_header):
        f = open('./static/json/dict_amazon.json', 'r')
        dict_amazon = json.load(f)
        list_msg = []
        if parent in dict_amazon and shop in dict_amazon[parent]:
            k = 0
            for i in dict_amazon[parent][shop]:
                list_msg.append([])
                list_msg[k] = [parent, i]
                for j in list_header:
                    list_msg[k].append(dict_amazon[parent][shop][i][j])
                k += 1
            return list_msg
        else:
            return False

    def update_json(self):
        try:
            print("更新本地父体数据")
            f1 = open('./static/json/sku_info.json', 'r')
            dict_sku = json.load(f1)
            # f2 = open('./static/json/asin_info.json', 'r')
            # dict_asin = json.load(f2)
            f3 = open('./static/json/order_info.json', 'r')
            dict_order = json.load(f3)
            self.parent_json(dict_sku, dict_order)
            # self.asin_json(dict_asin, dict_order)
            return True, True
        except Exception as e:
            print(e)
            return False, str(e)

    def read_json(self, asin_son_json):
        time_now = datetime.now().strftime('%Y%m%d%H%M%S')
        f = open(f"./static/父ASIN详情/asin_son_info-{time_now}.json", 'w')
        f.write(asin_son_json)
        filename = "./static/父ASIN详情"
        list_file = os.listdir(filename)
        filelist = []
        for i in range(0, len(list_file)):
            path = os.path.join(filename, list_file[i])
            if os.path.isfile(path):
                filelist.append(list_file[i])
        time_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        time_now = datetime.strptime(time_now, "%Y-%m-%d %H:%M:%S")
        for i in range(0, len(filelist)):
            path = os.path.join(filename, filelist[i])
            if os.path.isdir(path):
                continue
            timestamp = os.path.getmtime(path)
            date = datetime.fromtimestamp(timestamp)
            flie_time = date.strftime('%Y-%m-%d %H:%M:%S')
            flie_time = datetime.strptime(flie_time, "%Y-%m-%d %H:%M:%S")
            file_day = (time_now - flie_time).days
            if file_day >= 30:
                os.remove(path)

    def get_list_parent(self, list_data, length):
        list_parent = []
        for i in range(1, len(list_data)):
            if list_data[i][1] != "...":
                list_col = []
                for j in range(1, length+3):
                    list_col.append(list_data[i][j])
                list_parent.append(list_col)
        return list_parent

    def get_list_amazon(self, list_data, length):
        list_amazon = []
        for i in range(1, len(list_data)):
            if list_data[i][length+3] != "...":
                list_col = []
                for j in range(length+3, length+length+5):
                    list_col.append(list_data[i][j])
                list_amazon.append(list_col)
        return list_amazon

    def string_splicing(self, list_parent, list_amazon, list_header, parent_asin, amazon_asin):
        if list_parent and list_amazon:
            if len(list_amazon) < len(list_parent):
                length = len(list_amazon)
            else:
                length = len(list_parent)
            str_header1 = string_html.str_header_html1()
            str_header2 = string_html.str_header_html2()
            str_header3 = ''
            str_header4 = ''
            for i in list_header:
                str_header_col = string_html.str_header_html3(i, parent_asin)
                str_header3 += str_header_col
            for i in list_header:
                str_header_col = string_html.str_header_html3(i, amazon_asin)
                str_header4 += str_header_col
            str_header = f'<tr class="list_control_header">{str_header1}{str_header3}{str_header2}{str_header4}</tr>'
            str_control = ''
            for i in range(0, length):
                str_control_col1 = string_html.str_content_html(i+1, "list_control4")
                for j in range(0, len(list_parent[i])):
                    if j == 0:
                        str_control_col2 = string_html.str_content_html(list_parent[i][j], "list_control2")
                    elif j == 1:
                        str_control_col2 = string_html.str_content_name(list_parent[i][j], parent_asin)
                    else:
                        str_control_col2 = string_html.str_content_html(list_parent[i][j], "list_control1")
                    str_control_col1 += str_control_col2
                for j in range(0, len(list_amazon[i])):
                    if j == 0:
                        str_control_col2 = string_html.str_content_html(list_amazon[i][j], "list_control2")
                    elif j == 1:
                        str_control_col2 = string_html.str_content_name(list_amazon[i][j], amazon_asin)
                    else:
                        str_control_col2 = string_html.str_content_html(list_amazon[i][j], "list_control1")
                    str_control_col1 += str_control_col2
                str_control_tr = f'<tr class="list_control_tr">{str_control_col1}</tr>'
                str_control += str_control_tr
            str_control2 = ''
            if len(list_amazon) != len(list_parent):
                if length == len(list_parent):
                    for i in range(length, len(list_amazon)):
                        str_control_col1 = string_html.str_content_html(i+1, "list_control4")
                        for j in range(0, len(list_amazon[i])):
                            if j == 0:
                                str_control_col2 = string_html.str_content_html("...", "list_control2")
                            elif j == 1:
                                str_control_col2 = string_html.str_content_html("...", "list_control3")
                            else:
                                str_control_col2 = string_html.str_content_html("...", "list_control1")
                            str_control_col1 += str_control_col2
                        for j in range(0, len(list_amazon[i])):
                            if j == 0:
                                str_control_col2 = string_html.str_content_html(list_amazon[i][j], "list_control2")
                            elif j == 1:
                                str_control_col2 = string_html.str_content_name(list_amazon[i][j], amazon_asin)
                            else:
                                str_control_col2 = string_html.str_content_html(list_amazon[i][j], "list_control1")
                            str_control_col1 += str_control_col2
                        str_control_tr = f'<tr class="list_control_tr">{str_control_col1}</tr>'
                        str_control2 += str_control_tr
                else:
                    for i in range(length, len(list_parent)):
                        str_control_col1 = string_html.str_content_html(i+1, "list_control4")
                        for j in range(0, len(list_parent[i])):
                            if j == 0:
                                str_control_col2 = string_html.str_content_html(list_parent[i][j], "list_control2")
                            elif j == 1:
                                str_control_col2 = string_html.str_content_name(list_parent[i][j], parent_asin)
                            else:
                                str_control_col2 = string_html.str_content_html(list_parent[i][j], "list_control1")
                            str_control_col1 += str_control_col2
                        for j in range(0, len(list_parent[i])):
                            if j == 0:
                                str_control_col2 = string_html.str_content_html("...", "list_control2")
                            elif j == 1:
                                str_control_col2 = string_html.str_content_html("...", "list_control3")
                            else:
                                str_control_col2 = string_html.str_content_html("...", "list_control1")
                            str_control_col1 += str_control_col2
                        str_control_tr = f'<tr class="list_control_tr">{str_control_col1}</tr>'
                        str_control2 += str_control_tr
            str_html = str_header + str_control + str_control2
            return str_html
        else:
            if list_amazon:
                str_header1 = string_html.str_header_html1()
                str_header2 = string_html.str_header_html2()
                str_header3 = ''
                str_header4 = ''
                for i in list_header:
                    str_header_col = string_html.str_header_html3(i, "")
                    str_header3 += str_header_col
                for i in list_header:
                    str_header_col = string_html.str_header_html3(i, amazon_asin)
                    str_header4 += str_header_col
                str_header = f'<tr class="list_control_header">{str_header1}{str_header3}{str_header2}{str_header4}</tr>'
                str_control = ''
                for i in range(0, len(list_amazon)):
                    str_control_col1 = string_html.str_content_html(i + 1, "list_control4")
                    for j in range(0, len(list_amazon[i])):
                        if j == 0:
                            str_control_col2 = string_html.str_content_html("...", "list_control2")
                        elif j == 1:
                            str_control_col2 = string_html.str_content_html("...", "list_control3")
                        else:
                            str_control_col2 = string_html.str_content_html("...", "list_control1")
                        str_control_col1 += str_control_col2
                    for j in range(0, len(list_amazon[i])):
                        if j == 0:
                            str_control_col2 = string_html.str_content_html(list_amazon[i][j], "list_control2")
                        elif j == 1:
                            str_control_col2 = string_html.str_content_name(list_amazon[i][j], parent_asin)
                        else:
                            str_control_col2 = string_html.str_content_html(list_amazon[i][j], "list_control1")
                        str_control_col1 += str_control_col2
                    str_control_tr = f'<tr class="list_control_tr">{str_control_col1}</tr>'
                    str_control += str_control_tr
                str_html = str_header + str_control
                return str_html
            else:
                str_header1 = string_html.str_header_html1()
                str_header2 = string_html.str_header_html2()
                str_header3 = ''
                str_header4 = ''
                for i in list_header:
                    str_header_col = string_html.str_header_html3(i, parent_asin)
                    str_header3 += str_header_col
                for i in list_header:
                    str_header_col = string_html.str_header_html3(i, '')
                    str_header4 += str_header_col
                str_header = f'<tr class="list_control_header">{str_header1}{str_header3}{str_header2}{str_header4}</tr>'
                str_control = ''
                for i in range(0, len(list_parent)):
                    str_control_col1 = string_html.str_content_html(i + 1, "list_control4")
                    for j in range(0, len(list_parent[i])):
                        if j == 0:
                            str_control_col2 = string_html.str_content_html(list_parent[i][j], "list_control2")
                        elif j == 1:
                            str_control_col2 = string_html.str_content_name(list_parent[i][j], amazon_asin)
                        else:
                            str_control_col2 = string_html.str_content_html(list_parent[i][j], "list_control1")
                        str_control_col1 += str_control_col2
                    for j in range(0, len(list_parent[i])):
                        if j == 0:
                            str_control_col2 = string_html.str_content_html("...", "list_control2")
                        elif j == 1:
                            str_control_col2 = string_html.str_content_html("...", "list_control3")
                        else:
                            str_control_col2 = string_html.str_content_html("...", "list_control1")
                        str_control_col1 += str_control_col2
                    str_control_tr = f'<tr class="list_control_tr">{str_control_col1}</tr>'
                    str_control += str_control_tr
                str_html = str_header + str_control
                return str_html

    def get_list_header(self, username):
        self.sql()
        sql = "select * from `amazon_form`.`user_headers` where `账号` = '%s'" % username
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            self.sql_close()
            str_header = result[0]['list_header']
            list_header = str_header.split('.')
            return list_header
        else:
            str_header = "FBA库存.FBA在途.本地库存.预计库存.30天销量"
            list_header = str_header.split('.')
            sql2 = "insert into `amazon_form`.`user_headers`(`账号`, `list_header`)values('%s', '%s')" % (username, str_header)
            self.cursor.execute(sql2)
            self.connection.commit()
            self.sql_close()
            return list_header

    def change_list_header(self, username, list_header):
        str_header = '.'.join(list_header)
        sql = "update `amazon_form`.`user_headers` set `list_header` = '%s' where `账号` = '%s'" % (str_header, username)
        try:
            self.sql()
            self.cursor.execute(sql)
            self.connection.commit()
            self.sql_close()
            return True, True
        except Exception as e:
            print(e)
            return False, str(e)

    def amazon_parent(self, male_parent):
        self.sql()
        sql = "select * from `amazon_form`.`male_amazon` where `本地父体` = '%s'" % male_parent
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            list_amazon = []
            for i in result:
                list_amazon.append(i['亚马逊父体'])
            return list_amazon
        else:
            return False

    def contrast_parent(self, list1, list2):
        # print(list2)
        dict1 = {}
        dict2 = {}
        list_male = []
        list_asin = []
        for i in list1:
            if i[1]:
                if i[1] == '本地品名':
                    continue
                if i[1] == '...':
                    continue
                dict1[i[1]] = i
        for i in list2:
            if i[1]:
                if i[1] == '本地品名':
                    continue
                if i[1] == '...':
                    continue
                dict2[i[1]] = i
        list_sku = []
        for i in dict1:
            if i in dict2:
                list_i_1 = [dict2[i][0]]
                list_i_2 = [dict1[i][0]]
                for j in range(1, len(dict1[i])):
                    list_i_1.append(dict1[i][j])
                    list_i_2.append(dict2[i][j])
                list_male.append(list_i_1)
                list_sku.append(i)
                list_asin.append(list_i_2)
        # print(list_sku)
        # print(list_male)
        list_null = []
        for i in range(1, len(list1[0])):
            list_null.append('...')
        list_male.append(list_null)
        list_asin.append(list_null)
        for i in dict1:
            if i in list_sku:
                continue
            else:
                list_i = ['无亚马逊父体']
                for j in range(1, len(dict1[i])):
                    list_i.append(dict1[i][j])
                list_male.append(list_i)
        for i in dict2:
            if i in list_sku:
                continue
            else:
                list_i = ['无本地父体']
                for j in range(1, len(dict2[i])):
                    list_i.append(dict2[i][j])
                list_male.append(list_i)
        return list_male, list_asin

    def windows_msg(self, sku, asin, country):
        if asin.find('B0') >= 0:
            list_sku = self.get_male_sku(sku)
            f1 = open('./static/json/asin_info.json', 'r')
            str_asin = json.load(f1)
            dict_sku = str_asin[asin]
            f2 = open('./static/json/fnsku_info.json', 'r')
            dict_order = json.load(f2)
            list_fnsku = []
            for i in list_sku:
                if i in dict_sku and i in dict_order:
                    if country in dict_sku[i]:
                        for j in dict_sku[i][country]:
                            if j[0] in dict_order[i]:
                                list_warehouse = []
                                for k in dict_order[i][j[0]]:
                                    if k[2] == country:
                                        list_warehouse = [i, j[0], j[5], j[6], j[1], j[2], j[3], k[0], k[1], k[2], j[7]]
                                if list_warehouse:
                                    list_fnsku.append(list_warehouse)
                                else:
                                    list_fnsku.append([i, j[0], j[5], j[6], j[1], j[2], j[3], 0, 0, j[4], j[7]])
                            else:
                                list_fnsku.append([i, j[0], j[5], j[6], j[1], j[2], j[3], 0, 0, j[4], j[7]])
                else:
                    continue
            return list_fnsku
        else:
            list_sku = self.get_male_sku(sku)
            f1 = open('./static/json/sku_info.json', 'r')
            dict_sku = json.load(f1)
            f2 = open('./static/json/fnsku_info.json', 'r')
            dict_order = json.load(f2)
            list_fnsku = []
            for i in list_sku:
                if i in dict_sku and i in dict_order:
                    if country in dict_sku[i]:
                        for j in dict_sku[i][country]:
                            if j[0] in dict_order[i]:
                                list_warehouse = []
                                for k in dict_order[i][j[0]]:
                                    if j[4] == k[2]:
                                        list_warehouse = [i, j[0], j[5], j[6], j[1], j[2], j[3], k[0], k[1], k[2], j[7]]
                                if list_warehouse:
                                    list_fnsku.append(list_warehouse)
                                else:
                                    list_fnsku.append([i, j[0], j[5], j[6], j[1], j[2], j[3], 0, 0, j[4], j[7]])
                            else:
                                list_fnsku.append([i, j[0], j[5], j[6], j[1], j[2], j[3], 0, 0, j[4], j[7]])
                else:
                    continue
            return list_fnsku

    def get_male_sku(self, sku):
        list_sku = []
        self.sql()
        sql = f"select * from `amazon_form`.`male_sku` where `本地品名` = '{sku}'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        if result:
            for i in result:
                list_sku.append(i['SKU'])
        else:
            list_sku = [sku]
        return list_sku

    def ascending_sort(self, list1, list2, index, sort_asin):
        list1_pop = []
        if list1:
            for i in list1:
                if i[1]:
                    if i[2] == '...':
                        continue
                    list_i = []
                    for j in range(0, len(i)):
                        list_i.append(i[j])
                    list1_pop.append(list_i)
        list2_pop = []
        if list2:
            for i in list2:
                if i[1]:
                    if i[2] == '...':
                        continue
                    list_i = []
                    for j in range(0, len(i)):
                        list_i.append(i[j])
                    list1_pop.append(list_i)
        list1_new = []
        list2_new = []
        if sort_asin.find('B0') >= 0:
            list2_new = sorted(list2_pop, key=lambda d: int(d[index]), reverse=False)
            if list1_pop:
                for i in list2_new:
                    flag = 1
                    for j in list1_pop:
                        if j[1] == i[1]:
                            flag = 0
                            list1_new.append(j)
                    if flag:
                        list_null = []
                        for j in range(0, len(list1[0])):
                            list_null.append('...')
                        list1_new.append(list_null)
                for i in list1_pop:
                    flag = 1
                    for j in list1_new:
                        if j[1] == i[1]:
                            flag = 0
                    if flag:
                        list1_new.append(i)
        else:
            list1_new = sorted(list1_pop, key=lambda d: int(d[index]), reverse=False)
            if list2_pop:
                for i in list1_new:
                    flag = 1
                    for j in list2_pop:
                        if j[1] == i[1]:
                            flag = 0
                            list2_new.append(j)
                    if flag:
                        list_null = []
                        for j in range(0, len(list1[0])):
                            list_null.append('...')
                        list2_new.append(list_null)
                for i in list2_pop:
                    flag = 1
                    for j in list2_new:
                        if j[1] == i[1]:
                            flag = 0
                    if flag:
                        list2_new.append(i)
        return list1_new, list2_new

    def descending_sort(self, list1, list2, index, sort_asin):
        list1_pop = []
        if list1:
            for i in list1:
                if i[1]:
                    if i[2] == '...':
                        continue
                    list_i = []
                    for j in range(0, len(i)):
                        list_i.append(i[j])
                    list1_pop.append(list_i)
        list2_pop = []
        if list2:
            for i in list2:
                if i[1]:
                    if i[2] == '...':
                        continue
                    list_i = []
                    for j in range(0, len(i)):
                        list_i.append(i[j])
                    list2_pop.append(list_i)
        list1_new = []
        list2_new = []
        if sort_asin.find('B0') >= 0:
            list2_new = sorted(list2_pop, key=lambda d: int(d[index]), reverse=True)
            if list1_pop:
                for i in list2_new:
                    flag = 1
                    for j in list1_pop:
                        if j[1] == i[1]:
                            flag = 0
                            list1_new.append(j)
                    if flag:
                        list_null = []
                        for j in range(0, len(list1[0])):
                            list_null.append('...')
                        list1_new.append(list_null)
                for i in list1_pop:
                    flag = 1
                    for j in list1_new:
                        if j[1] == i[1]:
                            flag = 0
                    if flag:
                        list1_new.append(i)
        else:
            list1_new = sorted(list1_pop, key=lambda d: int(d[index]), reverse=True)
            if list2_pop:
                for i in list1_new:
                    flag = 1
                    for j in list2_pop:
                        if j[1] == i[1]:
                            flag = 0
                            list2_new.append(j)
                    if flag:
                        list_null = []
                        for j in range(0, len(list1[0])):
                            list_null.append('...')
                        list2_new.append(list_null)
                for i in list2_pop:
                    flag = 1
                    for j in list2_new:
                        if j[1] == i[1]:
                            flag = 0
                    if flag:
                        list2_new.append(i)
        return list1_new, list2_new

    def ascending_sort2(self, list_msg, index):
        list_pop = []
        if list_msg:
            list_msg.pop(0)
            for i in list_msg:
                if i[1]:
                    if i[2] == '...':
                        continue
                    list_pop.append([i[1], i[2], i[3], i[4], i[5], i[6], i[7], i[8], i[9], i[10], i[11]])
        list_new = []
        if index == 'FBA库存':
            list_new = sorted(list_pop, key=lambda d: int(d[4]), reverse=False)
        if index == 'FBA在途':
            list_new = sorted(list_pop, key=lambda d: int(d[5]), reverse=False)
        if index == '30天销量':
            list_new = sorted(list_pop, key=lambda d: int(d[6]), reverse=False)
        if index == '本地库存':
            list_new = sorted(list_pop, key=lambda d: int(d[7]), reverse=False)
        if index == '预计库存':
            list_new = sorted(list_pop, key=lambda d: int(d[8]), reverse=False)
        return list_new

    def descending_sort2(self, list_msg, index):
        list_pop = []
        if list_msg:
            list_msg.pop(0)
            for i in list_msg:
                if i[1]:
                    if i[2] == '...':
                        continue
                    list_pop.append([i[1], i[2], i[3], i[4], i[5], i[6], i[7], i[8], i[9], i[10], i[11]])
        list_new = []
        if index == 'FBA库存':
            list_new = sorted(list_pop, key=lambda d: int(d[4]), reverse=True)
        if index == 'FBA在途':
            list_new = sorted(list_pop, key=lambda d: int(d[5]), reverse=True)
        if index == '30天销量':
            list_new = sorted(list_pop, key=lambda d: int(d[6]), reverse=True)
        if index == '本地库存':
            list_new = sorted(list_pop, key=lambda d: int(d[7]), reverse=True)
        if index == '预计库存':
            list_new = sorted(list_pop, key=lambda d: int(d[8]), reverse=True)
        return list_new

    def sql_parent_test(self):
        list_parent = []
        dict_parent = {}
        self.sql()
        dict_sku = {}
        sql = "select * from `amazon_form`.`male_sku`"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        for i in result:
            if i['本地品名'] in dict_sku:
                dict_sku[i['本地品名']].append(i['SKU'])
            else:
                dict_sku[i['本地品名']] = []
                dict_sku[i['本地品名']].append(i['SKU'])
        for i in dict_sku:
            if i in dict_parent:
                continue
            else:
                dict_parent[i] = []
            for j in dict_sku[i]:
                sql1 = "select * from `amazon_form`.`male_parent` where `SKU` = '%s'" % j
                self.cursor.execute(sql1)
                result1 = self.cursor.fetchall()
                for k in result1:
                    if k['本地父体'] in dict_parent[i]:
                        continue
                    else:
                        dict_parent[i].append(k['本地父体'])
        for i in dict_parent:
            for j in dict_parent[i]:
                sql2 = "select * from `amazon_form`.`male_parent` where `本地父体` = '%s'" % j
                self.cursor.execute(sql2)
                result2 = self.cursor.fetchall()
                list_sku = []
                for k in result2:
                    list_sku.append(k['SKU'])
                list_male = []
                for k in dict_sku[i]:
                    list_male.append(k)
                for k in list_male:
                    if k in list_sku:
                        continue
                    else:
                        list_parent.append([j, i, k])
        self.sql_close()
        print(list_parent)

    def write_sql(self, list_sku):
        self.sql()
        for i in list_sku:
            sql = "insert into `amazon_form`.`male_parent`(`本地父体`, `SKU`)values('%s', '%s')" % (i[0], i[2])
            self.cursor.execute(sql)
            self.connection.commit()
        self.sql_close()

    def download_parent(self, asin, data_time):
        if data_time:
            list_file = os.listdir("./static/父ASIN详情")
            filelist = []
            for i in range(0, len(list_file)):
                path = os.path.join("./static/父ASIN详情", list_file[i])
                if os.path.isfile(path):
                    filelist.append(list_file[i])
            filename = ''
            for i in range(0, len(filelist)):
                path = os.path.join("./static/父ASIN详情", filelist[i])
                if os.path.isdir(path):
                    continue
                if path.find('asin_son_info') >= 0:
                    timestamp = os.path.getmtime(path)
                    date = datetime.fromtimestamp(timestamp)
                    flie_time = date.strftime('%Y-%m-%d')
                    if flie_time == data_time:
                        filename = path
            if filename:
                # print(filename)
                f1 = open(filename, 'r')
                dict_asin = json.load(f1)
                if asin in dict_asin:
                    wb = openpyxl.Workbook()
                    wb_sheet = wb.active
                    wb_sheet.append(['ASIN'])
                    for i in dict_asin[asin]:
                        wb_sheet.append([i])
                    time_now = datetime.now().strftime("%Y%m%d%H%M%S")
                    path = f"./static/父ASIN详情/{asin}详情-{time_now}.xlsx"
                    wb.save(path)
                    return True, [f"../父ASIN详情/{asin}详情-{time_now}.xlsx", f"{asin}详情-{time_now}"]
                else:
                    return False, f"没有找到{asin}父ASIN的详情，请检查"
            else:
                 return False, f"{data_time}无父ASIN详情"
        else:
            f1 = open('./static/json/asin_son_info.json', 'r')
            dict_asin = json.load(f1)
            if asin in dict_asin:
                wb = openpyxl.Workbook()
                wb_sheet = wb.active
                wb_sheet.append(['ASIN'])
                for i in dict_asin[asin]:
                    wb_sheet.append([i])
                time_now = datetime.now().strftime("%Y%m%d%H%M%S")
                path = f"./static/父ASIN详情/{asin}详情-{time_now}.xlsx"
                wb.save(path)
                return True, [f"../父ASIN详情/{asin}详情-{time_now}.xlsx", f"{asin}详情-{time_now}"]
            else:
                return False, f"没有找到{asin}父ASIN的详情，请检查"

    def find_parent_sku(self, list_male, list_sku, local_sku, shop, parent_asin, amazon_asin, list_header):
        if list_sku or list_male:
            list_sku_new = []
            if list_sku:
                list_amazon = self.get_amazon_msg(amazon_asin, shop, list_header)
                for i in list_amazon:
                    if i[1] and i[1].find(local_sku) >= 0:
                        list_sku_new.append(i)
            list_male_new = []
            if list_male:
                list_male_parent = self.get_male_msg(parent_asin, shop, list_header)
                for i in list_male_parent:
                    if i[1] and i[1].find(local_sku) >= 0:
                        list_male_new.append(i)
            return list_male_new, list_sku_new
        else:
            return list_male, list_sku

    def get_json(self):
        f1 = open('./static/json/asin_info.json', 'r')
        dict_asin = json.load(f1)
        list_msg = dict_asin['B0C4GJR2GZ']['[PDD]400']
        print(list_msg)


if __name__ == '__main__':
    quantity = Quantity()
    # quantity.get_json()
    quantity.download_asin()
    # str_header = "FBA库存.FBA在途.本地库存.预计库存.30天销量"
    # list_header = str_header.split('.')
    # quantity.get_male_msg('KPW5黑色壳子款父体', 'CoBak_US', list_header)
    # quantity.update_json()
    # quantity.write_asin('525688927394889728')
    # quantity.write_order()
    # quantity.find_parent_sku([['0'], ['Fire10父体']], [['0'], ['B0BRRTKF11']], 'Fire8')
    # quantity.read_json()
    # quantity.write_asin('527448560848609280')
    # quantity.write_order()
    # quantity.find_excl('B0C3J7MCNW')
    # quantity.windows_msg('K22-壳子款-玻纤板-十字纹-黑色', 'B0C3J7MCNW')
    # quantity.sql_parent_test()

