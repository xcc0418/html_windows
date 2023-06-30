import openpyxl
import pymysql
import requests
import json
import datetime


class Quantity():
    def __init__(self):
        self.s = requests.Session()
        login_url = 'https://erp.lingxing.com/api/passport/login'
        # 请求头
        headers = {'Host': 'erp.lingxing.com',
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
            , 'Referer': 'https://erp.lingxing.com/login',
                   'Accept': 'application/json, text/plain, */*',
                   'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                   'Accept-Encoding': 'gzip, deflate, br',
                   'Content-Type': 'application/json;charset=utf-8',
                   'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                   'X-AK-Company-Id': '90136229927150080',
                   'X-AK-Request-Source': 'erp',
                   'X-AK-ENV-KEY': 'SAAS-10',
                   'X-AK-Version': '1.0.0.0.0.023',
                   'X-AK-Zid': '109810',
                   'Content-Length': '114',
                   'Origin': 'https://erp.lingxing.com',
                   'Connection': 'keep-alive'}
        # 传递用户名和密码
        data = {'account': 'IT-Test', 'pwd': 'IT-Test'}
        data = json.dumps(data)
        self.s.post(login_url, headers=headers, data=data)
        self.auth_token = None

    def sql(self):
        self.connection = pymysql.connect(host='3354n8l084.goho.co',  # 数据库地址
                                          port=24824,
                                          user='test_user',  # 数据库用户名
                                          password='a123456',  # 数据库密码
                                          db='storage',  # 数据库名称
                                          charset='utf8',
                                          cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor()

    def sql_close(self):
        self.cursor.close()
        self.connection.close()

    def get_permanent_store(self, index=None, index_data=None):
        self.sql()
        if index:
            if index == 'SKU':
                sql = f"select * from `storage`.`common_supplier` where `SKU` like '%{index_data}%'"
            else:
                sql = f"select * from `storage`.`common_supplier` where `品名` like '%{index_data}%'"
        else:
            sql = "select * from `storage`.`common_supplier`"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.sql_close()
        list_data = []
        for i in result:
            list_data.append([i['SKU'], i['品名'], i['触发库存'], i['安全库存'], i['采购基数']])
        return list_data

    def add_common_supplier(self, sku, safe_num, touch_num, purchase_num):
        try:
            self.sql()
            product_name = self.get_product_name(sku)
            if product_name:
                sql1 = f"select * from `storage`.`common_supplier` where `SKU` = '{sku}'"
                self.cursor.execute(sql1)
                result = self.cursor.fetchall()
                if result:
                    return False, f'已有{sku}数据, 新增失败！'
                sql = "insert into `storage`.`common_supplier`(`SKU`, `品名`, `触发库存`, `安全库存`, `采购基数`)" \
                      "values('%s', '%s', %d, %d, %d)" % (sku, product_name, safe_num, touch_num, purchase_num)
                self.cursor.execute(sql)
                self.connection.commit()
                self.sql_close()
                return True, True
            else:
                self.sql_close()
                return False, f'没有找到{sku}相关物料信息，请检查'
        except Exception as e:
            print(e)
            return False, str(e)

    def upload_add_common_supplier(self, filename):
        wb = openpyxl.load_workbook(filename)
        wb_sheet = wb.active
        max_row = wb_sheet.max_row
        for i in range(max_row, 0, -1):
            cell_value1 = wb_sheet.cell(row=i, column=1).value
            if cell_value1:
                max_row = i
                break
        list_error = []
        message_data = '新增成功'
        for i in range(2, max_row+1):
            sku = wb_sheet.cell(row=i, column=1).value
            safe_num = wb_sheet.cell(row=i, column=2).value
            touch_num = wb_sheet.cell(row=i, column=3).value
            purchase_num = wb_sheet.cell(row=i, column=4).value
            msg, message = self.add_common_supplier(sku, int(safe_num), int(touch_num), int(purchase_num))
            list_error.append([sku, message])
            if msg:
                continue
            else:
                message_data = '新增失败'
        ws = openpyxl.Workbook()
        ws_sheet = ws.active
        ws_sheet.append(['SKU', '状态'])
        for i in list_error:
            ws_sheet.append(i)
        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        path = f'./static/工厂常备物料管理/工厂常备物料新增详情_{time_now}.xlsx'
        file = f'../工厂常备物料管理/工厂常备物料新增详情_{time_now}.xlsx'
        filename = f'工厂常备物料新增详情_{time_now}'
        download_name = "工厂常备物料新增详情"
        ws.save(path)
        return file, filename, download_name, message_data

    def delete_common_supplier(self, delete_data):
        try:
            self.sql()
            if delete_data.find('GCWL') >= 0:
                sql1 = f"select * from `storage`.`common_supplier` where `SKU` = '{delete_data}'"
                self.cursor.execute(sql1)
                result = self.cursor.fetchall()
                if result:
                    sql2 = "DELETE FROM `storage`.`common_supplier` WHERE `SKU` = '%s'" % delete_data
                else:
                    return False, f'没有找到{delete_data}相关物料信息，请检查'
            else:
                sql1 = f"select * from `storage`.`common_supplier` where `SKU` = '{delete_data}'"
                self.cursor.execute(sql1)
                result = self.cursor.fetchall()
                if result:
                    sql2 = "DELETE FROM `storage`.`common_supplier` WHERE `品名` = '%s'" % delete_data
                else:
                    return False, f'没有找到{delete_data}相关物料信息，请检查'
            self.cursor.execute(sql2)
            self.connection.commit()
            self.sql_close()
            return True, True
        except Exception as e:
            print(e)
            return False, str(e)

    def upload_delete_common_supplier(self, filename):
        wb = openpyxl.load_workbook(filename)
        wb_sheet = wb.active
        max_row = wb_sheet.max_row
        for i in range(max_row, 0, -1):
            cell_value1 = wb_sheet.cell(row=i, column=1).value
            if cell_value1:
                max_row = i
                break
        list_error = []
        message_data = '删除成功'
        for i in range(2, max_row + 1):
            delete_data = wb_sheet.cell(row=i, column=1).value
            msg, message = self.delete_common_supplier(delete_data)
            list_error.append([delete_data, message])
            if msg:
                continue
            else:
                message_data = '删除失败'
        ws = openpyxl.Workbook()
        ws_sheet = ws.active
        ws_sheet.append(['SKU/品名', '状态'])
        for i in list_error:
            ws_sheet.append(i)
        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        path = f'./static/工厂常备物料管理/工厂常备物料删除详情_{time_now}.xlsx'
        file = f'../工厂常备物料管理/工厂常备物料删除详情_{time_now}.xlsx'
        filename = f'工厂常备物料删除详情_{time_now}'
        download_name = "工厂常备物料删除详情"
        ws.save(path)
        return file, filename, download_name, message_data

    def update_common_supplier(self, list_data):
        try:
            list_sql = self.get_permanent_store()
            dict_sql = {}
            dict_data = {}
            for i in list_sql:
                dict_sql[i[0]] = i
            list_data.pop(0)
            for i in list_data:
                dict_data[i[0]] = i
            self.sql()
            flag = 0
            for i in dict_data:
                if int(dict_data[i][2]) == dict_sql[i][2] and int(dict_data[i][3]) == dict_sql[i][3] and int(dict_data[i][4]) == dict_sql[i][4]:
                    continue
                else:
                    flag = 1
                    sql = "update `storage`.`common_supplier` set `触发库存` = %d, `安全库存` = %d, `采购基数` = %d " \
                          "where `SKU` = '%s'" % (int(dict_data[i][2]), int(dict_data[i][3]), int(dict_data[i][4]), i)
                    self.cursor.execute(sql)
            self.connection.commit()
            self.sql_close()
            if flag:
                return True, True
            else:
                return False, '当前没有要修改的物料'
        except Exception as e:
            print(e)
            return False, str(e)

    def upload_update_common_supplier(self, filename):
        wb = openpyxl.load_workbook(filename)
        wb_sheet = wb.active
        max_row = wb_sheet.max_row
        for i in range(max_row, 0, -1):
            cell_value1 = wb_sheet.cell(row=i, column=1).value
            if cell_value1:
                max_row = i
                break
        list_error = []
        list_sql = self.get_permanent_store()
        dict_sql = {}
        for i in list_sql:
            dict_sql[i[0]] = i
        self.sql()
        message_data = '修改成功'
        for i in range(2, max_row + 1):
            sku = wb_sheet.cell(row=i, column=1).value
            update_sale_num = dict_sql[sku][2]
            update_touch_num = dict_sql[sku][3]
            update_purchase_num = dict_sql[sku][4]
            safe_num = wb_sheet.cell(row=i, column=2).value
            touch_num = wb_sheet.cell(row=i, column=3).value
            purchase_num = wb_sheet.cell(row=i, column=4).value
            if safe_num:
                update_sale_num = int(safe_num)
            if touch_num:
                update_touch_num = int(touch_num)
            if purchase_num:
                update_purchase_num = int(purchase_num)
            msg, message = self.update_common_supplier_sql(sku, update_sale_num, update_touch_num, update_purchase_num)
            if msg:
                continue
            else:
                message_data = '修改失败'
            list_error.append([sku, message])
        self.connection.commit()
        self.sql_close()
        ws = openpyxl.Workbook()
        ws_sheet = ws.active
        ws_sheet.append(['SKU', '状态'])
        for i in list_error:
            ws_sheet.append(i)
        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        path = f'./static/工厂常备物料管理/工厂常备物料修改详情_{time_now}.xlsx'
        file = f'../工厂常备物料管理/工厂常备物料修改详情_{time_now}.xlsx'
        filename = f'工厂常备物料修改详情_{time_now}'
        download_name = "工厂常备物料修改详情"
        ws.save(path)
        return file, filename, download_name, message_data

    def update_common_supplier_sql(self, sku, safe_num, touch_num, purchase_num):
        try:
            sql = "update `storage`.`common_supplier` set `触发库存` = %d, `安全库存` = %d, `采购基数` = %d " \
                  "where `SKU` = '%s'" % (safe_num, touch_num, purchase_num, sku)
            self.cursor.execute(sql)
            return True, True
        except Exception as e:
            print(e)
            return False, str(e)

    def get_product_name(self, sku):
        sql = "select * from `data_read`.`product_id` where `SKU` = '%s'" % sku
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        if result:
            return result[0]['品名']
        else:
            return False